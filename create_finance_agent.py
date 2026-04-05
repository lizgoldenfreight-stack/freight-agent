"""
Finance Agent - 财务管理
记账、发票管理、应收应付、现金流跟踪
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime

header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
income_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
expense_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
input_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
highlight_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
bold_font = Font(name='Arial', bold=True, size=10)
big_bold = Font(name='Arial', bold=True, size=13, color='2F5496')
title_font = Font(name='Arial', bold=True, size=18, color='2F5496')
normal_font = Font(name='Arial', size=10)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_wrap = Alignment(horizontal='left', vertical='center', wrap_text=True)

wb = openpyxl.Workbook()

# ============================================================
# Sheet 1: 收支记录 Income & Expenses
# ============================================================
ws = wb.active
ws.title = "收支记录 Ledger"

ws.merge_cells('A1:I1')
ws.cell(row=1, column=1, value="💰 收支记录 Income & Expense Ledger").font = title_font

headers = ["日期\nDate", "类型\nType", "类别\nCategory", "客户/供应商\nParty",
           "订单号\nRef", "描述\nDescription", "收入 Income\n(USD/MYR)",
           "支出 Expense\n(USD/MYR)", "余额\nBalance"]
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=3, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = center

for r in range(4, 104):
    for c in range(1, 10):
        cell = ws.cell(row=r, column=c)
        cell.border = thin_border
        cell.alignment = center if c not in [6] else left_wrap
        cell.font = normal_font
    ws.cell(row=r, column=1).fill = input_fill
    ws.cell(row=r, column=7).number_format = '#,##0.00'
    ws.cell(row=r, column=8).number_format = '#,##0.00'
    ws.cell(row=r, column=9).number_format = '#,##0.00'

dv_type = DataValidation(type="list", formula1='"收入 Income,支出 Expense"', allow_blank=True)
ws.add_data_validation(dv_type)
dv_type.add('B4:B103')

dv_cat = DataValidation(
    type="list",
    formula1='"海运费收入,空运费收入,附加费收入,代理费收入,运费支出,报关费,文件费,拖车费,保险费,工资,租金,水电,办公用品,营销费,其他支出"',
    allow_blank=True
)
ws.add_data_validation(dv_cat)
dv_cat.add('C4:C103')

widths = [12, 10, 14, 18, 14, 25, 16, 16, 16]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = 'A4'

# ============================================================
# Sheet 2: 应收账款 Accounts Receivable
# ============================================================
ws_ar = wb.create_sheet("应收账款 AR")

ws_ar.merge_cells('A1:I1')
ws_ar.cell(row=1, column=1, value="📥 应收账款 Accounts Receivable").font = title_font

ar_headers = ["客户\nCustomer", "订单号\nRef", "Invoice No", "金额\nAmount",
              "币种\nCcy", "发票日\nInv Date", "到期日\nDue Date",
              "状态\nStatus", "备注\nNotes"]
for col, h in enumerate(ar_headers, 1):
    cell = ws_ar.cell(row=3, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = center

for r in range(4, 34):
    for c in range(1, 10):
        cell = ws_ar.cell(row=r, column=c)
        cell.border = thin_border
        cell.alignment = center if c not in [9] else left_wrap
        cell.font = normal_font
    ws_ar.cell(row=r, column=4).number_format = '#,##0.00'
    ws_ar.cell(row=r, column=4).fill = input_fill
    ws_ar.cell(row=r, column=6).fill = input_fill
    ws_ar.cell(row=r, column=7).fill = input_fill
    ws_ar.cell(row=r, column=8).fill = input_fill

dv_ar_status = DataValidation(
    type="list",
    formula1='"✅ 已收款,⏳ 未到期,⚠️ 逾期,🔴 严重逾期"',
    allow_blank=True
)
ws_ar.add_data_validation(dv_ar_status)
dv_ar_status.add('H4:H33')

dv_ccy = DataValidation(type="list", formula1='"USD,MYR"', allow_blank=True)
ws_ar.add_data_validation(dv_ccy)
dv_ccy.add('E4:E33')

# Total
ws_ar.cell(row=34, column=1, value="合计 TOTAL").font = bold_font
ws_ar.cell(row=34, column=1).border = thin_border
ws_ar.cell(row=34, column=4).border = thin_border
ws_ar.cell(row=34, column=4).fill = highlight_fill
ws_ar.cell(row=34, column=4).font = Font(name='Arial', bold=True, size=12)
ws_ar.cell(row=34, column=4).number_format = '#,##0.00'
ws_ar.cell(row=34, column=4).alignment = center

ar_widths = [18, 14, 14, 14, 8, 12, 12, 12, 20]
for i, w in enumerate(ar_widths, 1):
    ws_ar.column_dimensions[get_column_letter(i)].width = w
ws_ar.freeze_panes = 'A4'

# ============================================================
# Sheet 3: 应付账款 Accounts Payable
# ============================================================
ws_ap = wb.create_sheet("应付账款 AP")

ws_ap.merge_cells('A1:I1')
ws_ap.cell(row=1, column=1, value="📤 应付账款 Accounts Payable").font = title_font

ap_headers = ["供应商\nSupplier", "订单号\nRef", "Invoice No", "金额\nAmount",
              "币种\nCcy", "发票日\nInv Date", "到期日\nDue Date",
              "状态\nStatus", "备注\nNotes"]
for col, h in enumerate(ap_headers, 1):
    cell = ws_ap.cell(row=3, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = center

for r in range(4, 34):
    for c in range(1, 10):
        cell = ws_ap.cell(row=r, column=c)
        cell.border = thin_border
        cell.alignment = center if c not in [9] else left_wrap
        cell.font = normal_font
    ws_ap.cell(row=r, column=4).number_format = '#,##0.00'
    ws_ap.cell(row=r, column=4).fill = input_fill
    ws_ap.cell(row=r, column=6).fill = input_fill
    ws_ap.cell(row=r, column=7).fill = input_fill
    ws_ap.cell(row=r, column=8).fill = input_fill

dv_ap_status = DataValidation(
    type="list",
    formula1='"✅ 已付款,⏳ 未到期,⚠️ 待付款,🔴 已逾期"',
    allow_blank=True
)
ws_ap.add_data_validation(dv_ap_status)
dv_ap_status.add('H4:H33')

ws_ap.add_data_validation(dv_ccy)
dv_ccy2 = DataValidation(type="list", formula1='"USD,MYR"', allow_blank=True)
ws_ap.add_data_validation(dv_ccy2)
dv_ccy2.add('E4:E33')

ws_ap.cell(row=34, column=1, value="合计 TOTAL").font = bold_font
ws_ap.cell(row=34, column=1).border = thin_border
ws_ap.cell(row=34, column=4).border = thin_border
ws_ap.cell(row=34, column=4).fill = expense_fill
ws_ap.cell(row=34, column=4).font = Font(name='Arial', bold=True, size=12)
ws_ap.cell(row=34, column=4).number_format = '#,##0.00'
ws_ap.cell(row=34, column=4).alignment = center

for i, w in enumerate(ar_widths, 1):
    ws_ap.column_dimensions[get_column_letter(i)].width = w
ws_ap.freeze_panes = 'A4'

# ============================================================
# Sheet 4: 月度报表 Monthly Report
# ============================================================
ws_month = wb.create_sheet("月度报表 Monthly")

ws_month.merge_cells('A1:E1')
ws_month.cell(row=1, column=1, value="📊 月度财务报表 Monthly Financial Report").font = title_font

m_headers = ["月份\nMonth", "收入 Income\n(USD)", "支出 Expense\n(USD)",
             "净利润\nNet Profit (USD)", "利润率\nMargin %"]
for col, h in enumerate(m_headers, 1):
    cell = ws_month.cell(row=3, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = center

months = ["2026-01", "2026-02", "2026-03", "2026-04", "2026-05", "2026-06",
          "2026-07", "2026-08", "2026-09", "2026-10", "2026-11", "2026-12"]
for r, m in enumerate(months, 4):
    ws_month.cell(row=r, column=1, value=m).font = normal_font
    ws_month.cell(row=r, column=1).border = thin_border
    ws_month.cell(row=r, column=1).alignment = center
    for c in range(2, 6):
        cell = ws_month.cell(row=r, column=c)
        cell.border = thin_border
        cell.alignment = center
        cell.font = normal_font
        cell.fill = input_fill
        if c in [2, 3, 4]:
            cell.number_format = '#,##0.00'
        elif c == 5:
            cell.number_format = '0.0%'

# Year total
ws_month.cell(row=16, column=1, value="年度合计 Annual").font = bold_font
ws_month.cell(row=16, column=1).border = thin_border
for c in range(2, 5):
    cell = ws_month.cell(row=16, column=c)
    cell.font = Font(name='Arial', bold=True, size=12)
    cell.border = thin_border
    cell.fill = highlight_fill
    cell.alignment = center
    cell.number_format = '#,##0.00'

m_widths = [12, 16, 16, 18, 12]
for i, w in enumerate(m_widths, 1):
    ws_month.column_dimensions[get_column_letter(i)].width = w

# ============================================================
# Sheet 5: 财务看板 Dashboard
# ============================================================
ws_dash = wb.create_sheet("财务看板 Dashboard")

ws_dash.merge_cells('A1:F1')
ws_dash.cell(row=1, column=1, value="📊 财务看板 Financial Dashboard").font = title_font

ws_dash.cell(row=3, column=1, value="💰 现金流 Cash Flow").font = big_bold

cf_stats = [
    (4, "总收入 Total Income", '=SUMIF(收支记录 Ledger!B4:B103,"收入 Income",收支记录 Ledger!G4:G103)'),
    (5, "总支出 Total Expense", '=SUMIF(收支记录 Ledger!B4:B103,"支出 Expense",收支记录 Ledger!H4:H103)'),
    (6, "净利润 Net Profit", '=B4-B5'),
    (7, "利润率 Profit Margin", '=IF(B4=0,0,B6/B4)'),
]
for row, label, formula in cf_stats:
    ws_dash.cell(row=row, column=1, value=label).font = bold_font
    ws_dash.cell(row=row, column=1).border = thin_border
    ws_dash.cell(row=row, column=2, value=formula).font = Font(name='Arial', bold=True, size=14, color='2F5496')
    ws_dash.cell(row=row, column=2).border = thin_border
    ws_dash.cell(row=row, column=2).alignment = center
    ws_dash.cell(row=row, column=2).number_format = '#,##0.00'
ws_dash.cell(row=7, column=2).number_format = '0.0%'

ws_dash.cell(row=9, column=1, value="📥 应收账款 AR Summary").font = big_bold

ar_stats = [
    (10, "应收账款总额", '=SUM(应收账款 AR!D4:D33)'),
    (11, "未到期", '=SUMIF(应收账款 AR!H4:H33,"*未到期*",应收账款 AR!D4:D33)'),
    (12, "逾期 ⚠️", '=SUMIF(应收账款 AR!H4:H33,"*逾期*",应收账款 AR!D4:D33)'),
]
for row, label, formula in ar_stats:
    ws_dash.cell(row=row, column=1, value=label).font = bold_font
    ws_dash.cell(row=row, column=1).border = thin_border
    ws_dash.cell(row=row, column=2, value=formula).font = Font(name='Arial', bold=True, size=14)
    ws_dash.cell(row=row, column=2).border = thin_border
    ws_dash.cell(row=row, column=2).alignment = center
    ws_dash.cell(row=row, column=2).number_format = '#,##0.00'

ws_dash.cell(row=14, column=1, value="📤 应付账款 AP Summary").font = big_bold

ap_stats = [
    (15, "应付账款总额", '=SUM(应付账款 AP!D4:D33)'),
    (16, "未到期", '=SUMIF(应付账款 AP!H4:H33,"*未到期*",应付账款 AP!D4:D33)'),
    (17, "待付款 ⚠️", '=SUMIF(应付账款 AP!H4:H33,"*待付款*",应付账款 AP!D4:D33)'),
]
for row, label, formula in ap_stats:
    ws_dash.cell(row=row, column=1, value=label).font = bold_font
    ws_dash.cell(row=row, column=1).border = thin_border
    ws_dash.cell(row=row, column=2, value=formula).font = Font(name='Arial', bold=True, size=14)
    ws_dash.cell(row=row, column=2).border = thin_border
    ws_dash.cell(row=row, column=2).alignment = center
    ws_dash.cell(row=row, column=2).number_format = '#,##0.00'

ws_dash.column_dimensions['A'].width = 22
ws_dash.column_dimensions['B'].width = 16

# Save
wb.save('/root/.openclaw/workspace/freight-agent/Finance_Agent.xlsx')
print("✅ Finance_Agent.xlsx created")
