"""
Operations Agent - 运营操作核心
负责：运单全流程跟踪、文件收集、送柜安排、K1/K2、每日状态更新
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime

# ============================================================
# Styles
# ============================================================
header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
input_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
blue_fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
grey_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
orange_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
purple_fill = PatternFill(start_color='E2D0F0', end_color='E2D0F0', fill_type='solid')
bold_font = Font(name='Arial', bold=True, size=10)
big_bold = Font(name='Arial', bold=True, size=13, color='2F5496')
title_font = Font(name='Arial', bold=True, size=18, color='2F5496')
subtitle_font = Font(name='Arial', bold=True, size=12, color='2F5496')
normal_font = Font(name='Arial', size=10)
small_font = Font(name='Arial', size=9, color='666666')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_wrap = Alignment(horizontal='left', vertical='center', wrap_text=True)

wb = openpyxl.Workbook()

# ============================================================
# Sheet 1: Shipment Tracker 运单全流程跟踪
# ============================================================
ws1 = wb.active
ws1.title = "运单跟踪 Shipment"

ws1.merge_cells('A1:P1')
ws1.cell(row=1, column=1, value="🚢 运单全流程跟踪 Shipment Tracker").font = title_font
ws1.row_dimensions[1].height = 35

ws1.merge_cells('A2:P2')
ws1.cell(row=2, column=1,
    value="黄色格子填资料 → 状态自动显示颜色 → 每天更新 Status 和日期").font = small_font

headers1 = [
    "No.",                    # A - 序号
    "Ref No.\n运单号",        # B
    "Customer\n客户名",        # C
    "Contact\n联系方式",       # D
    "POL\n装货港",             # E
    "POD\n卸货港",             # F
    "Container\n柜型",         # G
    "Container No.\n柜号",     # H
    "Vessel / Voyage\n船名航次", # I
    "ETD\n预计离港",           # J
    "ETA\n预计到港",           # K
    "Current Status\n当前状态", # L
    "Last Updated\n上次更新",   # M
    "Next Action\n下一步动作",  # N
    "Customs Agent\n报关行",    # O
    "Remarks\n备注",           # P
]

for col, h in enumerate(headers1, 1):
    cell = ws1.cell(row=4, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = center

# Sample data rows (10 rows for input)
status_options = (
    "Booking Confirmed 订舱确认,"
    "Awaiting Docs 等文件,"
    "Customs Filing 报关中,"
    "Container Pickup 提柜中,"
    "Stuffing 装柜中,"
    "Gate In 已进场,"
    "Vessel Departed 已离港,"
    "In Transit 运输中,"
    "Arrived at POD 已到港,"
    "Customs Clearance 清关中,"
    "Released 已放行,"
    "Delivered 已派送,"
    "Completed 已完成,"
    "On Hold 暂停,"
    "Cancelled 已取消"
)

dv_status = DataValidation(type="list", formula1=f'"{status_options}"', allow_blank=True)
dv_status.error = "请从下拉菜单选择状态"
dv_status.errorTitle = "无效状态"
ws1.add_data_validation(dv_status)

# Container type dropdown
container_options = "20GP,40GP,40HQ,45HQ,20RF,40RF,20OT,40OT,20FR,40FR,LCL"
dv_container = DataValidation(type="list", formula1=f'"{container_options}"', allow_blank=True)
ws1.add_data_validation(dv_container)

for row in range(5, 25):
    ws1.cell(row=row, column=1, value=row - 4).font = normal_font
    ws1.cell(row=row, column=1).alignment = center
    for col in range(1, 17):
        ws1.cell(row=row, column=col).border = thin_border
        ws1.cell(row=row, column=col).font = normal_font
        if col in [2, 3, 5, 6, 8, 9, 13, 14, 15, 16]:
            ws1.cell(row=row, column=col).fill = input_fill
        if col in [4, 7, 10, 11, 12]:
            ws1.cell(row=row, column=col).fill = input_fill
    # Status dropdown
    dv_status.add(ws1.cell(row=row, column=12))
    # Container dropdown
    dv_container.add(ws1.cell(row=row, column=7))

# Column widths
col_widths_1 = [5, 16, 16, 16, 12, 12, 12, 16, 18, 13, 13, 20, 14, 20, 16, 20]
for i, w in enumerate(col_widths_1, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# Freeze top rows
ws1.freeze_panes = 'A5'

# Conditional formatting notes (we'll add color logic below)
# Color legend
ws1.cell(row=26, column=1, value="📊 状态颜色说明:").font = bold_font
status_colors = [
    ("Booking Confirmed ~ Awaiting Docs", "🟡 黄色 - 待处理"),
    ("Customs Filing ~ Gate In", "🔵 蓝色 - 进行中"),
    ("Vessel Departed ~ In Transit", "🟠 橙色 - 运输中"),
    ("Arrived ~ Delivered", "🟢 绿色 - 接近完成"),
    ("Completed", "✅ 完成"),
    ("On Hold / Cancelled", "🔴 红色 - 异常"),
]
for i, (status, color) in enumerate(status_colors):
    ws1.cell(row=27 + i, column=1, value=status).font = small_font
    ws1.cell(row=27 + i, column=3, value=color).font = small_font

# ============================================================
# Sheet 2: Document Checklist 文件收集清单
# ============================================================
ws2 = wb.create_sheet("文件清单 Documents")

ws2.merge_cells('A1:N1')
ws2.cell(row=1, column=1, value="📋 文件收集清单 Document Checklist").font = title_font
ws2.row_dimensions[1].height = 35

ws2.merge_cells('A2:N2')
ws2.cell(row=2, column=1,
    value="跟踪每个运单的文件收集状态 → 确认齐全后才安排出货").font = small_font

# Sub-section: Customer Docs
ws2.cell(row=4, column=1, value="📤 需从客户收集 Collect from Customer").font = big_bold

cust_headers = [
    "No.",
    "Ref No.\n运单号",
    "Customer\n客户名",
    "Packing List\n装箱单",
    "Commercial Invoice\n商业发票",
    "Draft BL Sent\n提单草稿已发",
    "Draft BL Confirmed\n客户确认提单",
    "Remarks\n备注",
]

for col, h in enumerate(cust_headers, 1):
    cell = ws2.cell(row=6, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = center

doc_status = "Pending 待收,Received 已收到,N/A 不适用"
dv_doc = DataValidation(type="list", formula1=f'"{doc_status}"', allow_blank=True)

bl_options = "Not Sent 未发,Sent 已发送,Confirmed 已确认,Revision 需修改"
dv_bl = DataValidation(type="list", formula1=f'"{bl_options}"', allow_blank=True)

for row in range(7, 27):
    ws2.cell(row=row, column=1, value=row - 6).font = normal_font
    ws2.cell(row=row, column=1).alignment = center
    for col in range(1, 9):
        ws2.cell(row=row, column=col).border = thin_border
        ws2.cell(row=row, column=col).font = normal_font
        if col >= 4:
            ws2.cell(row=row, column=col).fill = input_fill
    dv_doc.add(ws2.cell(row=row, column=4))
    dv_doc.add(ws2.cell(row=row, column=5))
    dv_bl.add(ws2.cell(row=row, column=6))
    dv_bl.add(ws2.cell(row=row, column=7))

# Sub-section: Customs Docs (K1, K2)
ws2.cell(row=28, column=1, value="🛃 报关文件 Customs Documents").font = big_bold

customs_headers = [
    "No.",
    "Ref No.\n运单号",
    "Customer\n客户名",
    "K1 Form\nK1表格",
    "K2 Form\nK2表格",
    "Customs Agent\n报关行",
    "Filing Status\n报关状态",
    "Remarks\n备注",
]

for col, h in enumerate(customs_headers, 1):
    cell = ws2.cell(row=30, column=col, value=h)
    cell.font = header_font
    cell.fill = PatternFill(start_color='7B2D8E', end_color='7B2D8E', fill_type='solid')
    cell.border = thin_border
    cell.alignment = center

filing_status = "Preparing 准备中,Submitted 已提交,Approved 已批准,Query 有问题,Released 已放行"
dv_filing = DataValidation(type="list", formula1=f'"{filing_status}"', allow_blank=True)

for row in range(31, 51):
    ws2.cell(row=row, column=1, value=row - 30).font = normal_font
    ws2.cell(row=row, column=1).alignment = center
    for col in range(1, 9):
        ws2.cell(row=row, column=col).border = thin_border
        ws2.cell(row=row, column=col).font = normal_font
        if col >= 4:
            ws2.cell(row=row, column=col).fill = input_fill
    dv_doc.add(ws2.cell(row=row, column=4))
    dv_doc.add(ws2.cell(row=row, column=5))
    dv_filing.add(ws2.cell(row=row, column=7))

col_widths_2 = [5, 16, 16, 16, 18, 18, 18, 20]
for i, w in enumerate(col_widths_2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

ws2.freeze_panes = 'A7'

# ============================================================
# Sheet 3: Container Delivery 送柜安排
# ============================================================
ws3 = wb.create_sheet("送柜安排 Container")

ws3.merge_cells('A1:L1')
ws3.cell(row=1, column=1, value="🚛 送柜安排 Container Delivery Schedule").font = title_font
ws3.row_dimensions[1].height = 35

ws3.merge_cells('A2:L2')
ws3.cell(row=2, column=1,
    value="安排提柜→送柜→装柜→还柜全流程 → 追踪每一柜的状态").font = small_font

headers3 = [
    "No.",
    "Ref No.\n运单号",
    "Customer\n客户名",
    "Container No.\n柜号",
    "Container Type\n柜型",
    "Pickup Date\n提柜日期",
    "Pickup Location\n提柜地点",
    "Delivery Date\n送柜日期",
    "Delivery Addr\n送柜地址",
    "Stuffing Date\n装柜日期",
    "Return Date\n还柜日期",
    "Status\n状态",
]

for col, h in enumerate(headers3, 1):
    cell = ws3.cell(row=4, column=col, value=h)
    cell.font = header_font
    cell.fill = PatternFill(start_color='548235', end_color='548235', fill_type='solid')
    cell.border = thin_border
    cell.alignment = center

container_status = (
    "Pending 待安排,"
    "Pickup Scheduled 已安排提柜,"
    "Picked Up 已提柜,"
    "Delivered 已送达客户,"
    "Stuffing 装柜中,"
    "Stuffed 已装柜,"
    "Returned 已还柜,"
    "On Hold 暂停"
)
dv_cstatus = DataValidation(type="list", formula1=f'"{container_status}"', allow_blank=True)
ws3.add_data_validation(dv_cstatus)

for row in range(5, 25):
    ws3.cell(row=row, column=1, value=row - 4).font = normal_font
    ws3.cell(row=row, column=1).alignment = center
    for col in range(1, 13):
        ws3.cell(row=row, column=col).border = thin_border
        ws3.cell(row=row, column=col).font = normal_font
        if col >= 4:
            ws3.cell(row=row, column=col).fill = input_fill
    dv_cstatus.add(ws3.cell(row=row, column=12))

col_widths_3 = [5, 16, 16, 16, 14, 14, 16, 14, 18, 14, 14, 18]
for i, w in enumerate(col_widths_3, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

ws3.freeze_panes = 'A5'

# ============================================================
# Sheet 4: Daily Update Log 每日状态更新日志
# ============================================================
ws4 = wb.create_sheet("每日更新 Daily Log")

ws4.merge_cells('A1:H1')
ws4.cell(row=1, column=1, value="📅 每日状态更新日志 Daily Status Update Log").font = title_font
ws4.row_dimensions[1].height = 35

ws4.merge_cells('A2:H2')
ws4.cell(row=2, column=1,
    value="每天更新每票运单的状态 → 复制到 WhatsApp 发给客户").font = small_font

headers4 = [
    "Date\n日期",
    "Ref No.\n运单号",
    "Customer\n客户名",
    "Status Update\n状态更新",
    "Action Taken\n已执行动作",
    "Next Step\n下一步",
    "Updated By\n更新人",
    "Customer Notified\n已通知客户",
]

for col, h in enumerate(headers4, 1):
    cell = ws4.cell(row=4, column=col, value=h)
    cell.font = header_font
    cell.fill = PatternFill(start_color='BF8F00', end_color='BF8F00', fill_type='solid')
    cell.border = thin_border
    cell.alignment = center

notify_options = "Yes 已通知,No 未通知,Pending 待通知"
dv_notify = DataValidation(type="list", formula1=f'"{notify_options}"', allow_blank=True)
ws4.add_data_validation(dv_notify)

for row in range(5, 55):
    for col in range(1, 9):
        ws4.cell(row=row, column=col).border = thin_border
        ws4.cell(row=row, column=col).font = normal_font
        ws4.cell(row=row, column=col).fill = input_fill
    dv_notify.add(ws4.cell(row=row, column=8))

col_widths_4 = [14, 16, 16, 25, 25, 20, 12, 14]
for i, w in enumerate(col_widths_4, 1):
    ws4.column_dimensions[get_column_letter(i)].width = w

ws4.freeze_panes = 'A5'

# ============================================================
# Sheet 5: Customer Comms 客户沟通模板
# ============================================================
ws5 = wb.create_sheet("客户沟通模板 Comms")

ws5.merge_cells('A1:C1')
ws5.cell(row=1, column=1, value="💬 客户沟通话术模板 Customer Communication Templates").font = title_font
ws5.row_dimensions[1].height = 35

ws5.merge_cells('A2:C2')
ws5.cell(row=2, column=1,
    value="复制话术 → 替换 [xxx] 变量 → WhatsApp/微信 发给客户").font = small_font

headers5 = ["场景 Scenario", "英文模板 EN Template", "中文模板 CN Template"]
for col, h in enumerate(headers5, 1):
    cell = ws5.cell(row=4, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = center

templates = [
    [
        "📦 请客户提供 PL & Invoice",
        "Hi [Customer],\n\n"
        "For shipment [Ref No.], could you please send me:\n"
        "1. Packing List\n"
        "2. Commercial Invoice\n\n"
        "I need these to proceed with customs declaration. "
        "Deadline: [Date]. Thanks!",

        "Hi [Customer],\n\n"
        "关于运单 [Ref No.]，请提供以下文件：\n"
        "1. 装箱单 (Packing List)\n"
        "2. 商业发票 (Commercial Invoice)\n\n"
        "需要用来安排报关，截止日期：[Date]。谢谢！"
    ],
    [
        "📄 Draft BL 确认",
        "Hi [Customer],\n\n"
        "Please find attached the Draft Bill of Lading for [Ref No.].\n\n"
        "Kindly review and confirm:\n"
        "✅ OK as is — I'll arrange the original BL\n"
        "🔄 Need changes — please highlight the corrections\n\n"
        "Please confirm by [Date]. Thanks!",

        "Hi [Customer],\n\n"
        "请查收运单 [Ref No.] 的提单草稿。\n\n"
        "请确认：\n"
        "✅ 没问题 — 我安排出正本提单\n"
        "🔄 需要修改 — 请标注要改的地方\n\n"
        "请在 [Date] 前确认，谢谢！"
    ],
    [
        "🚢 货已离港",
        "Hi [Customer],\n\n"
        "Good news! Your shipment [Ref No.] has departed:\n"
        "🚢 Vessel: [Vessel/Voyage]\n"
        "📅 ETD: [Date]\n"
        "📅 ETA: [Date]\n"
        "📦 Container: [Container No.]\n\n"
        "I'll keep you updated on the transit. Let me know if you need anything!",

        "Hi [Customer],\n\n"
        "好消息！您的货物 [Ref No.] 已经离港：\n"
        "🚢 船名航次：[Vessel/Voyage]\n"
        "📅 预计离港：[Date]\n"
        "📅 预计到港：[Date]\n"
        "📦 柜号：[Container No.]\n\n"
        "途中我会持续更新状态，有需要请随时联系！"
    ],
    [
        "✅ 货已到港",
        "Hi [Customer],\n\n"
        "Your shipment [Ref No.] has arrived at [Port]!\n"
        "📅 Arrival: [Date]\n"
        "📦 Container: [Container No.]\n\n"
        "We're now arranging customs clearance. "
        "Will update you once cleared.",

        "Hi [Customer],\n\n"
        "您的货物 [Ref No.] 已到达 [Port]！\n"
        "📅 到港日期：[Date]\n"
        "📦 柜号：[Container No.]\n\n"
        "正在安排清关，完成后会立即通知您。"
    ],
    [
        "🚚 送柜安排",
        "Hi [Customer],\n\n"
        "We're arranging container delivery for [Ref No.]:\n"
        "📦 Container: [Container No.]\n"
        "📅 Delivery Date: [Date]\n"
        "📍 Address: [Address]\n\n"
        "Please ensure the warehouse is ready for stuffing. "
        "Let me know if the timing works!",

        "Hi [Customer],\n\n"
        "关于运单 [Ref No.]，送柜安排如下：\n"
        "📦 柜号：[Container No.]\n"
        "📅 送柜日期：[Date]\n"
        "📍 地址：[Address]\n\n"
        "请确保仓库准备好装柜，时间确认OK吗？"
    ],
    [
        "📊 每日状态更新",
        "Hi [Customer],\n\n"
        "📦 Shipment Update [Ref No.]:\n"
        "Status: [Current Status]\n"
        "📍 [Location/Details]\n"
        "📅 Updated: [Today's Date]\n"
        "➡️ Next: [Next Action]\n\n"
        "Will keep you posted!",

        "Hi [Customer],\n\n"
        "📦 运单状态更新 [Ref No.]:\n"
        "状态：[Current Status]\n"
        "📍 [位置/详情]\n"
        "📅 更新时间：[Today's Date]\n"
        "➡️ 下一步：[Next Action]\n\n"
        "会持续为您跟进！"
    ],
    [
        "⚠️ 延迟通知",
        "Hi [Customer],\n\n"
        "Heads up — there's a delay on [Ref No.]:\n"
        "❌ Original ETA: [Old Date]\n"
        "✅ New ETA: [New Date]\n"
        "📝 Reason: [Reason]\n\n"
        "Sorry for the inconvenience. I'm monitoring closely "
        "and will update you if anything changes.",

        "Hi [Customer],\n\n"
        "通知您一下，运单 [Ref No.] 有延迟：\n"
        "❌ 原预计到港：[Old Date]\n"
        "✅ 新预计到港：[New Date]\n"
        "📝 原因：[Reason]\n\n"
        "非常抱歉，我会密切关注并及时更新。"
    ],
    [
        "🎉 运单完成",
        "Hi [Customer],\n\n"
        "Great news! Shipment [Ref No.] is now COMPLETE! 🎉\n"
        "📦 Delivered to: [Address]\n"
        "📅 Delivery Date: [Date]\n\n"
        "Thank you for shipping with us! If you have another "
        "shipment coming up, I'm here to help. 🚢",

        "Hi [Customer],\n\n"
        "好消息！运单 [Ref No.] 已全部完成！🎉\n"
        "📦 派送至：[Address]\n"
        "📅 派送日期：[Date]\n\n"
        "感谢您的信任！有新货要走随时找我 🚢"
    ],
]

for i, row_data in enumerate(templates):
    row = 5 + i
    for col, val in enumerate(row_data, 1):
        cell = ws5.cell(row=row, column=col, value=val)
        cell.font = normal_font
        cell.border = thin_border
        cell.alignment = left_wrap
    ws5.row_dimensions[row].height = 100

col_widths_5 = [28, 55, 55]
for i, w in enumerate(col_widths_5, 1):
    ws5.column_dimensions[get_column_letter(i)].width = w

# ============================================================
# Sheet 6: Dashboard 看板
# ============================================================
ws6 = wb.create_sheet("看板 Dashboard")

ws6.merge_cells('A1:F1')
ws6.cell(row=1, column=1, value="📊 Operations 看板 Dashboard").font = title_font
ws6.row_dimensions[1].height = 35

ws6.merge_cells('A2:F2')
ws6.cell(row=2, column=1,
    value="手动填写 → 每周更新一次 → 快速了解运营状况").font = small_font

# Section 1: Weekly Summary
ws6.cell(row=4, column=1, value="📈 本周概览 Weekly Summary").font = big_bold

summary_items = [
    (5, "总运单数 Total Shipments:", ""),
    (6, "待处理 Pending:", ""),
    (7, "运输中 In Transit:", ""),
    (8, "已完成 Completed:", ""),
    (9, "异常/暂停 On Hold:", ""),
]

for row, label, default in summary_items:
    ws6.cell(row=row, column=1, value=label).font = bold_font
    ws6.cell(row=row, column=1).border = thin_border
    ws6.cell(row=row, column=2, value=default).font = normal_font
    ws6.cell(row=row, column=2).fill = input_fill
    ws6.cell(row=row, column=2).border = thin_border

# Section 2: Pending Documents
ws6.cell(row=11, column=1, value="⚠️ 待收文件 Pending Documents").font = big_bold

ws6.cell(row=12, column=1, value="运单 Ref").font = bold_font
ws6.cell(row=12, column=1).fill = yellow_fill
ws6.cell(row=12, column=1).border = thin_border
ws6.cell(row=12, column=2, value="缺什么文件 Missing").font = bold_font
ws6.cell(row=12, column=2).fill = yellow_fill
ws6.cell(row=12, column=2).border = thin_border
ws6.cell(row=12, column=3, value="截止日 Deadline").font = bold_font
ws6.cell(row=12, column=3).fill = yellow_fill
ws6.cell(row=12, column=3).border = thin_border
ws6.cell(row=12, column=4, value="状态 Status").font = bold_font
ws6.cell(row=12, column=4).fill = yellow_fill
ws6.cell(row=12, column=4).border = thin_border

for row in range(13, 23):
    for col in range(1, 5):
        ws6.cell(row=row, column=col).border = thin_border
        ws6.cell(row=row, column=col).fill = input_fill

# Section 3: Container Schedule this week
ws6.cell(row=25, column=1, value="🚛 本周送柜 Container This Week").font = big_bold

ws6.cell(row=26, column=1, value="运单 Ref").font = bold_font
ws6.cell(row=26, column=1).fill = green_fill
ws6.cell(row=26, column=1).border = thin_border
ws6.cell(row=26, column=2, value="客户 Customer").font = bold_font
ws6.cell(row=26, column=2).fill = green_fill
ws6.cell(row=26, column=2).border = thin_border
ws6.cell(row=26, column=3, value="柜号 Container No.").font = bold_font
ws6.cell(row=26, column=3).fill = green_fill
ws6.cell(row=26, column=3).border = thin_border
ws6.cell(row=26, column=4, value="日期 Date").font = bold_font
ws6.cell(row=26, column=4).fill = green_fill
ws6.cell(row=26, column=4).border = thin_border
ws6.cell(row=26, column=5, value="动作 Action").font = bold_font
ws6.cell(row=26, column=5).fill = green_fill
ws6.cell(row=26, column=5).border = thin_border

for row in range(27, 37):
    for col in range(1, 6):
        ws6.cell(row=row, column=col).border = thin_border
        ws6.cell(row=row, column=col).fill = input_fill

col_widths_6 = [25, 25, 25, 16, 16, 16]
for i, w in enumerate(col_widths_6, 1):
    ws6.column_dimensions[get_column_letter(i)].width = w

# ============================================================
# Save
# ============================================================
output_path = "operations/Operations_Agent.xlsx"
wb.save(output_path)
print(f"✅ Operations Agent saved to {output_path}")
print(f"   Sheets: {[ws.title for ws in wb.worksheets]}")
print(f"   Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
