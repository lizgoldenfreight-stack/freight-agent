"""
Freight Quotation Generator - 一键生成报价单
填黄色格子 → 自动算总价 → 切到报价单页直接发
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

wb = openpyxl.Workbook()

# Styles
header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
input_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
result_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
highlight_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
warn_fill = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')
bold_font = Font(name='Arial', bold=True, size=10)
big_bold = Font(name='Arial', bold=True, size=13, color='2F5496')
title_font = Font(name='Arial', bold=True, size=18, color='2F5496')
subtitle_font = Font(name='Arial', size=11, color='666666')
normal_font = Font(name='Arial', size=10)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_wrap = Alignment(horizontal='left', vertical='center', wrap_text=True)

# ============================================================
# Sheet 1: 生成器 (Generator) - 用户在这里填数据
# ============================================================
ws = wb.active
ws.title = "生成器 Generator"

# Title
ws.merge_cells('A1:F1')
ws.cell(row=1, column=1, value="🚢 报价单生成器 Quotation Generator")
ws.cell(row=1, column=1).font = title_font
ws.row_dimensions[1].height = 35

ws.merge_cells('A2:F2')
ws.cell(row=2, column=1, value="填写黄色格子 → 自动生成报价单 → 切换到「报价单」Sheet 查看 & 发送")
ws.cell(row=2, column=1).font = subtitle_font

# ===== SECTION 1: 报价单信息 =====
r = 4
ws.cell(row=r, column=1, value="📋 报价单信息 Quote Info").font = big_bold

info_fields = [
    (5, "报价单号 Quote No:", "QT-2026-001", "自定义编号，如 QT-2026-001"),
    (6, "报价日期 Date:", datetime.now().strftime("%Y-%m-%d"), ""),
    (7, "有效期（天）Valid Days:", 7, "报价单有效期，默认7天"),
    (8, "业务员 Sales:", "Brandon", ""),
    (9, "公司名称 Your Company:", "[Your Company]", "你的公司名"),
    (10, "公司电话 Company Phone:", "+601x-xxxxxxx", ""),
    (11, "公司邮箱 Company Email:", "", ""),
]
for row, label, default, note in info_fields:
    ws.cell(row=row, column=1, value=label).font = bold_font
    ws.cell(row=row, column=1).border = thin_border
    ws.cell(row=row, column=2, value=default).font = normal_font
    ws.cell(row=row, column=2).fill = input_fill
    ws.cell(row=row, column=2).border = thin_border
    ws.cell(row=row, column=2).alignment = center
    if note:
        ws.cell(row=row, column=3, value=note).font = Font(name='Arial', italic=True, size=9, color='999999')

# ===== SECTION 2: 客户信息 =====
r = 13
ws.cell(row=r, column=1, value="👤 客户信息 Customer Info").font = big_bold

cust_fields = [
    (14, "公司名 Company:", ""),
    (15, "联系人 Contact:", ""),
    (16, "电话 Phone:", ""),
    (17, "微信/WhatsApp:", ""),
    (18, "邮箱 Email:", ""),
]
for row, label, default in cust_fields:
    ws.cell(row=row, column=1, value=label).font = bold_font
    ws.cell(row=row, column=1).border = thin_border
    ws.cell(row=row, column=2, value=default).font = normal_font
    ws.cell(row=row, column=2).fill = input_fill
    ws.cell(row=row, column=2).border = thin_border
    ws.cell(row=row, column=2).alignment = center

# ===== SECTION 3: 货物信息 =====
r = 20
ws.cell(row=r, column=1, value="📦 货物信息 Shipment Details").font = big_bold

ship_fields = [
    (21, "运输方式 Mode:", "Sea FCL", "Sea FCL / Sea LCL / Air"),
    (22, "起运地 POL:", "Shanghai", ""),
    (23, "目的地 POD:", "Port Klang", "Port Klang / Kuching / KK / Bintulu / Miri / Sibu"),
    (24, "柜型 Container:", "40HQ", "20GP / 40GP / 40HQ (FCL) 或 CBM/KG"),
    (25, "数量 Quantity:", 1, "柜数 / CBM / KG"),
    (26, "货品描述 Cargo:", "", "如: Electronics / Furniture"),
    (27, "货值 Cargo Value (USD):", 0, "用于计算保险"),
]
for row, label, default, note in ship_fields:
    ws.cell(row=row, column=1, value=label).font = bold_font
    ws.cell(row=row, column=1).border = thin_border
    ws.cell(row=row, column=2, value=default).font = normal_font
    ws.cell(row=row, column=2).fill = input_fill
    ws.cell(row=row, column=2).border = thin_border
    ws.cell(row=row, column=2).alignment = center
    if note:
        ws.cell(row=row, column=3, value=note).font = Font(name='Arial', italic=True, size=9, color='999999')

# ===== SECTION 4: 费用明细（核心！） =====
r = 29
ws.cell(row=r, column=1, value="💰 费用明细 Cost Breakdown").font = big_bold
ws.merge_cells('A30:F30')
ws.cell(row=30, column=1, value="在「成本 Cost」列填入你的采购成本，在「报价 Selling」列填入给客户的报价（或留空自动算）").font = Font(name='Arial', italic=True, size=10, color='E65100')

cost_headers = ["费用项目 Item", "币种 Ccy", "成本 Cost", "报价 Selling", "自动Margin%", "备注 Notes"]
for col, h in enumerate(cost_headers, 1):
    cell = ws.cell(row=31, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = center

# Cost line items
cost_items = [
    (32, "海运费 Ocean Freight", "USD", "", "", "", "从供应商拿的运价"),
    (33, "THC 码头费", "USD", 150, "", "", ""),
    (34, "DOC 文件费", "USD", 50, "", "", ""),
    (35, "BAF 燃油附加费", "USD", "", "", "", "按实际或估算"),
    (36, "ISPS 安保费", "USD", 15, "", "", ""),
    (37, "报关费 Customs (MYR)", "MYR", 200, "", "", ""),
    (38, "本地运费/中转 (MYR)", "MYR", "", "", "", "东马转运填这里"),
    (39, "报关费 East MY (MYR)", "MYR", 250, "", "", "东马需要"),
    (40, "拖车费 Trucking (MYR)", "MYR", "", "", "", "如需送到门"),
    (41, "保险 Insurance", "USD", "", "", "", "0.3% cargo value"),
    (42, "其他 Others", "", "", "", "", ""),
]

for row, item, ccy, cost, sell, margin, note in cost_items:
    ws.cell(row=row, column=1, value=item).font = normal_font
    ws.cell(row=row, column=1).border = thin_border
    ws.cell(row=row, column=2, value=ccy).font = normal_font
    ws.cell(row=row, column=2).border = thin_border
    ws.cell(row=row, column=2).alignment = center
    # Cost column (user input)
    ws.cell(row=row, column=3, value=cost).font = normal_font
    ws.cell(row=row, column=3).fill = input_fill
    ws.cell(row=row, column=3).border = thin_border
    ws.cell(row=row, column=3).alignment = center
    ws.cell(row=row, column=3).number_format = '#,##0.00'
    # Selling column (user input or auto)
    ws.cell(row=row, column=4, value=sell).font = normal_font
    ws.cell(row=row, column=4).fill = input_fill
    ws.cell(row=row, column=4).border = thin_border
    ws.cell(row=row, column=4).alignment = center
    ws.cell(row=row, column=4).number_format = '#,##0.00'
    # Auto margin formula
    ws.cell(row=row, column=5).font = normal_font
    ws.cell(row=row, column=5).border = thin_border
    ws.cell(row=row, column=5).alignment = center
    ws.cell(row=row, column=5).number_format = '0.0%'
    # Note
    ws.cell(row=row, column=6, value=note).font = Font(name='Arial', italic=True, size=9, color='666666')
    ws.cell(row=row, column=6).border = thin_border

# ===== TOTALS =====
ws.cell(row=44, column=1, value="📊 汇总 SUMMARY").font = big_bold

# USD subtotal
ws.cell(row=45, column=1, value="小计 USD Subtotal").font = bold_font
ws.cell(row=45, column=1).border = thin_border
ws.cell(row=45, column=2).border = thin_border
ws.cell(row=45, column=3).border = thin_border
ws.cell(row=45, column=3).alignment = center
ws.cell(row=45, column=3).number_format = '#,##0.00'
ws.cell(row=45, column=4).border = thin_border
ws.cell(row=45, column=4).alignment = center
ws.cell(row=45, column=4).number_format = '#,##0.00'
ws.cell(row=45, column=5, value="Cost").font = bold_font
ws.cell(row=45, column=5).border = thin_border
ws.cell(row=45, column=6, value="USD 行合计").font = Font(name='Arial', italic=True, size=9, color='999999')
ws.cell(row=45, column=6).border = thin_border

# MYR subtotal
ws.cell(row=46, column=1, value="小计 MYR Subtotal").font = bold_font
ws.cell(row=46, column=1).border = thin_border
ws.cell(row=46, column=2).border = thin_border
ws.cell(row=46, column=3).border = thin_border
ws.cell(row=46, column=3).alignment = center
ws.cell(row=46, column=3).number_format = '#,##0.00'
ws.cell(row=46, column=4).border = thin_border
ws.cell(row=46, column=4).alignment = center
ws.cell(row=46, column=4).number_format = '#,##0.00'
ws.cell(row=46, column=5, value="Cost").font = bold_font
ws.cell(row=46, column=5).border = thin_border
ws.cell(row=46, column=6, value="MYR 行合计").font = Font(name='Arial', italic=True, size=9, color='999999')
ws.cell(row=46, column=6).border = thin_border

# Exchange rate
ws.cell(row=47, column=1, value="汇率 Exchange Rate").font = bold_font
ws.cell(row=47, column=1).border = thin_border
ws.cell(row=47, column=2, value="MYR→USD").font = normal_font
ws.cell(row=47, column=2).border = thin_border
ws.cell(row=47, column=2).alignment = center
ws.cell(row=47, column=3, value=4.5).font = normal_font
ws.cell(row=47, column=3).fill = input_fill
ws.cell(row=47, column=3).border = thin_border
ws.cell(row=47, column=3).alignment = center
ws.cell(row=47, column=6, value="填当前汇率").font = Font(name='Arial', italic=True, size=9, color='999999')
ws.cell(row=47, column=6).border = thin_border

# Margin input
ws.cell(row=48, column=1, value="整体 Margin %").font = bold_font
ws.cell(row=48, column=1).border = thin_border
ws.cell(row=48, column=2).border = thin_border
ws.cell(row=48, column=3, value=20).font = Font(name='Arial', bold=True, size=14)
ws.cell(row=48, column=3).fill = input_fill
ws.cell(row=48, column=3).border = thin_border
ws.cell(row=48, column=3).alignment = center
ws.cell(row=48, column=3).number_format = '0"%"'
ws.cell(row=48, column=6, value="默认20%，可调整 10%-30%").font = Font(name='Arial', italic=True, size=9, color='999999')
ws.cell(row=48, column=6).border = thin_border

# Grand total - highlight
ws.cell(row=50, column=1, value="💵 总成本 Total Cost (USD)").font = Font(name='Arial', bold=True, size=12)
ws.cell(row=50, column=1).fill = highlight_fill
ws.cell(row=50, column=1).border = thin_border
ws.cell(row=50, column=2).border = thin_border
ws.cell(row=50, column=3).fill = highlight_fill
ws.cell(row=50, column=3).border = thin_border
ws.cell(row=50, column=3).alignment = center
ws.cell(row=50, column=3).number_format = '#,##0.00'
ws.cell(row=50, column=3).font = Font(name='Arial', bold=True, size=14)

ws.cell(row=51, column=1, value="💵 报价 Selling Price (USD)").font = Font(name='Arial', bold=True, size=14, color='E65100')
ws.cell(row=51, column=1).fill = highlight_fill
ws.cell(row=51, column=1).border = thin_border
ws.cell(row=51, column=2).border = thin_border
ws.cell(row=51, column=3).fill = highlight_fill
ws.cell(row=51, column=3).border = thin_border
ws.cell(row=51, column=3).alignment = center
ws.cell(row=51, column=3).number_format = '#,##0.00'
ws.cell(row=51, column=3).font = Font(name='Arial', bold=True, size=16, color='E65100')

ws.cell(row=52, column=1, value="📈 实际 Margin").font = bold_font
ws.cell(row=52, column=1).border = thin_border
ws.cell(row=52, column=2).border = thin_border
ws.cell(row=52, column=3).border = thin_border
ws.cell(row=52, column=3).alignment = center
ws.cell(row=52, column=3).number_format = '0.0%'
ws.cell(row=52, column=3).font = Font(name='Arial', bold=True, size=12, color='2F5496')

# Column widths
ws.column_dimensions['A'].width = 32
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 16
ws.column_dimensions['D'].width = 16
ws.column_dimensions['E'].width = 14
ws.column_dimensions['F'].width = 30

# ============================================================
# Sheet 2: 报价单 (Quotation - 美观的报价单，给客户看)
# ============================================================
ws_q = wb.create_sheet("报价单 Quotation")

# Company Header
ws_q.merge_cells('A1:E1')
ws_q.cell(row=1, column=1, value="=生成器 Generator!B9")  # Company name
ws_q.cell(row=1, column=1).font = Font(name='Arial', bold=True, size=20, color='2F5496')
ws_q.row_dimensions[1].height = 40

ws_q.merge_cells('A2:E2')
ws_q.cell(row=2, column=1, value="Freight Forwarding & Logistics Services")
ws_q.cell(row=2, column=1).font = Font(name='Arial', size=12, color='666666')
ws_q.cell(row=2, column=1).alignment = Alignment(horizontal='left')

ws_q.merge_cells('A3:E3')
ws_q.cell(row=3, column=1, value='=CONCATENATE("Tel: ",\'生成器 Generator\'!B10,"  |  Email: ",\'生成器 Generator\'!B11)')
ws_q.cell(row=3, column=1).font = Font(name='Arial', size=10, color='888888')

# Divider
ws_q.merge_cells('A4:E4')
ws_q.cell(row=4, column=1, value="━" * 60)
ws_q.cell(row=4, column=1).font = Font(color='2F5496')

# Quote header
ws_q.cell(row=6, column=1, value="QUOTATION").font = Font(name='Arial', bold=True, size=18, color='2F5496')

# Quote info (right side)
ws_q.cell(row=6, column=4, value="Quote No:").font = bold_font
ws_q.cell(row=6, column=4).alignment = Alignment(horizontal='right')
ws_q.cell(row=6, column=5, value="=生成器 Generator!B5").font = bold_font

ws_q.cell(row=7, column=4, value="Date:").font = normal_font
ws_q.cell(row=7, column=4).alignment = Alignment(horizontal='right')
ws_q.cell(row=7, column=5, value="=生成器 Generator!B6").font = normal_font

ws_q.cell(row=8, column=4, value="Valid Until:").font = normal_font
ws_q.cell(row=8, column=4).alignment = Alignment(horizontal='right')
ws_q.cell(row=8, column=5, value='=生成器 Generator!B6 + 生成器 Generator!B7').font = normal_font

# Customer info
ws_q.cell(row=10, column=1, value="TO:").font = Font(name='Arial', bold=True, size=11, color='2F5496')
ws_q.cell(row=11, column=1, value="=生成器 Generator!B14").font = Font(name='Arial', bold=True, size=11)  # Company
ws_q.cell(row=12, column=1, value='=CONCATENATE("Attn: ",\'生成器 Generator\'!B15)').font = normal_font  # Contact
ws_q.cell(row=13, column=1, value='=CONCATENATE("Tel: ",\'生成器 Generator\'!B16)').font = normal_font  # Phone
ws_q.cell(row=14, column=1, value='=CONCATENATE("Email: ",\'生成器 Generator\'!B18)').font = normal_font  # Email

# Shipment summary
ws_q.cell(row=16, column=1, value="SHIPMENT DETAILS").font = Font(name='Arial', bold=True, size=12, color='2F5496')

detail_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

details = [
    (17, "Service:", '=生成器 Generator!B21'),
    (18, "Route:", '=CONCATENATE(生成器 Generator!B22," → ",生成器 Generator!B23)'),
    (19, "Container:", '=CONCATENATE(生成器 Generator!B25," x ",生成器 Generator!B24)'),
    (20, "Cargo:", '=生成器 Generator!B26'),
]
for row, label, formula in details:
    ws_q.cell(row=row, column=1, value=label).font = bold_font
    ws_q.cell(row=row, column=2, value=formula).font = normal_font
    ws_q.cell(row=row, column=2).fill = detail_fill

# Cost breakdown table
ws_q.cell(row=22, column=1, value="COST BREAKDOWN").font = Font(name='Arial', bold=True, size=12, color='2F5496')

q_headers = ["Item", "Currency", "Amount", "Remarks"]
for col, h in enumerate(q_headers, 1):
    cell = ws_q.cell(row=23, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = center

# Map generator rows to quote rows (only non-empty)
q_items = [
    (24, "Ocean Freight", '=IF(生成器 Generator!D32="","",生成器 Generator!D32)', "USD", ""),
    (25, "THC", '=IF(生成器 Generator!D33="","",生成器 Generator!D33)', "USD", ""),
    (26, "Documentation Fee", '=IF(生成器 Generator!D34="","",生成器 Generator!D34)', "USD", ""),
    (27, "BAF / Fuel Surcharge", '=IF(生成器 Generator!D35="","",生成器 Generator!D35)', "USD", ""),
    (28, "ISPS", '=IF(生成器 Generator!D36="","",生成器 Generator!D36)', "USD", ""),
    (29, "Customs Clearance", '=IF(生成器 Generator!D37="","",生成器 Generator!D37)', "MYR", "MY"),
    (30, "Local Freight / Transshipment", '=IF(生成器 Generator!D38="","",生成器 Generator!D38)', "MYR", "MY"),
    (31, "Customs (East MY)", '=IF(生成器 Generator!D39="","",生成器 Generator!D39)', "MYR", "East MY"),
    (32, "Trucking / Delivery", '=IF(生成器 Generator!D40="","",生成器 Generator!D40)', "MYR", ""),
    (33, "Insurance", '=IF(生成器 Generator!D41="","",生成器 Generator!D41)', "USD", "Optional"),
    (34, "Others", '=IF(生成器 Generator!D42="","",生成器 Generator!D42)', "", ""),
]

for row, item, formula, ccy, remark in q_items:
    ws_q.cell(row=row, column=1, value=item).font = normal_font
    ws_q.cell(row=row, column=1).border = thin_border
    ws_q.cell(row=row, column=2, value=formula).font = normal_font
    ws_q.cell(row=row, column=2).border = thin_border
    ws_q.cell(row=row, column=2).alignment = center
    ws_q.cell(row=row, column=2).number_format = '#,##0.00'
    ws_q.cell(row=row, column=3, value=ccy).font = normal_font
    ws_q.cell(row=row, column=3).border = thin_border
    ws_q.cell(row=row, column=3).alignment = center
    ws_q.cell(row=row, column=4, value=remark).font = Font(name='Arial', italic=True, size=9, color='999999')
    ws_q.cell(row=row, column=4).border = thin_border

# Total
ws_q.cell(row=36, column=1, value="TOTAL (USD)").font = Font(name='Arial', bold=True, size=13)
ws_q.cell(row=36, column=1).fill = highlight_fill
ws_q.cell(row=36, column=1).border = thin_border
ws_q.cell(row=36, column=2, value="=生成器 Generator!C51").font = Font(name='Arial', bold=True, size=16, color='E65100')
ws_q.cell(row=36, column=2).fill = highlight_fill
ws_q.cell(row=36, column=2).border = thin_border
ws_q.cell(row=36, column=2).alignment = center
ws_q.cell(row=36, column=2).number_format = '#,##0.00'
ws_q.cell(row=36, column=3, value="USD").font = bold_font
ws_q.cell(row=36, column=3).fill = highlight_fill
ws_q.cell(row=36, column=3).border = thin_border
ws_q.cell(row=36, column=3).alignment = center

# Terms
ws_q.cell(row=38, column=1, value="TERMS & CONDITIONS").font = Font(name='Arial', bold=True, size=12, color='2F5496')

terms = [
    "1. Quotation valid for 7 days from the date of issue",
    "2. Rates subject to change without prior notice",
    "3. Customs inspection fees, storage charges not included",
    "4. Insurance is optional but strongly recommended",
    "5. Payment terms: Prepaid / Collect (to be confirmed)",
    "6. Transit time is estimated and not guaranteed",
]
for i, t in enumerate(terms):
    ws_q.cell(row=39 + i, column=1, value=t).font = normal_font

# Contact
ws_q.merge_cells('A47:E47')
ws_q.cell(row=47, column=1, value="For any inquiries, please don't hesitate to contact us!")
ws_q.cell(row=47, column=1).font = Font(name='Arial', italic=True, size=10, color='666666')
ws_q.cell(row=47, column=1).alignment = Alignment(horizontal='center')

ws_q.merge_cells('A48:E48')
ws_q.cell(row=48, column=1, value='=CONCATENATE(生成器 Generator!B8,"  |  ",生成器 Generator!B10,"  |  ",生成器 Generator!B11)')
ws_q.cell(row=48, column=1).font = Font(name='Arial', bold=True, size=10, color='2F5496')
ws_q.cell(row=48, column=1).alignment = Alignment(horizontal='center')

# Column widths
ws_q.column_dimensions['A'].width = 30
ws_q.column_dimensions['B'].width = 18
ws_q.column_dimensions['C'].width = 12
ws_q.column_dimensions['D'].width = 18
ws_q.column_dimensions['E'].width = 15

# Save
wb.save('/root/.openclaw/workspace/freight-agent/Quote_Generator.xlsx')
print("✅ Quote_Generator.xlsx created")
