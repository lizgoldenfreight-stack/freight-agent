"""
Marketing AI Agent Team — 营销 AI 团队
目标：马来西亚中小企业家 | 语言：双语 | 渠道：LinkedIn/Facebook/小红书/有机流量
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime

# ============================================================
# Styles
# ============================================================
header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
input_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
blue_fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
grey_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
orange_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
purple_fill = PatternFill(start_color='E2D0F0', end_color='E2D0F0', fill_type='solid')
bold_font = Font(name='Arial', bold=True, size=10)
big_bold = Font(name='Arial', bold=True, size=13, color='2F5496')
title_font = Font(name='Arial', bold=True, size=18, color='2F5496')
small_font = Font(name='Arial', size=9, color='666666')
normal_font = Font(name='Arial', size=10)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_wrap = Alignment(horizontal='left', vertical='center', wrap_text=True)


def style_header(ws, row, cols, fill_color='2F5496'):
    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = fill
        cell.border = thin_border
        cell.alignment = center


def add_data_rows(ws, start_row, count, cols):
    for row in range(start_row, start_row + count):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.font = normal_font
            cell.fill = input_fill
            cell.alignment = left_wrap


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ============================================================
# Agent 1: Content Marketing Agent
# ============================================================
def create_content_agent(wb):
    # --- Sheet 1: Keyword Research ---
    ws = wb.create_sheet("1.1 关键词研究 Keywords")
    ws.merge_cells('A1:G1')
    ws.cell(row=1, column=1, value="🔎 Content Agent — 关键词研究 Keyword Research").font = title_font
    ws.row_dimensions[1].height = 35
    ws.merge_cells('A2:G2')
    ws.cell(row=2, column=1, value="目标：马来西亚中小企业 → 双语关键词 → 有机搜索流量").font = small_font

    headers = ["Keyword 关键词", "Language 语言", "Intent 搜索意图", "Vol 月搜索量", "Difficulty 难度", "Content Type 内容类型", "Status 状态"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=h)
    style_header(ws, 4, 7)

    keywords = [
        # English Keywords
        ["freight forwarding Malaysia", "EN", "Commercial 商业", "300+", "Medium", "Landing Page", ""],
        ["shipping from China to Malaysia", "EN", "Informational 信息", "500+", "Low", "Blog Guide", ""],
        ["sea freight China Malaysia cost", "EN", "Commercial 商业", "200+", "Low", "Calculator", ""],
        ["air freight Malaysia", "EN", "Commercial 商业", "150+", "Medium", "Landing Page", ""],
        ["customs clearance Malaysia", "EN", "Commercial 商业", "200+", "Medium", "Service Page", ""],
        ["how long shipping China to Malaysia", "EN", "Informational 信息", "400+", "Low", "FAQ Page", ""],
        ["cheapest way to ship to Malaysia", "EN", "Transactional 交易", "300+", "Medium", "Comparison", ""],
        ["LCL vs FCL shipping Malaysia", "EN", "Informational 信息", "150+", "Low", "Blog Post", ""],
        ["import from China to Malaysia guide", "EN", "Informational 信息", "600+", "Medium", "Ultimate Guide", ""],
        ["3PL logistics Malaysia", "EN", "Commercial 商业", "200+", "Medium", "Service Page", ""],
        ["freight forwarder KL", "EN", "Local 本地", "100+", "Low", "GMB + Landing", ""],
        ["logistics company Penang", "EN", "Local 本地", "80+", "Low", "GMB + Landing", ""],
        # Chinese Keywords
        ["中国到马来西亚海运", "CN", "Commercial 商业", "200+", "Medium", "Landing Page 着陆页", ""],
        ["马来西亚货运代理", "CN", "Commercial 商业", "150+", "Medium", "Service Page 服务页", ""],
        ["中国到马来西亚运费多少钱", "CN", "Transactional 交易", "300+", "Low", "Calculator 计算器", ""],
        ["马来西亚清关流程", "CN", "Informational 信息", "200+", "Low", "Guide 指南", ""],
        ["海运到马来西亚要多久", "CN", "Informational 信息", "250+", "Low", "FAQ", ""],
        ["马来西亚进口中国货物", "CN", "Informational 信息", "400+", "Medium", "Blog 博客", ""],
        ["马来西亚仓储物流", "CN", "Commercial 商业", "100+", "Low", "Service Page", ""],
        ["从中国采购到马来西亚", "CN", "Informational 信息", "300+", "Medium", "Guide 指南", ""],
        ["马来西亚电商物流", "CN", "Commercial 商业", "150+", "Medium", "Landing Page", ""],
        ["吉隆坡货运公司", "CN", "Local 本地", "80+", "Low", "GMB + Landing", ""],
    ]
    for i, row_data in enumerate(keywords):
        row = 5 + i
        for col, val in enumerate(row_data, 1):
            ws.cell(row=row, column=col, value=val)
    add_data_rows(ws, 5, len(keywords), 7)
    set_widths(ws, [35, 10, 18, 14, 14, 20, 10])
    ws.freeze_panes = 'A5'

    # --- Sheet 2: Blog Content Plan ---
    ws2 = wb.create_sheet("1.2 博客计划 Blog Plan")
    ws2.merge_cells('A1:H1')
    ws2.cell(row=1, column=1, value="📝 Content Agent — 博客内容计划 Blog Content Plan").font = title_font
    ws2.row_dimensions[1].height = 35
    ws2.merge_cells('A2:H2')
    ws2.cell(row=2, column=1, value="每周 1-2 篇 → SEO 优化 → 双语 → 引流到 WhatsApp/表单").font = small_font

    headers2 = ["Week 周", "Title (EN)", "Title (CN)", "Target Keyword", "Language 语言", "Word Count 字数", "Status 状态", "URL"]
    for col, h in enumerate(headers2, 1):
        ws2.cell(row=4, column=col, value=h)
    style_header(ws2, 4, 8)

    blogs = [
        ["Week 1", "Complete Guide: Import from China to Malaysia",
         "终极指南：从中国进口到马来西亚", "import China Malaysia", "EN+CN", "2000+", "", ""],
        ["Week 1", "How Much Does Shipping from China Cost?",
         "中国到马来西亚运费多少钱？", "shipping cost China MY", "EN+CN", "1500+", "", ""],
        ["Week 2", "FCL vs LCL: Which is Right for You?",
         "整柜 vs 拼柜：怎么选？", "FCL vs LCL Malaysia", "EN+CN", "1200+", "", ""],
        ["Week 2", "Customs Clearance in Malaysia: Step-by-Step",
         "马来西亚清关流程详解", "customs clearance MY", "EN+CN", "1800+", "", ""],
        ["Week 3", "Sea Freight vs Air Freight: Cost Comparison",
         "海运 vs 空运：费用对比", "sea vs air freight MY", "EN+CN", "1500+", "", ""],
        ["Week 3", "5 Common Mistakes When Shipping to Malaysia",
         "发马来西亚货常犯的 5 个错误", "shipping mistakes MY", "EN+CN", "1200+", "", ""],
        ["Week 4", "How to Track Your Shipment in Real Time",
         "如何实时追踪你的货物", "shipment tracking MY", "EN+CN", "1000+", "", ""],
        ["Week 4", "Case Study: How We Saved 30% on Shipping",
         "案例：帮客户省了30%运费", "shipping savings MY", "EN+CN", "1500+", "", ""],
    ]
    for i, row_data in enumerate(blogs):
        row = 5 + i
        for col, val in enumerate(row_data, 1):
            ws2.cell(row=row, column=col, value=val)
    add_data_rows(ws2, 5, len(blogs), 8)
    set_widths(ws2, [10, 40, 30, 25, 12, 12, 10, 30])
    ws2.freeze_panes = 'A5'

    # --- Sheet 3: AEO FAQ ---
    ws3 = wb.create_sheet("1.3 AEO 问答优化")
    ws3.merge_cells('A1:E1')
    ws3.cell(row=1, column=1, value="❓ Content Agent — AEO 答案引擎优化").font = title_font
    ws3.row_dimensions[1].height = 35
    ws3.merge_cells('A2:E2')
    ws3.cell(row=2, column=1, value="抢 Google 精选摘要 + ChatGPT/Perplexity 回答").font = small_font

    headers3 = ["Question (EN) 问题", "Question (CN)", "Answer 简短回答", "Target Page 目标页面", "Status 状态"]
    for col, h in enumerate(headers3, 1):
        ws3.cell(row=4, column=col, value=h)
    style_header(ws3, 4, 5, '7B2D8E')

    faqs = [
        ["How much does it cost to ship from China to Malaysia?",
         "从中国运到马来西亚多少钱？",
         "Sea freight from China to Malaysia starts from $XX/CBM for LCL and $XX/container for FCL. Air freight from $X/KG. Total cost depends on volume, weight, origin port, and delivery address.",
         "Calculator Page", ""],
        ["How long does shipping take from China to Malaysia?",
         "海运从中国到马来西亚要多久？",
         "Sea freight: 7-12 days (China to Port Klang). Air freight: 1-3 days. Plus 2-3 days for customs clearance and delivery.",
         "FAQ Section", ""],
        ["What documents do I need to import to Malaysia?",
         "进口到马来西亚需要什么文件？",
         "Commercial Invoice, Packing List, Bill of Lading/AWB, Certificate of Origin, Import License (if applicable). A freight forwarder can help prepare these.",
         "Guide Page", ""],
        ["What is the cheapest way to ship to Malaysia?",
         "最便宜的运输方式是什么？",
         "LCL sea freight is cheapest for small shipments. FCL is cheaper per unit for large volumes. Compare rates and consider consolidation options.",
         "Comparison Blog", ""],
        ["Do I need a freight forwarder to import from China?",
         "需要货代帮忙进口吗？",
         "Yes, a freight forwarder handles booking, documentation, customs clearance, and delivery — saving you time and avoiding costly mistakes.",
         "Service Page", ""],
        ["How to clear customs in Malaysia for imports?",
         "如何在马来西亚清关？",
         "Submit documents to customs, pay duties/taxes (0-30% depending on HS code), clear inspection, release goods. A licensed customs agent handles this process.",
         "Guide Page", ""],
        ["How to find a reliable shipping company in Malaysia?",
         "怎么找靠谱的马来西亚货代？",
         "Check customs license, MATRADE registration, Google reviews, years in business, and request references. Get multiple quotes and compare.",
         "Blog Post", ""],
        ["What is the difference between FCL and LCL?",
         "整柜和拼柜有什么区别？",
         "FCL (Full Container Load): You book the entire container. Best for large shipments. LCL (Less than Container Load): You share container space. Best for small shipments.",
         "Blog Post", ""],
    ]
    for i, row_data in enumerate(faqs):
        row = 5 + i
        for col, val in enumerate(row_data, 1):
            ws3.cell(row=row, column=col, value=val)
        ws3.row_dimensions[5 + i].height = 60
    add_data_rows(ws3, 5, len(faqs), 5)
    set_widths(ws3, [40, 30, 60, 18, 10])
    ws3.freeze_panes = 'A5'


# ============================================================
# Agent 2: Social Media Agent
# ============================================================
def create_social_agent(wb):
    # --- Sheet 1: Platform Strategy ---
    ws = wb.create_sheet("2.1 平台策略 Platform")
    ws.merge_cells('A1:F1')
    ws.cell(row=1, column=1, value="📱 Social Media Agent — 平台策略").font = title_font
    ws.row_dimensions[1].height = 35

    headers = ["Platform 平台", "Audience 目标受众", "Content Type 内容类型", "Post Frequency 频率", "Goal 目标", "Status 状态"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    style_header(ws, 3, 6)

    platforms = [
        ["LinkedIn", "MY SME owners, import/export managers, logistics professionals",
         "Industry insights, case studies, tips, company updates",
         "3-5 posts/week", "Brand authority + Lead generation", ""],
        ["Facebook", "MY SME owners, small traders, e-commerce sellers",
         "Tips, promotions, customer stories, live Q&A",
         "3-5 posts/week", "Community + Customer engagement", ""],
        ["小红书 (RED)", "MY Chinese-speaking entrepreneurs, China importers",
         "Import tutorials, cost breakdowns, how-to guides, visual content",
         "3-5 posts/week", "Chinese market awareness + Leads", ""],
        ["Instagram", "Younger business owners, visual brand building",
         "Behind-the-scenes, infographics, reels",
         "2-3 posts/week", "Brand awareness (secondary)", ""],
    ]
    for i, row_data in enumerate(platforms):
        row = 4 + i
        for col, val in enumerate(row_data, 1):
            ws.cell(row=row, column=col, value=val)
    add_data_rows(ws, 4, len(platforms), 6)
    set_widths(ws, [16, 35, 35, 18, 28, 10])

    # --- Sheet 2: Post Templates ---
    ws2 = wb.create_sheet("2.2 内容模板 Templates")
    ws2.merge_cells('A1:E1')
    ws2.cell(row=1, column=1, value="📋 Social Media Agent — 内容模板库").font = title_font
    ws2.row_dimensions[1].height = 35
    ws2.merge_cells('A2:E2')
    ws2.cell(row=2, column=1, value="复制模板 → 替换 [变量] → 发布").font = small_font

    headers2 = ["Platform 平台", "Type 类型", "Post (EN)", "Post (CN)", "Hashtags"]
    for col, h in enumerate(headers2, 1):
        ws2.cell(row=4, column=col, value=h)
    style_header(ws2, 4, 5, '0077B5')

    templates = [
        ["LinkedIn", "Tip 技巧",
         "🚢 3 Things to Check Before Shipping from China:\n\n1️⃣ Verify your supplier's export license\n2️⃣ Confirm Incoterms (FOB? CIF? DDP?)\n3️⃣ Get a freight quote BEFORE placing the order\n\nMost importers skip step 3 and end up paying 20-30% more.\n\nSave this for your next shipment! 💡\n\n#FreightForwarding #Malaysia #ImportFromChina #Logistics",
         "🚢 从中国发货前必查 3 件事：\n\n1️⃣ 确认供应商有出口资质\n2️⃣ 确认贸易条款（FOB? CIF? DDP?）\n\n3️⃣ 下单前先拿运费报价\n\n很多进口商跳过第3步，结果多付了20-30%的运费。\n\n收藏这条！💡\n\n#货运代理 #马来西亚 #中国进口 #物流",
         "#FreightForwarding #Malaysia #Logistics #ImportExport"],
        ["LinkedIn", "Case Study 案例",
         "📊 Case Study: How We Helped [Client] Save [X]% on Shipping\n\nChallenge: [Client] was overpaying for LCL shipments from [City], China.\n\nSolution: We consolidated shipments and optimized routing.\n\nResult:\n✅ [X]% cost reduction\n✅ [X] days faster delivery\n✅ Zero customs issues\n\nWant similar results? DM me for a free consultation.\n\n#CaseStudy #FreightForwarding #Malaysia",
         "📊 案例：帮 [Client] 省了 [X]% 的运费\n\n挑战：[Client] 从中国 [City] 拼柜运输成本过高。\n\n方案：我们整合货量 + 优化路线。\n\n结果：\n✅ 费用降 [X]%\n✅ 快了 [X] 天\n✅ 零清关问题\n\n想省运费？私信我免费咨询。\n\n#案例分享 #货运代理 #马来西亚",
         "#CaseStudy #Malaysia #Logistics"],
        ["Facebook", "Promo 促销",
         "🔥 LIMITED OFFER: Free Customs Clearance!\n\nBook your first shipment with us this month and we'll handle your customs clearance at NO extra cost.\n\n🚢 Sea freight China → MY\n✈️ Air freight available\n📍 Door-to-door delivery\n\nDM us or WhatsApp: [Number]\n\nHurry — only [X] spots left! 🏃‍♂️",
         "🔥 限时优惠：免费清关！\n\n本月第一次走货，清关费用全免！\n\n🚢 海运中国→马来西亚\n✈️ 空运可选\n📍 门到门派送\n\n私信或 WhatsApp：[Number]\n\n名额有限，先到先得！🏃‍♂️",
         "#Promotion #Malaysia #FreightForwarding"],
        ["小红书", "Tutorial 教程",
         "📦 中国到马来西亚海运全攻略！\n\n新手必看！手把手教你从中国进货到马来西亚 👇\n\n1️⃣ 找供应商（1688/阿里国际站）\n2️⃣ 确认价格 + 贸易条款\n3️⃣ 找货代订舱\n4️⃣ 供应商发货到仓库\n5️⃣ 报关 → 装柜 → 发货\n6️⃣ 清关 → 派送到门\n\n⏰ 海运 7-12天\n💰 拼柜 XX/CBM起\n\n有疑问评论区问我！💬",
         "",
         "#马来西亚海运 #中国进口 #跨境电商 #外贸"],
        ["小红书", "Cost Breakdown 费用",
         "💰 从中国进货到马来西亚，运费到底多少钱？\n\n以一个 20GP 柜子为例：\n\n📦 海运费：$XXX\n🛃 清关费：$XXX\n🚚 派送费：$XXX\n📄 文件费：$XXX\n—————————\n💵 总计：$XXX\n\n⚠️ 注意：不同货物、不同港口价格不一样！\n\n想拿准确报价？评论区扣 1 👇",
         "",
         "#马来西亚 #运费 #中国进货 #货代"],
    ]
    for i, row_data in enumerate(templates):
        row = 5 + i
        for col, val in enumerate(row_data, 1):
            ws2.cell(row=row, column=col, value=val)
        ws2.row_dimensions[5 + i].height = 150
    add_data_rows(ws2, 5, len(templates), 5)
    set_widths(ws2, [14, 18, 55, 55, 35])

    # --- Sheet 3: Content Calendar ---
    ws3 = wb.create_sheet("2.3 内容日历 Calendar")
    ws3.merge_cells('A1:H1')
    ws3.cell(row=1, column=1, value="📅 Social Media Agent — 月度内容日历").font = title_font
    ws3.row_dimensions[1].height = 35

    days = ["Mon 周一", "Tue 周二", "Wed 周三", "Thu 周四", "Fri 周五"]
    headers3 = ["Week 周"] + days
    for col, h in enumerate(headers3, 1):
        ws3.cell(row=3, column=col, value=h)
    style_header(ws3, 3, 6, 'BF8F00')

    calendar = [
        ["Week 1", "LinkedIn: Tips\n小红书: 教程",
         "Facebook: Promo\nLinkedIn: Insight",
         "Blog: Publish\nLinkedIn: Share",
         "小红书: 费用拆解\nFacebook: Tips",
         "LinkedIn: Case Study\nAll: Weekend recap"],
        ["Week 2", "LinkedIn: Tips\n小红书: 教程",
         "Facebook: Q&A\nLinkedIn: Industry",
         "Blog: Publish\nLinkedIn: Share",
         "小红书: 避坑指南\nFacebook: Tips",
         "LinkedIn: Case Study\nAll: Weekend recap"],
        ["Week 3", "LinkedIn: Tips\n小红书: 教程",
         "Facebook: Promo\nLinkedIn: Insight",
         "Blog: Publish\nLinkedIn: Share",
         "小红书: 费用对比\nFacebook: Tips",
         "LinkedIn: Case Study\nAll: Weekend recap"],
        ["Week 4", "LinkedIn: Tips\n小红书: 教程",
         "Facebook: Q&A\nLinkedIn: Industry",
         "Blog: Publish\nLinkedIn: Share",
         "小红书: 供应商推荐\nFacebook: Tips",
         "LinkedIn: Case Study\nAll: Month recap"],
    ]
    for i, row_data in enumerate(calendar):
        row = 4 + i
        for col, val in enumerate(row_data, 1):
            ws3.cell(row=row, column=col, value=val)
    add_data_rows(ws3, 4, len(calendar), 6)
    set_widths(ws3, [10, 25, 25, 25, 25, 25])
    for r in range(4, 8):
        ws3.row_dimensions[r].height = 50


# ============================================================
# Agent 3: Email Marketing Agent
# ============================================================
def create_email_agent(wb):
    # --- Sheet 1: Email Sequences ---
    ws = wb.create_sheet("3.1 邮件序列 Sequences")
    ws.merge_cells('A1:F1')
    ws.cell(row=1, column=1, value="📧 Email Marketing Agent — 自动邮件序列").font = title_font
    ws.row_dimensions[1].height = 35
    ws.merge_cells('A2:F2')
    ws.cell(row=2, column=1, value="Lead 进入 CRM → 自动触发邮件序列 → 培育 → 转化").font = small_font

    headers = ["Sequence 序列", "Day 天数", "Subject (EN)", "Subject (CN)", "Content Summary 内容概要", "CTA 行动号召"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=h)
    style_header(ws, 4, 6, 'D4444E')

    sequences = [
        # Welcome Sequence
        ["1. Welcome 欢迎", "Day 0", "Welcome! Here's your China-MY shipping guide",
         "欢迎！这是你的中马货运指南", "Welcome + free guide download + company intro", "Download guide"],
        ["1. Welcome 欢迎", "Day 2", "5 tips to save money on shipping",
         "5个省钱运输技巧", "Value-add tips + case study link", "Read case study"],
        ["1. Welcome 欢迎", "Day 5", "How much does shipping really cost?",
         "运费到底多少钱？", "Cost breakdown + calculator link", "Get a quote"],
        # Nurture Sequence
        ["2. Nurture 培育", "Day 10", "Why Malaysian importers choose us",
         "为什么马来进口商选择我们", "Testimonials + differentiators", "Book consultation"],
        ["2. Nurture 培育", "Day 15", "Common shipping mistakes (and how to avoid them)",
         "常见运输错误（及如何避免）", "Educational content + offer", "Avoid mistakes →"],
        ["2. Nurture 培育", "Day 22", "Your shipping questions answered",
         "你的运输问题解答", "FAQ + AEO content repurposed", "Ask your question"],
        # Conversion Sequence
        ["3. Convert 转化", "Day 28", "Special offer: [X]% off your first shipment",
         "首单优惠：运费立减 [X]%", "Limited-time offer + urgency", "Claim offer now"],
        ["3. Convert 转化", "Day 35", "Last chance: your discount expires tomorrow",
         "最后机会：优惠明天到期", "FOMO + reminder", "Use discount →"],
        # Newsletter
        ["4. Newsletter", "Weekly", "🚢 Shipping Update: [Month] rates & news",
         "🚢 货运资讯：[月] 运价&新闻", "Monthly rate updates + industry news + tips", "Read more"],
    ]
    for i, row_data in enumerate(sequences):
        row = 5 + i
        for col, val in enumerate(row_data, 1):
            ws.cell(row=row, column=col, value=val)
    add_data_rows(ws, 5, len(sequences), 6)
    set_widths(ws, [20, 10, 40, 30, 40, 18])
    ws.freeze_panes = 'A5'

    # --- Sheet 2: Email Templates ---
    ws2 = wb.create_sheet("3.2 邮件模板 Templates")
    ws2.merge_cells('A1:D1')
    ws2.cell(row=1, column=1, value="📧 Email Marketing Agent — 邮件模板").font = title_font
    ws2.row_dimensions[1].height = 35

    headers2 = ["Template Name 模板名", "Subject 主题", "Body (EN)", "Body (CN)"]
    for col, h in enumerate(headers2, 1):
        ws2.cell(row=3, column=col, value=h)
    style_header(ws2, 3, 4, 'D4444E')

    email_templates = [
        ["Welcome Email",
         "Welcome to [Company] — Your China-MY Shipping Partner! 🚢",
         "Hi [Name],\n\nWelcome! I'm Brandon, founder of [Company].\n\nWe help Malaysian businesses import from China — fast, reliable, and affordable shipping.\n\nHere's what you'll get from us:\n✅ Competitive sea & air freight rates\n✅ Full customs clearance\n✅ Door-to-door delivery\n✅ Real-time tracking\n\nAs a welcome gift, here's our free guide:\n📘 [Link to shipping guide]\n\nHave questions? Just reply to this email or WhatsApp me at [Number].\n\nLooking forward to working with you!\n\nBest,\nBrandon",
         "Hi [Name]，\n\n欢迎！我是 [Company] 的创始人 Brandon。\n\n我们帮马来西亚企业从中国进口 — 快速、可靠、实惠的货运服务。\n\n我们提供：\n✅ 优惠的海运/空运价格\n✅ 全程清关\n✅ 门到门派送\n✅ 实时追踪\n\n见面礼：免费货运指南 📘\n[Link]\n\n有问题直接回复或 WhatsApp：[Number]\n\n期待合作！\n\nBrandon"],
        ["Quote Follow-up",
         "Your shipping quote from [Company] 📋",
         "Hi [Name],\n\nThanks for reaching out! Here's your quote for:\n\n📦 From: [Origin]\n📍 To: [Destination]\n🚢 Method: [Sea/Air]\n💰 Rate: [Price]\n\nThis quote is valid for [X] days.\n\nReady to book? Reply 'YES' or WhatsApp [Number].\n\nBest,\nBrandon",
         "Hi [Name]，\n\n感谢询价！以下是您的报价：\n\n📦 从：[Origin]\n📍 到：[Destination]\n🚢 方式：[海运/空运]\n💰 价格：[Price]\n\n报价 [X] 天内有效。\n\n要下单？回复 'YES' 或 WhatsApp：[Number]。\n\nBrandon"],
    ]
    for i, row_data in enumerate(email_templates):
        row = 4 + i
        for col, val in enumerate(row_data, 1):
            ws2.cell(row=row, column=col, value=val)
        ws2.row_dimensions[4 + i].height = 200
    add_data_rows(ws2, 4, len(email_templates), 4)
    set_widths(ws2, [22, 40, 60, 60])


# ============================================================
# Agent 4: Brand Agent
# ============================================================
def create_brand_agent(wb):
    # --- Sheet 1: Brand Story ---
    ws = wb.create_sheet("4.1 品牌故事 Story")
    ws.merge_cells('A1:C1')
    ws.cell(row=1, column=1, value="🌟 Brand Agent — 品牌故事与定位").font = title_font
    ws.row_dimensions[1].height = 35

    headers = ["Element 元素", "Content (EN)", "Content (CN)"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    style_header(ws, 3, 3, '548235')

    brand_elements = [
        ["Mission 使命",
         "To make China-Malaysia shipping simple, affordable, and reliable for every SME.",
         "让每个中小企业的中马货运变得简单、实惠、可靠。"],
        ["Vision 愿景",
         "Become the #1 trusted freight partner for Malaysian SMEs importing from China.",
         "成为马来西亚中小企业从中国进口的第一信赖货运伙伴。"],
        ["Brand Promise 承诺",
         "Ship with confidence. We handle the complexity so you can focus on growing your business.",
         "放心发货。我们处理复杂问题，您专注发展业务。"],
        ["Unique Value 独特价值",
         "Bilingual team (EN/CN), deep China-MY expertise, dedicated account manager, transparent pricing.",
         "双语团队（英/中），深耕中马专线，专属客户经理，价格透明。"],
        ["Brand Voice 品牌语气",
         "Professional but approachable. Expert but not arrogant. Helpful, clear, and direct.",
         "专业但亲切。有经验但不傲慢。乐于助人、清晰、直接。"],
        ["Brand Story 品牌故事",
         "[Tell your story — why you started, what drives you, what makes you different]",
         "[讲述你的故事 — 为什么创业、动力是什么、什么让你与众不同]"],
    ]
    for i, row_data in enumerate(brand_elements):
        row = 4 + i
        for col, val in enumerate(row_data, 1):
            ws.cell(row=row, column=col, value=val)
        ws.row_dimensions[4 + i].height = 60
    add_data_rows(ws, 4, len(brand_elements), 3)
    set_widths(ws, [20, 60, 50])

    # --- Sheet 2: Case Studies ---
    ws2 = wb.create_sheet("4.2 案例研究 Cases")
    ws2.merge_cells('A1:G1')
    ws2.cell(row=1, column=1, value="📊 Brand Agent — 案例研究 Case Studies").font = title_font
    ws2.row_dimensions[1].height = 35

    headers2 = ["Client 客户", "Industry 行业", "Challenge 挑战", "Solution 方案", "Result 结果", "Status 状态", "Published 发布"]
    for col, h in enumerate(headers2, 1):
        ws2.cell(row=3, column=col, value=h)
    style_header(ws2, 3, 7, '548235')

    add_data_rows(ws2, 4, 10, 7)
    set_widths(ws2, [16, 14, 30, 30, 25, 12, 14])

    # --- Sheet 3: Testimonials ---
    ws3 = wb.create_sheet("4.3 客户评价 Testimonials")
    ws3.merge_cells('A1:E1')
    ws3.cell(row=1, column=1, value="💬 Brand Agent — 客户评价收集").font = title_font
    ws3.row_dimensions[1].height = 35

    headers3 = ["Client 客户", "Quote (EN) 评价", "Quote (CN)", "Platform 平台", "Status 状态"]
    for col, h in enumerate(headers3, 1):
        ws3.cell(row=3, column=col, value=h)
    style_header(ws3, 3, 5, '548235')

    add_data_rows(ws3, 4, 15, 5)
    set_widths(ws3, [16, 50, 40, 14, 12])


# ============================================================
# Agent 5: Paid Ads Agent (Ready for future)
# ============================================================
def create_ads_agent(wb):
    ws = wb.create_sheet("5.1 广告准备 Ads Ready")
    ws.merge_cells('A1:F1')
    ws.cell(row=1, column=1, value="💰 Paid Ads Agent — 广告投放准备（有机流量优先，广告备用）").font = title_font
    ws.row_dimensions[1].height = 35
    ws.merge_cells('A2:F2')
    ws.cell(row=2, column=1, value="先做有机流量 → 积累数据 → 准备好广告素材 → 未来投入付费广告").font = small_font

    headers = ["Platform 平台", "Ad Type 广告类型", "Target Audience 目标受众", "Ad Copy (EN)", "Ad Copy (CN)", "Status 状态"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=h)
    style_header(ws, 4, 6, 'E67E22')

    ads = [
        ["Google Search", "Search Ad 搜索广告",
         "MY businesses searching 'freight forwarder Malaysia'",
         "Ship from China to Malaysia — Fast & Affordable\nSea/Air Freight | Customs Clearance | Door-to-Door\nGet a Free Quote Today! 🚢",
         "中国到马来西亚货运 — 快速实惠\n海运/空运 | 清关 | 门到门\n免费报价！🚢", ""],
        ["Google Search", "Search Ad 搜索广告",
         "MY businesses searching 'shipping cost China Malaysia'",
         "China to Malaysia Shipping — Transparent Pricing\nNo Hidden Fees | Real-Time Tracking\nSee Our Rates Now →",
         "中国到马来西亚运费 — 价格透明\n无隐藏费 | 实时追踪\n查看费率 →", ""],
        ["Facebook", "Lead Gen 潜在客户",
         "MY SME owners, 25-55, interested in import/export",
         "Importing from China? Let us handle the logistics.\n🚢 Competitive rates\n🛃 Full customs clearance\n📍 Door-to-door delivery\nGet your free quote in 24 hours!",
         "从中国进口？我们帮你搞定物流。\n🚢 优惠运价\n🛃 全程清关\n📍 门到门派送\n24小时内免费报价！", ""],
        ["LinkedIn", "Sponsored Post",
         "MY logistics managers, import/export directors",
         "Stop overpaying for shipping from China.\nWe help Malaysian SMEs save [X]% on freight costs.\nFree consultation →",
         "别再为中国运输多花钱了。\n我们帮马来西亚中小企业省 [X]% 运费。\n免费咨询 →", ""],
    ]
    for i, row_data in enumerate(ads):
        row = 5 + i
        for col, val in enumerate(row_data, 1):
            ws.cell(row=row, column=col, value=val)
    add_data_rows(ws, 5, len(ads), 6)
    set_widths(ws, [16, 18, 35, 50, 40, 10])

    # Sheet 2: Budget Planner
    ws2 = wb.create_sheet("5.2 预算规划 Budget")
    ws2.merge_cells('A1:F1')
    ws2.cell(row=1, column=1, value="💰 Paid Ads Agent — 预算规划（备用）").font = title_font
    ws2.row_dimensions[1].height = 35

    headers2 = ["Month 月份", "Platform 平台", "Budget 预算 (MYR)", "Goal 目标", "Expected Leads 预期leads", "Status 状态"]
    for col, h in enumerate(headers2, 1):
        ws2.cell(row=3, column=col, value=h)
    style_header(ws2, 3, 6, 'E67E22')

    add_data_rows(ws2, 4, 12, 6)
    set_widths(ws2, [14, 16, 16, 25, 16, 10])


# ============================================================
# Agent Dashboard
# ============================================================
def create_dashboard(wb):
    ws = wb.create_sheet("📊 Dashboard 看板")
    ws.merge_cells('A1:F1')
    ws.cell(row=1, column=1, value="📊 Marketing Team Dashboard 营销团队看板").font = title_font
    ws.row_dimensions[1].height = 35

    # KPI Overview
    ws.cell(row=3, column=1, value="📈 本月 KPI").font = big_bold
    kpis = [
        (4, "Blog Posts Published 博客发布:", ""),
        (5, "Social Posts Published 社媒发布:", ""),
        (6, "Emails Sent 邮件发送:", ""),
        (7, "New Leads 新 Leads:", ""),
        (8, "Website Visits 网站访问:", ""),
        (9, "Social Followers 社媒粉丝:", ""),
        (10, "WhatsApp Inquiries WhatsApp 咨询:", ""),
    ]
    for row, label, val in kpis:
        ws.cell(row=row, column=1, value=label).font = bold_font
        ws.cell(row=row, column=1).border = thin_border
        ws.cell(row=row, column=2, value=val).fill = input_fill
        ws.cell(row=row, column=2).border = thin_border

    # Weekly checklist
    ws.cell(row=12, column=1, value="✅ 每周任务 Weekly Tasks").font = big_bold
    tasks = [
        (13, "Publish 2 blog posts (EN+CN) 发布2篇博客", ""),
        (14, "Post 5x on LinkedIn LinkedIn发5条", ""),
        (15, "Post 5x on Facebook Facebook发5条", ""),
        (16, "Post 3-5x on 小红书 小红书发3-5条", ""),
        (17, "Send newsletter 发newsletter", ""),
        (18, "Review analytics 复盘数据", ""),
        (19, "Update content calendar 更新内容日历", ""),
    ]
    for row, task, done in tasks:
        ws.cell(row=row, column=1, value=task).font = normal_font
        ws.cell(row=row, column=1).border = thin_border
        ws.cell(row=row, column=2, value=done).fill = input_fill
        ws.cell(row=row, column=2).border = thin_border

    set_widths(ws, [45, 15, 15, 15, 15, 15])


# ============================================================
# Main — Create all agent workbooks
# ============================================================

# === Workbook 1: Content Agent ===
wb1 = openpyxl.Workbook()
wb1.remove(wb1.active)
create_content_agent(wb1)
wb1.save("marketing/1_Content_Agent.xlsx")
print(f"✅ Content Agent → marketing/1_Content_Agent.xlsx ({len(wb1.worksheets)} sheets)")

# === Workbook 2: Social Media Agent ===
wb2 = openpyxl.Workbook()
wb2.remove(wb2.active)
create_social_agent(wb2)
wb2.save("marketing/2_Social_Media_Agent.xlsx")
print(f"✅ Social Media Agent → marketing/2_Social_Media_Agent.xlsx ({len(wb2.worksheets)} sheets)")

# === Workbook 3: Email Marketing Agent ===
wb3 = openpyxl.Workbook()
wb3.remove(wb3.active)
create_email_agent(wb3)
wb3.save("marketing/3_Email_Marketing_Agent.xlsx")
print(f"✅ Email Marketing Agent → marketing/3_Email_Marketing_Agent.xlsx ({len(wb3.worksheets)} sheets)")

# === Workbook 4: Brand Agent ===
wb4 = openpyxl.Workbook()
wb4.remove(wb4.active)
create_brand_agent(wb4)
wb4.save("marketing/4_Brand_Agent.xlsx")
print(f"✅ Brand Agent → marketing/4_Brand_Agent.xlsx ({len(wb4.worksheets)} sheets)")

# === Workbook 5: Paid Ads Agent ===
wb5 = openpyxl.Workbook()
wb5.remove(wb5.active)
create_ads_agent(wb5)
wb5.save("marketing/5_Paid_Ads_Agent.xlsx")
print(f"✅ Paid Ads Agent → marketing/5_Paid_Ads_Agent.xlsx ({len(wb5.worksheets)} sheets)")

# === Workbook 6: Marketing Dashboard ===
wb6 = openpyxl.Workbook()
wb6.remove(wb6.active)
create_dashboard(wb6)
wb6.save("marketing/0_Marketing_Dashboard.xlsx")
print(f"✅ Marketing Dashboard → marketing/0_Marketing_Dashboard.xlsx ({len(wb6.worksheets)} sheets)")

print(f"\n🎉 Marketing AI Agent Team Complete!")
print(f"   Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
print(f"   Total files: 6 Excel workbooks")
