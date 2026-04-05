"""
Document Agent - 货运文档自动生成器
自动生成：Commercial Invoice, Packing List, Booking Confirmation, Shipping Order
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime

header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
input_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
result_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
highlight_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
bold_font = Font(name='Arial', bold=True, size=10)
big_bold = Font(name='Arial', bold=True, size=13, color='2F5496')
title_font = Font(name='Arial', bold=True, size=18, color='2F5496')
normal_font = Font(name='Arial', size=10)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_wrap = Alignment(horizontal='left', vertical='center', wrap_text=True)

wb = openpyxl.Workbook()

# ============================================================
# Sheet 1: Shipment Input (填一次，所有文档自动拉数据)
# ============================================================
ws = wb.active
ws.title = "货物信息 Shipment"

ws.merge_cells('A1:F1')
ws.cell(row=1, column=1, value="📦 货物信息录入 Shipment Information")
ws.cell(row=1, column=1).font = title_font
ws.row_dimensions[1].height = 35

ws.merge_cells('A2:F2')
ws.cell(row=2, column=1, value="填黄色格子 → 切换到其他 Sheet 自动出文档 → 打印/导出 PDF 发给相关方")
ws.cell(row=2, column=1).font = Font(name='Arial', size=10, color='666666')

# === Shipper Info ===
ws.cell(row=4, column=1, value="📤 发货人 SHIPPER").font = big_bold

shipper = [
    (5, "公司名 Company:", ""),
    (6, "地址 Address:", ""),
    (7, "联系人 Contact:", ""),
    (8, "电话 Tel:", ""),
    (9, "邮箱 Email:", ""),
]
for row, label, default in shipper:
    ws.cell(row=row, column=1, value=label).font = bold_font
    ws.cell(row=row, column=1).border = thin_border
    ws.cell(row=row, column=2, value=default).font = normal_font
    ws.cell(row=row, column=2).fill = input_fill
    ws.cell(row=row, column=2).border = thin_border
ws.merge_cells('B6:F6')

# === Consignee Info ===
ws.cell(row=11, column=1, value="📥 收货人 CONSIGNEE").font = big_bold

consignee = [
    (12, "公司名 Company:", ""),
    (13, "地址 Address:", ""),
    (14, "联系人 Contact:", ""),
    (15, "电话 Tel:", ""),
    (16, "邮箱 Email:", ""),
]
for row, label, default in consignee:
    ws.cell(row=row, column=1, value=label).font = bold_font
    ws.cell(row=row, column=1).border = thin_border
    ws.cell(row=row, column=2, value=default).font = normal_font
    ws.cell(row=row, column=2).fill = input_fill
    ws.cell(row=row, column=2).border = thin_border
ws.merge_cells('B13:F13')

# === Shipment Details ===
ws.cell(row=18, column=1, value="🚢 运输信息 SHIPMENT DETAILS").font = big_bold

details = [
    (19, "运输方式 Mode:", "Sea FCL", "Sea FCL / Sea LCL / Air"),
    (20, "起运港 POL:", "", ""),
    (21, "卸货港 POD:", "", ""),
    (22, "最终目的地 Final Dest:", "", "East MY 填这里"),
    (23, "船名/航次 Vessel/Voy:", "", "确认后填"),
    (24, "提单号 B/L No:", "", "船公司给的"),
    (25, "订舱号 Booking Ref:", "", ""),
    (26, "柜号 Container No:", "", "装柜后填"),
    (27, "封号 Seal No:", "", "装柜后填"),
    (28, "截关日 Cut-off Date:", "", "Cutoff date"),
    (29, "ETD 预计离港:", "", "Estimated departure"),
    (30, "ETA 预计到港:", "", "Estimated arrival"),
    (31, "贸易条款 Incoterms:", "FOB", "FOB / CIF / EXW / DDP"),
    (32, "付款方式 Payment:", "T/T", "T/T / L/C / D/P"),
]
for row, label, default, note in details:
    ws.cell(row=row, column=1, value=label).font = bold_font
    ws.cell(row=row, column=1).border = thin_border
    ws.cell(row=row, column=2, value=default).font = normal_font
    ws.cell(row=row, column=2).fill = input_fill
    ws.cell(row=row, column=2).border = thin_border
    ws.cell(row=row, column=2).alignment = center
    if note:
        ws.cell(row=row, column=3, value=note).font = Font(name='Arial', italic=True, size=9, color='999999')

# Data validation
dv_mode = DataValidation(type="list", formula1='"Sea FCL,Sea LCL,Air"', allow_blank=True)
ws.add_data_validation(dv_mode)
dv_mode.add('B19')
dv_incoterms = DataValidation(type="list", formula1='"FOB,CIF,EXW,DDP,CFR,FCA"', allow_blank=True)
ws.add_data_validation(dv_incoterms)
dv_incoterms.add('B31')

# === Cargo Items (货品明细) ===
ws.cell(row=34, column=1, value="📋 货品明细 CARGO ITEMS").font = big_bold

cargo_headers = ["序号\nNo.", "品名\nDescription", "HS Code", "数量\nQty",
                 "单位\nUnit", "净重 KG\nNet Wt", "毛重 KG\nGross Wt",
                 "体积 CBM\nMeas.", "单价\nUnit Price", "总价\nTotal"]
for col, h in enumerate(cargo_headers, 1):
    cell = ws.cell(row=35, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = center

for r in range(36, 46):  # 10 rows for cargo items
    ws.cell(row=r, column=1, value=r-35).font = normal_font
    ws.cell(row=r, column=1).border = thin_border
    ws.cell(row=r, column=1).alignment = center
    for c in range(2, 11):
        cell = ws.cell(row=r, column=c)
        cell.font = normal_font
        cell.fill = input_fill
        cell.border = thin_border
        cell.alignment = center
        if c >= 6:
            cell.number_format = '#,##0.00'

# Total row
ws.cell(row=46, column=1, value="合计 TOTAL").font = bold_font
ws.cell(row=46, column=1).border = thin_border
for c in [6, 7, 8, 10]:
    cell = ws.cell(row=46, column=c)
    cell.font = bold_font
    cell.border = thin_border
    cell.alignment = center
    cell.fill = highlight_fill
    cell.number_format = '#,##0.00'

ws.cell(row=48, column=1, value="币种 Currency:").font = bold_font
ws.cell(row=48, column=2, value="USD").font = normal_font
ws.cell(row=48, column=2).fill = input_fill
ws.cell(row=48, column=2).border = thin_border

# Column widths
widths = [20, 25, 14, 12, 10, 14, 14, 14, 14, 14]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ============================================================
# Sheet 2: Commercial Invoice (商业发票)
# ============================================================
ws_ci = wb.create_sheet("Invoice 发票")

ws_ci.merge_cells('A1:E1')
ws_ci.cell(row=1, column=1, value="COMMERCIAL INVOICE").font = Font(name='Arial', bold=True, size=20, color='2F5496')
ws_ci.row_dimensions[1].height = 40

# Invoice info
ws_ci.cell(row=3, column=1, value="Invoice No:").font = bold_font
ws_ci.cell(row=3, column=2, value='=CONCATENATE("INV-",货物信息 Shipment!B25)').font = normal_font
ws_ci.cell(row=3, column=4, value="Date:").font = bold_font
ws_ci.cell(row=3, column=5, value=datetime.now().strftime("%Y-%m-%d")).font = normal_font
ws_ci.cell(row=4, column=4, value="B/L No:").font = bold_font
ws_ci.cell(row=4, column=5, value="=货物信息 Shipment!B24").font = normal_font
ws_ci.cell(row=5, column=4, value="Booking Ref:").font = bold_font
ws_ci.cell(row=5, column=5, value="=货物信息 Shipment!B25").font = normal_font

# Shipper / Consignee
ws_ci.cell(row=7, column=1, value="SHIPPER / EXPORTER:").font = Font(name='Arial', bold=True, size=10, color='2F5496')
ws_ci.cell(row=8, column=1, value="=货物信息 Shipment!B5").font = bold_font
ws_ci.cell(row=9, column=1, value="=货物信息 Shipment!B6").font = normal_font
ws_ci.cell(row=10, column=1, value='=CONCATENATE("Tel: ",货物信息 Shipment!B8)').font = normal_font

ws_ci.cell(row=7, column=4, value="CONSIGNEE / IMPORTER:").font = Font(name='Arial', bold=True, size=10, color='2F5496')
ws_ci.cell(row=8, column=4, value="=货物信息 Shipment!B12").font = bold_font
ws_ci.cell(row=9, column=4, value="=货物信息 Shipment!B13").font = normal_font
ws_ci.cell(row=10, column=4, value='=CONCATENATE("Tel: ",货物信息 Shipment!B15)').font = normal_font

# Shipment info
ws_ci.cell(row=12, column=1, value="FROM:").font = bold_font
ws_ci.cell(row=12, column=2, value="=货物信息 Shipment!B20").font = normal_font  # POL
ws_ci.cell(row=12, column=4, value="TO:").font = bold_font
ws_ci.cell(row=12, column=5, value="=货物信息 Shipment!B21").font = normal_font  # POD
ws_ci.cell(row=13, column=1, value="VESSEL/VOY:").font = bold_font
ws_ci.cell(row=13, column=2, value="=货物信息 Shipment!B23").font = normal_font
ws_ci.cell(row=13, column=4, value="INCOTERMS:").font = bold_font
ws_ci.cell(row=13, column=5, value="=货物信息 Shipment!B31").font = normal_font

# Cargo table
ws_ci.cell(row=15, column=1, value="DESCRIPTION OF GOODS").font = bold_font

ci_headers = ["Description", "HS Code", "Qty", "Unit", "Gross Wt (KG)", "Meas (CBM)", "Unit Price", "Total"]
for col, h in enumerate(ci_headers, 1):
    cell = ws_ci.cell(row=16, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = center

for r in range(17, 27):  # 10 rows
    src = r - 17 + 36  # Source row in Shipment sheet
    ws_ci.cell(row=r, column=1, value=f"=货物信息 Shipment!B{src}").font = normal_font
    ws_ci.cell(row=r, column=2, value=f"=货物信息 Shipment!C{src}").font = normal_font
    ws_ci.cell(row=r, column=3, value=f"=货物信息 Shipment!D{src}").font = normal_font
    ws_ci.cell(row=r, column=4, value=f"=货物信息 Shipment!E{src}").font = normal_font
    ws_ci.cell(row=r, column=5, value=f"=货物信息 Shipment!G{src}").font = normal_font
    ws_ci.cell(row=r, column=6, value=f"=货物信息 Shipment!H{src}").font = normal_font
    ws_ci.cell(row=r, column=7, value=f"=货物信息 Shipment!I{src}").font = normal_font
    ws_ci.cell(row=r, column=8, value=f"=货物信息 Shipment!J{src}").font = normal_font
    for c in range(1, 9):
        cell = ws_ci.cell(row=r, column=c)
        cell.border = thin_border
        cell.alignment = center
        if c >= 5:
            cell.number_format = '#,##0.00'

# Total row
ws_ci.cell(row=27, column=1, value="TOTAL").font = bold_font
ws_ci.cell(row=27, column=1).border = thin_border
for c, col_letter in [(3, 'D'), (5, 'G'), (6, 'H'), (8, 'J')]:
    cell = ws_ci.cell(row=27, column=c)
    cell.font = bold_font
    cell.border = thin_border
    cell.fill = highlight_fill
    cell.alignment = center
    cell.number_format = '#,##0.00'

ws_ci.cell(row=29, column=1, value="TOTAL AMOUNT:").font = Font(name='Arial', bold=True, size=12)
ws_ci.cell(row=29, column=3, value="=货物信息 Shipment!B48").font = Font(name='Arial', bold=True, size=12)
ws_ci.cell(row=29, column=5, value="=货物信息 Shipment!J46").font = Font(name='Arial', bold=True, size=14, color='2F5496')
ws_ci.cell(row=29, column=5).number_format = '#,##0.00'

# CI column widths
ci_widths = [25, 14, 10, 10, 16, 14, 14, 14]
for i, w in enumerate(ci_widths, 1):
    ws_ci.column_dimensions[get_column_letter(i)].width = w

# ============================================================
# Sheet 3: Packing List (装箱单)
# ============================================================
ws_pl = wb.create_sheet("Packing List 装箱单")

ws_pl.merge_cells('A1:E1')
ws_pl.cell(row=1, column=1, value="PACKING LIST").font = Font(name='Arial', bold=True, size=20, color='2F5496')
ws_pl.row_dimensions[1].height = 40

ws_pl.cell(row=3, column=1, value="P/L No:").font = bold_font
ws_pl.cell(row=3, column=2, value='=CONCATENATE("PL-",货物信息 Shipment!B25)').font = normal_font
ws_pl.cell(row=3, column=4, value="Date:").font = bold_font
ws_pl.cell(row=3, column=5, value=datetime.now().strftime("%Y-%m-%d")).font = normal_font
ws_pl.cell(row=4, column=4, value="B/L No:").font = bold_font
ws_pl.cell(row=4, column=5, value="=货物信息 Shipment!B24").font = normal_font

ws_pl.cell(row=6, column=1, value="SHIPPER:").font = Font(name='Arial', bold=True, size=10, color='2F5496')
ws_pl.cell(row=7, column=1, value="=货物信息 Shipment!B5").font = bold_font
ws_pl.cell(row=8, column=1, value="=货物信息 Shipment!B6").font = normal_font

ws_pl.cell(row=6, column=4, value="CONSIGNEE:").font = Font(name='Arial', bold=True, size=10, color='2F5496')
ws_pl.cell(row=7, column=4, value="=货物信息 Shipment!B12").font = bold_font
ws_pl.cell(row=8, column=4, value="=货物信息 Shipment!B13").font = normal_font

ws_pl.cell(row=10, column=1, value="FROM:").font = bold_font
ws_pl.cell(row=10, column=2, value="=货物信息 Shipment!B20").font = normal_font
ws_pl.cell(row=10, column=4, value="TO:").font = bold_font
ws_pl.cell(row=10, column=5, value="=货物信息 Shipment!B21").font = normal_font
ws_pl.cell(row=11, column=1, value="CONTAINER:").font = bold_font
ws_pl.cell(row=11, column=2, value="=货物信息 Shipment!B26").font = normal_font
ws_pl.cell(row=11, column=4, value="SEAL:").font = bold_font
ws_pl.cell(row=11, column=5, value="=货物信息 Shipment!B27").font = normal_font

ws_pl.cell(row=13, column=1, value="PACKING DETAILS").font = bold_font

pl_headers = ["Description", "Qty", "Unit", "Net Wt (KG)", "Gross Wt (KG)", "CBM"]
for col, h in enumerate(pl_headers, 1):
    cell = ws_pl.cell(row=14, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = center

for r in range(15, 25):
    src = r - 15 + 36
    ws_pl.cell(row=r, column=1, value=f"=货物信息 Shipment!B{src}").font = normal_font
    ws_pl.cell(row=r, column=2, value=f"=货物信息 Shipment!D{src}").font = normal_font
    ws_pl.cell(row=r, column=3, value=f"=货物信息 Shipment!E{src}").font = normal_font
    ws_pl.cell(row=r, column=4, value=f"=货物信息 Shipment!F{src}").font = normal_font
    ws_pl.cell(row=r, column=5, value=f"=货物信息 Shipment!G{src}").font = normal_font
    ws_pl.cell(row=r, column=6, value=f"=货物信息 Shipment!H{src}").font = normal_font
    for c in range(1, 7):
        cell = ws_pl.cell(row=r, column=c)
        cell.border = thin_border
        cell.alignment = center
        if c >= 4:
            cell.number_format = '#,##0.00'

ws_pl.cell(row=25, column=1, value="TOTAL").font = bold_font
ws_pl.cell(row=25, column=1).border = thin_border
for c in [2, 4, 5, 6]:
    cell = ws_pl.cell(row=25, column=c)
    cell.font = bold_font
    cell.border = thin_border
    cell.fill = highlight_fill
    cell.alignment = center
    cell.number_format = '#,##0.00'

pl_widths = [25, 10, 10, 14, 14, 14]
for i, w in enumerate(pl_widths, 1):
    ws_pl.column_dimensions[get_column_letter(i)].width = w

# ============================================================
# Sheet 4: Document Checklist (文档清单)
# ============================================================
ws_dc = wb.create_sheet("文档清单 Checklist")

ws_dc.merge_cells('A1:F1')
ws_dc.cell(row=1, column=1, value="📋 文档清单 Document Checklist").font = title_font

ws_dc.cell(row=3, column=1, value="订单号 Booking Ref:").font = bold_font
ws_dc.cell(row=3, column=2, value="=货物信息 Shipment!B25").font = normal_font
ws_dc.cell(row=3, column=4, value="客户:").font = bold_font
ws_dc.cell(row=3, column=5, value="=货物信息 Shipment!B12").font = normal_font

dc_headers = ["序号", "文件 Document", "状态 Status", "负责方 Responsible", "截止日 Deadline", "备注 Notes"]
for col, h in enumerate(dc_headers, 1):
    cell = ws_dc.cell(row=5, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = center

docs = [
    [1, "Commercial Invoice 商业发票", "", "Shipper → You", "", "必须有"],
    [2, "Packing List 装箱单", "", "Shipper → You", "", "必须有"],
    [3, "Bill of Lading (B/L) 提单", "", "Shipping Line", "", "船公司出"],
    [4, "Booking Confirmation 订舱确认", "", "You", "", ""],
    [5, "Shipping Order (S/O) 托运单", "", "You → Shipping Line", "", ""],
    [6, "Customs Declaration 报关单", "", "Customs Broker", "", "需要报关行"],
    [7, "Certificate of Origin (CO) 原产地证", "", "Chamber of Commerce", "", "看客户要求"],
    [8, "Fumigation Certificate 熏蒸证书", "", "Fumigation Co.", "", "木质包装需要"],
    [9, "Insurance Certificate 保险单", "", "Insurance Co.", "", "如客户要求"],
    [10, "Inspection Report 验货报告", "", "Inspector", "", "如需要"],
]

for r, data in enumerate(docs, 6):
    for c, val in enumerate(data, 1):
        cell = ws_dc.cell(row=r, column=c, value=val)
        cell.font = normal_font
        cell.border = thin_border
        cell.alignment = center if c != 2 else left_wrap
    # Status column with dropdown
    ws_dc.cell(row=r, column=3).fill = input_fill

dv_status = DataValidation(
    type="list",
    formula1='"✅ 已完成,⏳ 进行中,⬜ 未开始,❌ 不需要,⚠️ 缺失"',
    allow_blank=True
)
ws_dc.add_data_validation(dv_status)
dv_status.add('C6:C15')

dc_widths = [5, 32, 14, 22, 14, 20]
for i, w in enumerate(dc_widths, 1):
    ws_dc.column_dimensions[get_column_letter(i)].width = w

# ============================================================
# Sheet 5: Shipping Schedule (船期表)
# ============================================================
ws_ss = wb.create_sheet("船期表 Schedule")

ws_ss.merge_cells('A1:G1')
ws_ss.cell(row=1, column=1, value="🚢 船期跟踪 Shipping Schedule Tracker").font = title_font

ss_headers = ["订单号\nBooking", "客户\nCustomer", "航线\nRoute",
              "截关日\nCut-off", "ETD\n离港", "ETA\n到港", "状态\nStatus"]
for col, h in enumerate(ss_headers, 1):
    cell = ws_ss.cell(row=3, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = center

for r in range(4, 24):  # 20 rows
    for c in range(1, 8):
        cell = ws_ss.cell(row=r, column=c)
        cell.border = thin_border
        cell.alignment = center
    ws_ss.cell(row=r, column=7).fill = input_fill

dv_ship_status = DataValidation(
    type="list",
    formula1='"📦 已订舱,🏭 装柜中,🚢 已离港,📍 在途,✅ 已到港,📋 已放行"',
    allow_blank=True
)
ws_ss.add_data_validation(dv_ship_status)
dv_ship_status.add('G4:G23')

ss_widths = [16, 18, 20, 14, 14, 14, 14]
for i, w in enumerate(ss_widths, 1):
    ws_ss.column_dimensions[get_column_letter(i)].width = w

# Save
wb.save('/root/.openclaw/workspace/freight-agent/Document_Agent.xlsx')
print("✅ Document_Agent.xlsx created")
