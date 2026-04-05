"""
SEO AI Agent — 专业搜索引擎优化
目标：马来西亚中小企业 | 双语 | 有机流量增长
覆盖：技术SEO + 站内优化 + 外链 + 本地SEO + 排名追踪 + 竞品分析
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime

# Styles
header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
input_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
blue_fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
grey_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
orange_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
purple_fill = PatternFill(start_color='E2D0F0', end_color='E2D0F0', fill_type='solid')
bold_font = Font(name='Arial', bold=True, size=10)
big_bold = Font(name='Arial', bold=True, size=13, color='2F5496')
title_font = Font(name='Arial', bold=True, size=18, color='2F5496')
subtitle_font = Font(name='Arial', bold=True, size=13, color='2F5496')
small_font = Font(name='Arial', size=9, color='666666')
normal_font = Font(name='Arial', size=10)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_wrap = Alignment(horizontal='left', vertical='center', wrap_text=True)

def style_header(ws, row, cols, fill_color='2F5496'):
    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = fill
        cell.border = thin_border
        cell.alignment = center

def add_rows(ws, start, count, cols):
    for row in range(start, start + count):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.font = normal_font
            cell.fill = input_fill
            cell.alignment = left_wrap

def widths(ws, w):
    for i, v in enumerate(w, 1):
        ws.column_dimensions[get_column_letter(i)].width = v

wb = openpyxl.Workbook()
wb.remove(wb.active)

# ============================================================
# Sheet 1: Technical SEO Audit 技术SEO审计
# ============================================================
ws1 = wb.create_sheet("技术SEO Audit")
ws1.merge_cells('A1:F1')
ws1.cell(row=1, column=1, value="🔧 SEO Agent — 技术SEO审计清单").font = title_font
ws1.row_dimensions[1].height = 35
ws1.merge_cells('A2:F2')
ws1.cell(row=2, column=1, value="每月检查一次 → 修好所有问题 → Google 才会爱你").font = small_font

h1 = ["Category 分类", "Check Item 检查项", "Tool 工具", "Result 结果", "Priority 优先级", "Status 状态"]
for col, h in enumerate(h1, 1):
    ws1.cell(row=4, column=col, value=h)
style_header(ws1, 4, 6)

priority = "High 高,Medium 中,Low 低"
dv_p = DataValidation(type="list", formula1=f'"{priority}"', allow_blank=True)
ws1.add_data_validation(dv_p)
status_opt = "Pass ✅,Fail ❌,Warning ⚠️,Fixed 🔧,N/A"
dv_s = DataValidation(type="list", formula1=f'"{status_opt}"', allow_blank=True)
ws1.add_data_validation(dv_s)

tech_checks = [
    # Site Speed
    ["⚡ Site Speed 网站速度", "Page load time < 3 seconds", "Google PageSpeed Insights", "", "High 高", ""],
    ["⚡ Site Speed 网站速度", "Core Web Vitals (LCP, FID, CLS)", "Google Search Console", "", "High 高", ""],
    ["⚡ Site Speed 网站速度", "Image compression & WebP format", "TinyPNG / Squoosh", "", "Medium 中", ""],
    ["⚡ Site Speed 网站速度", "Enable browser caching", "GTMetrix", "", "Medium 中", ""],
    ["⚡ Site Speed 网站速度", "Minify CSS/JS", "GTMetrix", "", "Medium 中", ""],
    ["⚡ Site Speed 网站速度", "Use CDN (Cloudflare)", "Cloudflare", "", "Medium 中", ""],

    # Mobile
    ["📱 Mobile 移动端", "Mobile responsive design", "Google Mobile Test", "", "High 高", ""],
    ["📱 Mobile 移动端", "Touch-friendly buttons", "Manual check", "", "High 高", ""],
    ["📱 Mobile 移动端", "No horizontal scrolling", "Manual check", "", "High 高", ""],

    # Crawlability
    ["🕷️ Crawl 抓取", "Robots.txt properly configured", "robots.txt checker", "", "High 高", ""],
    ["🕷️ Crawl 抓取", "XML Sitemap submitted to GSC", "Google Search Console", "", "High 高", ""],
    ["🕷️ Crawl 抓取", "No broken links (404s)", "Screaming Frog / Ahrefs", "", "High 高", ""],
    ["🕷️ Crawl 抓取", "Canonical tags set correctly", "Screaming Frog", "", "Medium 中", ""],
    ["🕷️ Crawl 抓取", "No duplicate content", "Siteliner", "", "Medium 中", ""],

    # HTTPS & Security
    ["🔒 Security 安全", "HTTPS enabled (SSL certificate)", "SSL Checker", "", "High 高", ""],
    ["🔒 Security 安全", "No mixed content warnings", "Why No Padlock", "", "High 高", ""],
    ["🔒 Security 安全", "HSTS header enabled", "Security Headers", "", "Low 低", ""],

    # Schema/Structured Data
    ["📊 Schema 结构化数据", "Organization schema markup", "Google Rich Results Test", "", "High 高", ""],
    ["📊 Schema 结构化数据", "LocalBusiness schema", "Schema Markup Validator", "", "High 高", ""],
    ["📊 Schema 结构化数据", "FAQ schema on FAQ pages", "Google Rich Results Test", "", "Medium 中", ""],
    ["📊 Schema 结构化数据", "Breadcrumb schema", "Schema Markup Validator", "", "Medium 中", ""],
    ["📊 Schema 结构化数据", "Service schema on service pages", "Schema Markup Validator", "", "Medium 中", ""],

    # International SEO
    ["🌏 i18n 国际化", "Hreflang tags (EN ↔ CN)", "Ahrefs / Screaming Frog", "", "High 高", ""],
    ["🌏 i18n 国际化", "Language switcher works", "Manual check", "", "High 高", ""],
    ["🌏 i18n 国际化", "URL structure (/en/ /cn/ or subdomain)", "Manual check", "", "Medium 中", ""],

    # Analytics
    ["📈 Analytics 分析", "Google Analytics 4 installed", "GA4", "", "High 高", ""],
    ["📈 Analytics 分析", "Google Search Console connected", "GSC", "", "High 高", ""],
    ["📈 Analytics 分析", "Goal/conversion tracking set up", "GA4", "", "High 高", ""],
]

for i, row_data in enumerate(tech_checks):
    row = 5 + i
    for col, val in enumerate(row_data, 1):
        ws1.cell(row=row, column=col, value=val)
    dv_p.add(ws1.cell(row=row, column=5))
    dv_s.add(ws1.cell(row=row, column=6))

add_rows(ws1, 5, len(tech_checks), 6)
widths(ws1, [22, 38, 25, 16, 14, 12])
ws1.freeze_panes = 'A5'


# ============================================================
# Sheet 2: On-Page SEO 站内优化
# ============================================================
ws2 = wb.create_sheet("站内优化 On-Page")
ws2.merge_cells('A1:I1')
ws2.cell(row=1, column=1, value="📄 SEO Agent — 站内SEO优化清单").font = title_font
ws2.row_dimensions[1].height = 35
ws2.merge_cells('A2:I2')
ws2.cell(row=2, column=1, value="每页都按这个标准优化 → 标题/描述/标题标签/内部链接/图片ALT").font = small_font

h2 = ["Page URL 页面", "Title Tag 标题(≤60字)", "Meta Description 描述(≤160字)",
      "H1 标题", "Target Keyword 目标关键词", "Internal Links 内链",
      "Image ALT 图片ALT", "Schema 结构化数据", "Status 状态"]
for col, h in enumerate(h2, 1):
    ws2.cell(row=4, column=col, value=h)
style_header(ws2, 4, 9)

add_rows(ws2, 5, 20, 9)
widths(ws2, [25, 35, 40, 25, 22, 18, 18, 16, 10])
ws2.freeze_panes = 'A5'

# On-page checklist below the table
check_row = 27
ws2.cell(row=check_row, column=1, value="✅ 每页优化清单 On-Page Checklist").font = big_bold
checks = [
    "Title tag includes primary keyword + brand name",
    "Meta description includes keyword + CTA",
    "Only ONE H1 per page, includes keyword",
    "H2/H3 hierarchy is logical",
    "Keyword appears in first 100 words",
    "Keyword density: 1-2% (natural, not stuffed)",
    "3+ internal links to related pages",
    "All images have descriptive ALT text",
    "URL is short, descriptive, includes keyword",
    "Page has FAQ section (for AEO / featured snippets)",
    "Schema markup added (Organization, FAQ, Breadcrumb)",
    "Content is 1500+ words for key pages",
    "Content is bilingual (EN + CN) with hreflang",
    "Social sharing meta tags (OG tags) set",
]
for i, check in enumerate(checks):
    ws2.cell(row=check_row + 1 + i, column=1, value=f"☐ {check}").font = normal_font


# ============================================================
# Sheet 3: Local SEO 本地SEO
# ============================================================
ws3 = wb.create_sheet("本地SEO Local")
ws3.merge_cells('A1:E1')
ws3.cell(row=1, column=1, value="📍 SEO Agent — 本地SEO (Google My Business)").font = title_font
ws3.row_dimensions[1].height = 35
ws3.merge_cells('A2:E2')
ws3.cell(row=2, column=1, value="本地搜索 'freight forwarder near me' → 让你在地图上出现").font = small_font

h3 = ["Task 任务", "Details 详情", "Tool 工具", "Done 完成", "Notes 备注"]
for col, h in enumerate(h3, 1):
    ws3.cell(row=4, column=col, value=h)
style_header(ws3, 4, 5, '548235')

local_tasks = [
    ["Google My Business 基础", "Claim & verify Google Business Profile", "Google Business", "", ""],
    ["Google My Business 基础", "Complete all business info (name, address, phone, hours)", "Google Business", "", ""],
    ["Google My Business 基础", "Add all service categories", "Google Business", "", ""],
    ["Google My Business 基础", "Write compelling business description (EN+CN)", "Google Business", "", ""],
    ["Google My Business 基础", "Add high-quality photos (office, team, shipments)", "Google Business", "", ""],

    ["Reviews 评价", "Get 10+ Google reviews (ask happy customers)", "Manual", "", ""],
    ["Reviews 评价", "Respond to ALL reviews (positive & negative)", "Google Business", "", ""],
    ["Reviews 评价", "Create review link to share with customers", "GMB Review Link", "", ""],

    ["Local Citations 本地引用", "List on Yellow Pages MY", "yellowpages.my", "", ""],
    ["Local Citations 本地引用", "List on FMM Directory", "fmm.org.my", "", ""],
    ["Local Citations 本地引用", "List on MATRADE Directory", "matrade.gov.my", "", ""],
    ["Local Citations 本地引用", "List on Malaysia Freight Forwarders Directory", "Various", "", ""],
    ["Local Citations 本地引用", "Ensure NAP (Name/Address/Phone) consistent everywhere", "Moz Local", "", ""],

    ["Local Content 本地内容", "Create KL-specific landing page", "Website", "", ""],
    ["Local Content 本地内容", "Create Penang-specific landing page", "Website", "", ""],
    ["Local Content 本地内容", "Create Johor-specific landing page", "Website", "", ""],
    ["Local Content 本地内容", "Add Google Maps embed to contact page", "Website", "", ""],
]

for i, row_data in enumerate(local_tasks):
    row = 5 + i
    for col, val in enumerate(row_data, 1):
        ws3.cell(row=row, column=col, value=val)
    ws3.cell(row=row, column=4).fill = input_fill
add_rows(ws3, 5, len(local_tasks), 5)
widths(ws3, [25, 50, 20, 8, 20])


# ============================================================
# Sheet 4: Link Building 外链建设
# ============================================================
ws4 = wb.create_sheet("外链建设 Backlinks")
ws4.merge_cells('A1:G1')
ws4.cell(row=1, column=1, value="🔗 SEO Agent — 外链建设策略").font = title_font
ws4.row_dimensions[1].height = 35
ws4.merge_cells('A2:G2')
ws4.cell(row=2, column=1, value="高质量外链 = Google 信任你 → 排名上升").font = small_font

h4 = ["Source 来源", "Type 类型", "URL", "DA 域名权重", "Anchor Text 锚文本", "Status 状态", "Notes 备注"]
for col, h in enumerate(h4, 1):
    ws4.cell(row=4, column=col, value=h)
style_header(ws4, 4, 7, 'E67E22')

link_types = "Guest Post,Directory,Citation,Partner,PR/News,Social Profile,Forum,Blog Comment"
dv_lt = DataValidation(type="list", formula1=f'"{link_types}"', allow_blank=True)
ws4.add_data_validation(dv_lt)

links = [
    # Directories
    ["Google My Business", "Citation", "", "—", "Brand name", "", ""],
    ["Yellow Pages MY", "Citation", "yellowpages.my", "50+", "Brand name", "", ""],
    ["MATRADE", "Directory", "matrade.gov.my", "70+", "Brand name", "", ""],
    ["FMM", "Directory", "fmm.org.my", "60+", "Brand name", "", ""],
    ["LinkedIn Company", "Social Profile", "linkedin.com/company", "90+", "Brand name", "", ""],
    ["Facebook Page", "Social Profile", "facebook.com", "90+", "Brand name", "", ""],
    ["小红书", "Social Profile", "xiaohongshu.com", "80+", "Brand name", "", ""],

    # Guest Posts / PR
    ["Logistics trade blog", "Guest Post", "", "30+", "freight forwarder Malaysia", "", ""],
    ["SME business blog MY", "Guest Post", "", "30+", "shipping China Malaysia", "", ""],
    ["Trade publication", "PR/News", "", "40+", "logistics Malaysia", "", ""],

    # Partner links
    ["Supplier websites", "Partner", "", "—", "shipping partner", "", ""],
    ["Customer testimonials", "Partner", "", "—", "freight forwarding", "", ""],
]

for i, row_data in enumerate(links):
    row = 5 + i
    for col, val in enumerate(row_data, 1):
        ws4.cell(row=row, column=col, value=val)
    dv_lt.add(ws4.cell(row=row, column=2))
add_rows(ws4, 5, len(links), 7)
widths(ws4, [22, 16, 28, 14, 24, 12, 20])
ws4.freeze_panes = 'A5'


# ============================================================
# Sheet 5: Ranking Tracker 排名追踪
# ============================================================
ws5 = wb.create_sheet("排名追踪 Rankings")
ws5.merge_cells('A1:K1')
ws5.cell(row=1, column=1, value="📊 SEO Agent — 关键词排名追踪").font = title_font
ws5.row_dimensions[1].height = 35
ws5.merge_cells('A2:K2')
ws5.cell(row=2, column=1, value="每周记录一次排名 → 看趋势 → 调策略").font = small_font

h5 = ["Keyword 关键词", "Language 语言", "Page URL",
      "Week 1", "Week 2", "Week 3", "Week 4",
      "Week 5", "Week 6", "Week 7", "Week 8"]
for col, h in enumerate(h5, 1):
    ws5.cell(row=4, column=col, value=h)
style_header(ws5, 4, 11, 'D4444E')

track_keywords = [
    # English
    ["freight forwarding Malaysia", "EN", ""],
    ["shipping from China to Malaysia", "EN", ""],
    ["sea freight Malaysia", "EN", ""],
    ["air freight Malaysia", "EN", ""],
    ["customs clearance Malaysia", "EN", ""],
    ["shipping cost China Malaysia", "EN", ""],
    ["freight forwarder KL", "EN", ""],
    ["import from China to Malaysia", "EN", ""],
    ["LCL shipping Malaysia", "EN", ""],
    ["3PL logistics Malaysia", "EN", ""],
    # Chinese
    ["中国到马来西亚海运", "CN", ""],
    ["马来西亚货运代理", "CN", ""],
    ["中国到马来西亚运费", "CN", ""],
    ["马来西亚清关", "CN", ""],
    ["海运到马来西亚", "CN", ""],
    ["马来西亚进口中国货物", "CN", ""],
    ["吉隆坡货运公司", "CN", ""],
    ["马来西亚电商物流", "CN", ""],
]

for i, row_data in enumerate(track_keywords):
    row = 5 + i
    for col, val in enumerate(row_data, 1):
        ws5.cell(row=row, column=col, value=val)
add_rows(ws5, 5, len(track_keywords), 11)
widths(ws5, [32, 10, 25, 10, 10, 10, 10, 10, 10, 10, 10])
ws5.freeze_panes = 'D5'

# Ranking color legend
ws5.cell(row=25, column=1, value="📊 排名颜色说明:").font = bold_font
ws5.cell(row=26, column=1, value="Top 3 (Position 1-3)").font = normal_font
ws5.cell(row=26, column=1).fill = green_fill
ws5.cell(row=27, column=1, value="Page 1 (Position 4-10)").font = normal_font
ws5.cell(row=27, column=1).fill = yellow_fill
ws5.cell(row=28, column=1, value="Page 2-3 (Position 11-30)").font = normal_font
ws5.cell(row=28, column=1).fill = orange_fill
ws5.cell(row=29, column=1, value="Not Ranking (30+)").font = normal_font
ws5.cell(row=29, column=1).fill = red_fill


# ============================================================
# Sheet 6: Competitor Analysis 竞品分析
# ============================================================
ws6 = wb.create_sheet("竞品分析 Competitors")
ws6.merge_cells('A1:H1')
ws6.cell(row=1, column=1, value="🕵️ SEO Agent — 竞品SEO分析").font = title_font
ws6.row_dimensions[1].height = 35
ws6.merge_cells('A2:H2')
ws6.cell(row=2, column=1, value="分析竞品的SEO策略 → 找到差距 → 超越他们").font = small_font

h6 = ["Competitor 竞品", "Website 网站", "DA 域名权重", "Top Keywords 排名关键词",
      "Content Strategy 内容策略", "Backlinks 外链数量", "Their Strength 优势", "Our Opportunity 我们的机会"]
for col, h in enumerate(h6, 1):
    ws6.cell(row=4, column=col, value=h)
style_header(ws6, 4, 8, '7B2D8E')

competitors = [
    ["Competitor 1", "", "", "", "", "", "", ""],
    ["Competitor 2", "", "", "", "", "", "", ""],
    ["Competitor 3", "", "", "", "", "", "", ""],
    ["Competitor 4", "", "", "", "", "", "", ""],
    ["Competitor 5", "", "", "", "", "", "", ""],
]
for i, row_data in enumerate(competitors):
    row = 5 + i
    for col, val in enumerate(row_data, 1):
        ws6.cell(row=row, column=col, value=val)
add_rows(ws6, 5, 5, 8)
widths(ws6, [18, 25, 14, 30, 25, 16, 25, 25])

# Competitor research checklist
ws6.cell(row=12, column=1, value="🔍 竞品分析步骤 Competitor Research Steps").font = big_bold
steps = [
    "1. Search your top 10 keywords → note who ranks on page 1",
    "2. Check their DA with Ahrefs/Moz/SEMrush",
    "3. Analyze their top pages (what content drives traffic?)",
    "4. Check their backlinks (where do they get links from?)",
    "5. Review their blog/content frequency & quality",
    "6. Check their Google My Business profile",
    "7. Analyze their social media presence",
    "8. Find gaps: what are they NOT covering that you can?",
    "9. Create better content for their top keywords",
    "10. Replicate their best backlink sources",
]
for i, step in enumerate(steps):
    ws6.cell(row=13 + i, column=1, value=step).font = normal_font


# ============================================================
# Sheet 7: Monthly SEO Report 月度报告
# ============================================================
ws7 = wb.create_sheet("月度报告 Monthly")
ws7.merge_cells('A1:D1')
ws7.cell(row=1, column=1, value="📊 SEO Agent — 月度SEO报告").font = title_font
ws7.row_dimensions[1].height = 35

h7 = ["Metric 指标", "Last Month 上月", "This Month 本月", "Change 变化"]
for col, h in enumerate(h7, 1):
    ws7.cell(row=3, column=col, value=h)
style_header(ws7, 3, 4, 'D4444E')

metrics = [
    ["Total Organic Sessions 总有机流量", "", "", ""],
    ["Total Keywords Ranking 排名关键词数", "", "", ""],
    ["Keywords on Page 1 首页关键词数", "", "", ""],
    ["Keywords in Top 3 前3关键词数", "", "", ""],
    ["Average Position 平均排名", "", "", ""],
    ["Total Backlinks 总外链数", "", "", ""],
    ["Domain Authority 域名权重", "", "", ""],
    ["Google My Business Views 地图浏览", "", "", ""],
    ["Google My Business Actions 地图行动", "", "", ""],
    ["Blog Posts Published 博客发布数", "", "", ""],
    ["Page Load Speed 页面速度", "", "", ""],
    ["Core Web Vitals 核心指标", "", "", ""],
    ["Bounce Rate 跳出率", "", "", ""],
    ["Average Session Duration 平均时长", "", "", ""],
    ["Contact Form Submissions 表单提交", "", "", ""],
    ["WhatsApp Clicks WhatsApp点击", "", "", ""],
]
for i, row_data in enumerate(metrics):
    row = 4 + i
    for col, val in enumerate(row_data, 1):
        ws7.cell(row=row, column=col, value=val)
add_rows(ws7, 4, len(metrics), 4)
widths(ws7, [38, 18, 18, 14])

# SEO Goals
ws7.cell(row=22, column=1, value="🎯 SEO Goals 目标").font = big_bold
goals = [
    (23, "Month 1-3: Get indexed for all target keywords", "第1-3月：所有目标关键词被收录"),
    (24, "Month 3-6: Rank on Page 2-3 for primary keywords", "第3-6月：主要关键词排名前2-3页"),
    (25, "Month 6-9: Rank on Page 1 for 5+ keywords", "第6-9月：5+关键词排名首页"),
    (26, "Month 9-12: Rank Top 3 for 3+ keywords", "第9-12月：3+关键词排名前3"),
    (27, "Ongoing: 500+ organic sessions/month by month 6", "持续：第6月有机流量500+/月"),
]
for row, en, cn in goals:
    ws7.cell(row=row, column=1, value=en).font = normal_font
    ws7.cell(row=row, column=3, value=cn).font = small_font


# ============================================================
# Sheet 8: SEO Tools Setup 工具设置
# ============================================================
ws8 = wb.create_sheet("工具设置 Tools")
ws8.merge_cells('A1:D1')
ws8.cell(row=1, column=1, value="🛠️ SEO Agent — 必备工具清单").font = title_font
ws8.row_dimensions[1].height = 35

h8 = ["Tool 工具", "Purpose 用途", "Cost 费用", "Setup Done 已设置"]
for col, h in enumerate(h8, 1):
    ws8.cell(row=3, column=col, value=h)
style_header(ws8, 3, 4, '548235')

tools = [
    ["Google Search Console", "Index monitoring, keyword data, crawl errors", "Free 免费", ""],
    ["Google Analytics 4", "Traffic analysis, user behavior, conversions", "Free 免费", ""],
    ["Google Business Profile", "Local SEO, maps visibility, reviews", "Free 免费", ""],
    ["Google PageSpeed Insights", "Site speed & Core Web Vitals", "Free 免费", ""],
    ["Google Keyword Planner", "Keyword research & volume estimates", "Free (w/ Ads account)", ""],
    ["Ahrefs / SEMrush", "Keyword tracking, backlink analysis, competitor research", "Paid $99+/mo", ""],
    ["Screaming Frog", "Technical SEO audit, crawl analysis", "Free (500 URLs)", ""],
    ["Ubersuggest", "Keyword research & SEO audit", "Free limited / $29/mo", ""],
    ["Schema Markup Validator", "Validate structured data", "Free 免费", ""],
    ["Cloudflare", "CDN, SSL, speed optimization", "Free tier", ""],
    ["Google Rich Results Test", "Test schema/structured data", "Free 免费", ""],
    ["Moz Local", "Local citation management", "Free check / Paid", ""],
]
for i, row_data in enumerate(tools):
    row = 4 + i
    for col, val in enumerate(row_data, 1):
        ws8.cell(row=row, column=col, value=val)
    ws8.cell(row=row, column=4).fill = input_fill
add_rows(ws8, 4, len(tools), 4)
widths(ws8, [28, 45, 20, 12])


# Save
wb.save("marketing/SEO_Agent.xlsx")
print(f"✅ SEO Agent → marketing/SEO_Agent.xlsx ({len(wb.worksheets)} sheets)")
for ws in wb.worksheets:
    print(f"   - {ws.title}")
print(f"   Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
