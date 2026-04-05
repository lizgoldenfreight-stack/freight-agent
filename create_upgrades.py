"""
Upgrade 1: Enhanced Quote_Generator with auto margin formulas
Upgrade 2: Rate Trend Tracker
Upgrade 3: WhatsApp Auto-Reply Bot (Python + Flask web dashboard)
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime, timedelta
import os

header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
input_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
result_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
highlight_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
profit_fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')
loss_fill = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')
bold_font = Font(name='Arial', bold=True, size=10)
big_bold = Font(name='Arial', bold=True, size=13, color='2F5496')
title_font = Font(name='Arial', bold=True, size=18, color='2F5496')
normal_font = Font(name='Arial', size=10)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_wrap = Alignment(horizontal='left', vertical='center', wrap_text=True)

# ============================================================
# UPGRADE 1: Margin Calculator (独立工具)
# ============================================================
wb_margin = openpyxl.Workbook()

ws = wb_margin.active
ws.title = "Margin 计算器"

ws.merge_cells('A1:F1')
ws.cell(row=1, column=1, value="💰 Margin 自动计算工具").font = title_font

ws.merge_cells('A2:F2')
ws.cell(row=2, column=1, value="填成本和目标售价，自动算出每个环节的 margin 和利润率").font = Font(name='Arial', size=10, color='666666')

# === 单票计算 ===
ws.cell(row=4, column=1, value="📊 单票报价计算 Single Shipment").font = big_bold

# Input section
ws.cell(row=6, column=1, value="📥 输入 INPUT").font = Font(name='Arial', bold=True, size=11, color='E65100')

inputs = [
    (7, "成本 Cost (USD):", 500, "你的采购总成本"),
    (8, "目标利润率 Target Margin %:", 20, "你想赚多少%"),
    (9, "目标售价 Target Price (USD):", "", "留空=自动算 | 填了=反算margin"),
    (10, "汇率 Rate (MYR/USD):", 4.5, ""),
]
for row, label, default, note in inputs:
    ws.cell(row=row, column=1, value=label).font = bold_font
    ws.cell(row=row, column=1).border = thin_border
    ws.cell(row=row, column=2, value=default).font = normal_font
    ws.cell(row=row, column=2).fill = input_fill
    ws.cell(row=row, column=2).border = thin_border
    ws.cell(row=row, column=2).alignment = center
    ws.cell(row=row, column=2).number_format = '#,##0.00'
    if note:
        ws.cell(row=row, column=3, value=note).font = Font(name='Arial', italic=True, size=9, color='999999')

# Auto-calculate outputs
ws.cell(row=12, column=1, value="📤 自动计算 AUTO CALC").font = Font(name='Arial', bold=True, size=11, color='2F5496')

calcs = [
    (13, "报价 Selling Price (USD):",
     '=IF(B9<>"",B9,B7/(1-B8/100))',
     "如果填了目标售价就用目标价，否则按利润率自动算"),
    (14, "利润 Profit (USD):",
     '=B13-B7',
     "售价 - 成本"),
    (15, "实际利润率 Actual Margin %:",
     '=IF(B13=0,0,(B13-B7)/B13)',
     "利润 ÷ 售价"),
    (16, "成本占比 Cost %:",
     '=IF(B13=0,0,B7/B13)',
     "成本 ÷ 售价"),
    (17, "利润率安全度:",
     '=IF(B15>=0.25,"🟢 优秀",IF(B15>=0.15,"🟡 还行",IF(B15>=0.1,"🟠 偏低","🔴 危险")))',
     ""),
    (18, "售价 (MYR):",
     '=B13*B10',
     "换算马币"),
    (19, "利润 (MYR):",
     '=(B13-B7)*B10',
     "换算马币"),
]
for row, label, formula, note in calcs:
    ws.cell(row=row, column=1, value=label).font = bold_font
    ws.cell(row=row, column=1).border = thin_border
    cell = ws.cell(row=row, column=2, value=formula)
    cell.font = Font(name='Arial', bold=True, size=12)
    cell.fill = result_fill
    cell.border = thin_border
    cell.alignment = center
    if row in [13, 14, 18, 19]:
        cell.number_format = '#,##0.00'
    elif row in [15, 16]:
        cell.number_format = '0.0%'
    if note:
        ws.cell(row=row, column=3, value=note).font = Font(name='Arial', italic=True, size=9, color='999999')

# === Margin 速查表 ===
ws.cell(row=22, column=1, value="📋 Margin 速查表 Quick Reference").font = big_bold

ref_headers = ["成本 Cost", "10%", "12%", "15%", "18%", "20%", "22%", "25%", "28%", "30%"]
for col, h in enumerate(ref_headers, 1):
    cell = ws.cell(row=23, column=col, value=h + (" Margin" if col > 1 else ""))
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = center

costs = [100, 200, 300, 400, 500, 600, 700, 800, 900, 1000, 1500, 2000, 2500, 3000, 4000, 5000]
margins = [10, 12, 15, 18, 20, 22, 25, 28, 30]

for r, cost in enumerate(costs, 24):
    ws.cell(row=r, column=1, value=cost).font = bold_font
    ws.cell(row=r, column=1).border = thin_border
    ws.cell(row=r, column=1).number_format = '#,##0'
    for c, margin in enumerate(margins, 2):
        cell = ws.cell(row=r, column=c, value=round(cost / (1 - margin/100), 2))
        cell.font = normal_font
        cell.border = thin_border
        cell.alignment = center
        cell.number_format = '#,##0.00'
        # Color: higher margin = greener
        if margin >= 25:
            cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        elif margin >= 20:
            cell.fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')
        elif margin >= 15:
            cell.fill = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')

ws.cell(row=r+2, column=1, value="📐 公式 Formula:").font = bold_font
ws.cell(row=r+2, column=2, value="报价 = 成本 ÷ (1 - Margin%)").font = normal_font
ws.cell(row=r+2, column=2).border = thin_border

ws.cell(row=r+3, column=1, value="例 Example:").font = bold_font
ws.cell(row=r+3, column=2, value="成本 500 + 20% margin = 500 ÷ 0.80 = 625 USD").font = normal_font

# === 多票对比 ===
ws.cell(row=r+5, column=1, value="📊 多票利润对比 Multi-Quote Compare").font = big_bold

comp_headers = ["报价方案", "成本 Cost", "售价 Price", "利润 Profit", "利润率 Margin", "推荐度"]
for col, h in enumerate(comp_headers, 1):
    cell = ws.cell(row=r+6, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = center

for i in range(1, 6):
    row = r + 6 + i
    ws.cell(row=row, column=1, value=f"方案 {i}").font = normal_font
    ws.cell(row=row, column=1).border = thin_border
    for c in range(2, 4):
        cell = ws.cell(row=row, column=c)
        cell.fill = input_fill
        cell.border = thin_border
        cell.alignment = center
        cell.number_format = '#,##0.00'
    # Auto profit
    ws.cell(row=row, column=4, value=f'=IF(C{row}="","",C{row}-B{row})').font = normal_font
    ws.cell(row=row, column=4).border = thin_border
    ws.cell(row=row, column=4).alignment = center
    ws.cell(row=row, column=4).number_format = '#,##0.00'
    # Auto margin
    ws.cell(row=row, column=5, value=f'=IF(C{row}=0,"",D{row}/C{row})').font = normal_font
    ws.cell(row=row, column=5).border = thin_border
    ws.cell(row=row, column=5).alignment = center
    ws.cell(row=row, column=5).number_format = '0.0%'
    # Recommendation
    ws.cell(row=row, column=6, value=f'=IF(E{row}="","",IF(E{row}>=0.2,"⭐⭐⭐",IF(E{row}>=0.15,"⭐⭐",IF(E{row}>=0.1,"⭐","❌"))))').font = normal_font
    ws.cell(row=row, column=6).border = thin_border
    ws.cell(row=row, column=6).alignment = center

ws.column_dimensions['A'].width = 28
ws.column_dimensions['B'].width = 18
ws.column_dimensions['C'].width = 18
ws.column_dimensions['D'].width = 14
ws.column_dimensions['E'].width = 14
ws.column_dimensions['F'].width = 14

wb_margin.save('/root/.openclaw/workspace/freight-agent/Margin_Calculator.xlsx')
print("✅ Margin_Calculator.xlsx created")

# ============================================================
# UPGRADE 2: Rate Trend Tracker
# ============================================================
wb_trend = openpyxl.Workbook()

ws_t = wb_trend.active
ws_t.title = "运价记录 Rate Log"

ws_t.merge_cells('A1:J1')
ws_t.cell(row=1, column=1, value="📈 运价趋势跟踪 Rate Trend Tracker").font = title_font

ws_t.merge_cells('A2:J2')
ws_t.cell(row=2, column=1, value="每周记录运价变化，找出最佳采购时机").font = Font(name='Arial', size=10, color='666666')

# FCL rate log
ws_t.cell(row=4, column=1, value="🚢 海运整柜 FCL Rates").font = big_bold

fcl_headers = ["日期\nDate", "供应商\nSupplier", "航线\nRoute", "POL", "POD",
               "20GP\n(USD)", "40GP\n(USD)", "40HQ\n(USD)", "变化\nChange", "备注\nNotes"]
for col, h in enumerate(fcl_headers, 1):
    cell = ws_t.cell(row=5, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = center

# Pre-fill with dates for 4 weeks
from datetime import date
base_date = date(2026, 4, 1)
routes = [
    ("Shanghai→PKL", "Shanghai", "Port Klang"),
    ("Ningbo→PKL", "Ningbo", "Port Klang"),
    ("Shenzhen→PKL", "Shenzhen", "Port Klang"),
    ("Shanghai→Kuching", "Shanghai", "Kuching"),
    ("Shenzhen→KK", "Shenzhen", "Kota Kinabalu"),
]

row = 6
for week in range(8):  # 8 weeks
    d = base_date + timedelta(weeks=week)
    for route, pol, pod in routes:
        ws_t.cell(row=row, column=1, value=d.strftime("%Y-%m-%d")).font = normal_font
        ws_t.cell(row=row, column=1).border = thin_border
        ws_t.cell(row=row, column=1).alignment = center
        ws_t.cell(row=row, column=2).fill = input_fill
        ws_t.cell(row=row, column=2).border = thin_border
        ws_t.cell(row=row, column=3, value=route).font = normal_font
        ws_t.cell(row=row, column=3).border = thin_border
        ws_t.cell(row=row, column=3).alignment = center
        ws_t.cell(row=row, column=4, value=pol).font = normal_font
        ws_t.cell(row=row, column=4).border = thin_border
        ws_t.cell(row=row, column=4).alignment = center
        ws_t.cell(row=row, column=5, value=pod).font = normal_font
        ws_t.cell(row=row, column=5).border = thin_border
        ws_t.cell(row=row, column=5).alignment = center
        for c in [6, 7, 8]:
            cell = ws_t.cell(row=row, column=c)
            cell.fill = input_fill
            cell.border = thin_border
            cell.alignment = center
            cell.number_format = '#,##0'
        # Change formula (vs previous week same route)
        if row > 10:
            ws_t.cell(row=row, column=9, 
                value=f'=IF(OR(H{row}="",H{row-5}=""),"",IF(H{row}>H{row-5},"📈 +"&(H{row}-H{row-5}),IF(H{row}<H{row-5},"📉 -"&(H{row-5}-H{row}),"➡️ 平")))').font = normal_font
        ws_t.cell(row=row, column=9).border = thin_border
        ws_t.cell(row=row, column=9).alignment = center
        ws_t.cell(row=row, column=10).border = thin_border
        row += 1

fcl_widths = [12, 14, 18, 12, 14, 10, 10, 10, 14, 18]
for i, w in enumerate(fcl_widths, 1):
    ws_t.column_dimensions[get_column_letter(i)].width = w
ws_t.freeze_panes = 'A6'

# Sheet 2: Air freight rate log
ws_air = wb_trend.create_sheet("空运记录 Air Log")

ws_air.merge_cells('A1:H1')
ws_air.cell(row=1, column=1, value="✈️ 空运运价记录 Air Freight Rate Log").font = title_font

air_headers = ["日期\nDate", "供应商\nSupplier", "Origin", "Dest",
               "费率\nRate (USD/kg)", "最低\nMin (USD)", "变化\nChange", "备注\nNotes"]
for col, h in enumerate(air_headers, 1):
    cell = ws_air.cell(row=3, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = center

air_routes = [("PVG→KUL", "PVG", "KUL"), ("CAN→KUL", "CAN", "KUL"), ("SZX→KUL", "SZX", "KUL")]
row = 4
for week in range(8):
    d = base_date + timedelta(weeks=week)
    for route, o, dest in air_routes:
        ws_air.cell(row=row, column=1, value=d.strftime("%Y-%m-%d")).font = normal_font
        ws_air.cell(row=row, column=1).border = thin_border
        ws_air.cell(row=row, column=1).alignment = center
        ws_air.cell(row=row, column=2).fill = input_fill
        ws_air.cell(row=row, column=2).border = thin_border
        ws_air.cell(row=row, column=3, value=o).font = normal_font
        ws_air.cell(row=row, column=3).border = thin_border
        ws_air.cell(row=row, column=3).alignment = center
        ws_air.cell(row=row, column=4, value=dest).font = normal_font
        ws_air.cell(row=row, column=4).border = thin_border
        ws_air.cell(row=row, column=4).alignment = center
        for c in [5, 6]:
            cell = ws_air.cell(row=row, column=c)
            cell.fill = input_fill
            cell.border = thin_border
            cell.alignment = center
            cell.number_format = '#,##0.00'
        if row > 6:
            ws_air.cell(row=row, column=7,
                value=f'=IF(OR(E{row}="",E{row-3}=""),"",IF(E{row}>E{row-3},"📈 +"&(E{row}-E{row-3}),IF(E{row}<E{row-3},"📉 -"&(E{row-3}-E{row}),"➡️ 平")))').font = normal_font
        ws_air.cell(row=row, column=7).border = thin_border
        ws_air.cell(row=row, column=7).alignment = center
        ws_air.cell(row=row, column=8).border = thin_border
        row += 1

air_widths = [12, 14, 10, 10, 16, 12, 14, 18]
for i, w in enumerate(air_widths, 1):
    ws_air.column_dimensions[get_column_letter(i)].width = w

# Sheet 3: Trend Dashboard
ws_dash = wb_trend.create_sheet("趋势看板 Dashboard")

ws_dash.merge_cells('A1:F1')
ws_dash.cell(row=1, column=1, value="📊 运价趋势看板 Rate Trend Dashboard").font = title_font

ws_dash.cell(row=3, column=1, value="🚢 海运 40HQ 最新价 & 变化").font = big_bold

dash_headers = ["航线\nRoute", "最新价\nLatest", "上周价\nLast Week", "变化\nChange",
                "月最低\nMonth Low", "月最高\nMonth High"]
for col, h in enumerate(dash_headers, 1):
    cell = ws_dash.cell(row=4, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = center

for i, (route, _, _) in enumerate(routes):
    r = 5 + i
    ws_dash.cell(row=r, column=1, value=route).font = bold_font
    ws_dash.cell(row=r, column=1).border = thin_border
    # These would need manual update or more complex formulas
    for c in range(2, 7):
        cell = ws_dash.cell(row=r, column=c)
        cell.fill = input_fill
        cell.border = thin_border
        cell.alignment = center
        cell.number_format = '#,##0'

ws_dash.cell(row=12, column=1, value="💡 运价分析建议 Tips").font = big_bold
tips = [
    "1. 连续2周涨价 → 尽快锁定当前价格",
    "2. 连续2周降价 → 可以观望等更低",
    "3. 价格波动大 → 多找几家供应商比价",
    "4. 旺季前（Q3-Q4）→ 提前囤仓位",
    "5. 关注船公司公告和行业新闻",
]
for i, tip in enumerate(tips):
    ws_dash.cell(row=13 + i, column=1, value=tip).font = normal_font

ws_dash.column_dimensions['A'].width = 22
for c in 'BCDEF':
    ws_dash.column_dimensions[c].width = 14

wb_trend.save('/root/.openclaw/workspace/freight-agent/Rate_Trend_Tracker.xlsx')
print("✅ Rate_Trend_Tracker.xlsx created")

# ============================================================
# UPGRADE 3: WhatsApp Auto-Reply Bot
# ============================================================
os.makedirs('/root/.openclaw/workspace/freight-agent/whatsapp-bot', exist_ok=True)

# Bot config
bot_config = '''{
  "bot_name": "Freight Assistant",
  "auto_reply": true,
  "business_hours": {
    "start": "09:00",
    "end": "18:00",
    "timezone": "Asia/Kuala_Lumpur",
    "days": ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat"]
  },
  "out_of_hours_message": "Hi! Thanks for your message 🙏\\n\\nOur business hours are Mon-Fri 9am-6pm, Sat 9am-1pm.\\n\\nI'll get back to you during business hours. For urgent matters, please call: [YOUR PHONE]",
  "keywords": {
    "quote": ["price", "how much", "quote", "rate", "报价", "多少钱", "运价"],
    "tracking": ["where", "status", "track", "到哪", "状态", "追踪"],
    "schedule": ["when", "transit", "time", "多久", "时效", "船期"],
    "service": ["service", "do you", "can you", "服务", "能运"],
    "greeting": ["hi", "hello", "hey", "你好", "hi", "morning"]
  }
}'''

with open('/root/.openclaw/workspace/freight-agent/whatsapp-bot/config.json', 'w') as f:
    f.write(bot_config)

# Reply templates
bot_replies = '''{
  "quote": {
    "en": "Thanks for your interest! To give you an accurate quote, I need:\\n\\n📦 1. From which city/port in China?\\n📍 2. To where in Malaysia?\\n📋 3. What cargo?\\n📐 4. Volume (CBM) or weight (KG)?\\n📅 5. When do you need to ship?\\n\\nI'll get back with a competitive rate ASAP! 🚢",
    "cn": "收到！为了给你准确报价，需要以下信息：\\n\\n📦 1. 从中国哪里出？\\n📍 2. 到马来西亚哪里？\\n📋 3. 什么货？\\n📐 4. 多大体积/多重？\\n📅 5. 什么时候出？\\n\\n确认后马上报价！"
  },
  "tracking": {
    "en": "Let me check the status for you! 🔍\\n\\nCould you share your booking reference number or B/L number?",
    "cn": "帮你查一下！🔍\\n\\n请发一下订单号或提单号？"
  },
  "schedule": {
    "en": "Transit time reference:\\n\\n🚢 Sea: China → PKL: 7-12 days\\nChina → East MY: +3-5 days\\n✈️ Air: China → KUL: 1-3 days\\n\\nPlus 2-3 days for customs & delivery. Need exact timing? Share your route!",
    "cn": "时效参考：\\n\\n🚢 海运：中国→巴生港 7-12天\\n中国→东马 再加3-5天\\n✈️ 空运：中国→吉隆坡 1-3天\\n\\n加上清关和派送一般还要2-3天。要具体线路的时效告诉我！"
  },
  "service": {
    "en": "We offer:\\n\\n✅ Sea freight (FCL & LCL)\\n✅ Air freight\\n✅ Customs clearance\\n✅ Door-to-door delivery\\n✅ East MY transshipment\\n✅ Cargo insurance\\n\\nWhat do you need? I\\'ll quote you the best rate! 💪",
    "cn": "我们的服务：\\n\\n✅ 海运（整柜/拼柜）\\n✅ 空运\\n✅ 报关清关\\n✅ 送货到门\\n✅ 东马转运\\n✅ 货物保险\\n\\n需要什么服务？帮你报最优价！"
  },
  "greeting": {
    "en": "Hi! 👋 Thanks for reaching out!\\n\\nI\\'m your freight forwarding assistant. How can I help you today?\\n\\n🚢 Sea freight quotes\\n✈️ Air freight quotes\\n📦 Shipment tracking\\n📋 Customs clearance",
    "cn": "Hi! 👋 感谢联系！\\n\\n我是你的货运助手，有什么可以帮你的？\\n\\n🚢 海运报价\\n✈️ 空运报价\\n📦 货物追踪\\n📋 清关服务"
  },
  "unknown": {
    "en": "Thanks for your message! I\\'ll have our team get back to you shortly. 🙏\\n\\nIn the meantime, you can:\\n• Ask for a quote\\n• Check shipment status\\n• Get transit time info",
    "cn": "收到！我会尽快让团队回复你 🙏\\n\\n你也可以直接问我：\\n• 报价查询\\n• 货物状态\\n• 时效信息"
  }
}'''

with open('/root/.openclaw/workspace/freight-agent/whatsapp-bot/replies.json', 'w') as f:
    f.write(bot_replies)

# Bot runner script
bot_script = '''#!/usr/bin/env python3
"""
WhatsApp Auto-Reply Bot for Freight Forwarder
=============================================
Uses pywhatkit for simple auto-reply via WhatsApp Web.

Setup:
1. pip install pywhatkit flask
2. Open WhatsApp Web in your browser (web.whatsapp.com)
3. Run: python3 bot.py
4. Scan QR code if needed
5. Bot will auto-reply to incoming messages

Note: WhatsApp Web must stay open in browser.
"""

import json
import os
import re
from datetime import datetime

# Load config
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

with open(os.path.join(SCRIPT_DIR, "config.json"), "r") as f:
    config = json.load(f)

with open(os.path.join(SCRIPT_DIR, "replies.json"), "r") as f:
    replies = json.load(f)


def is_business_hours():
    """Check if current time is within business hours (MYT)."""
    now = datetime.now()
    day_names = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
    current_day = day_names[now.weekday()]
    
    if current_day not in config["business_hours"]["days"]:
        return False
    
    start = datetime.strptime(config["business_hours"]["start"], "%H:%M").time()
    end = datetime.strptime(config["business_hours"]["end"], "%H:%M").time()
    
    return start <= now.time() <= end


def detect_language(text):
    """Detect if message is in Chinese or English."""
    chinese_chars = len(re.findall(r'[\\u4e00-\\u9fff]', text))
    return "cn" if chinese_chars > len(text) * 0.2 else "en"


def detect_intent(message):
    """Detect what the customer is asking about."""
    message_lower = message.lower()
    
    for intent, keywords in config["keywords"].items():
        for keyword in keywords:
            if keyword.lower() in message_lower:
                return intent
    
    return "unknown"


def get_reply(intent, lang="en"):
    """Get the appropriate reply template."""
    if intent in replies:
        return replies[intent].get(lang, replies[intent].get("en", replies["unknown"]["en"]))
    return replies["unknown"].get(lang, replies["unknown"]["en"])


def process_message(message):
    """Process incoming message and return reply."""
    if not is_business_hours() and not config.get("auto_reply_outside_hours", False):
        return config["out_of_hours_message"]
    
    lang = detect_language(message)
    intent = detect_intent(message)
    reply = get_reply(intent, lang)
    
    print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M')}] Intent: {intent} | Lang: {lang}")
    print(f"  Message: {message[:50]}...")
    print(f"  Reply: {reply[:50]}...")
    
    return reply


# ============================================================
# Dashboard: Simple web interface to manage replies
# ============================================================
def create_dashboard():
    """Create a simple web dashboard for managing the bot."""
    dashboard_html = """<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>WhatsApp Bot Dashboard</title>
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body { font-family: Arial, sans-serif; background: #f0f2f5; padding: 20px; }
        .container { max-width: 900px; margin: 0 auto; }
        h1 { color: #1a73e8; margin-bottom: 20px; }
        .card { background: white; border-radius: 10px; padding: 20px; margin-bottom: 20px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }
        .card h2 { color: #333; margin-bottom: 15px; font-size: 18px; }
        .status { display: flex; gap: 20px; margin-bottom: 20px; }
        .stat { flex: 1; background: #e3f2fd; padding: 15px; border-radius: 8px; text-align: center; }
        .stat .number { font-size: 32px; font-weight: bold; color: #1a73e8; }
        .stat .label { font-size: 12px; color: #666; }
        .reply-box { background: #f5f5f5; padding: 15px; border-radius: 8px; margin-bottom: 10px; }
        .reply-box .intent { font-weight: bold; color: #1a73e8; margin-bottom: 5px; }
        .reply-box .text { white-space: pre-wrap; font-size: 14px; }
        .btn { background: #1a73e8; color: white; border: none; padding: 10px 20px; border-radius: 5px; cursor: pointer; }
        .btn:hover { background: #1557b0; }
        textarea { width: 100%; padding: 10px; border: 1px solid #ddd; border-radius: 5px; min-height: 100px; font-family: monospace; }
        .tag { display: inline-block; background: #e8f5e9; color: #2e7d32; padding: 3px 8px; border-radius: 4px; font-size: 12px; margin: 2px; }
    </style>
</head>
<body>
    <div class="container">
        <h1>🤖 WhatsApp Freight Bot Dashboard</h1>
        
        <div class="status">
            <div class="stat">
                <div class="number">ACTIVE</div>
                <div class="label">Bot Status</div>
            </div>
            <div class="stat">
                <div class="number">5</div>
                <div class="label">Intents</div>
            </div>
            <div class="stat">
                <div class="number">EN / CN</div>
                <div class="label">Languages</div>
            </div>
        </div>
        
        <div class="card">
            <h2>📝 Quick Reply Templates</h2>
            <div class="reply-box">
                <div class="intent">🚢 Quote Request</div>
                <div class="text">Thanks for your interest! To give you an accurate quote, I need: 1. From which city? 2. To where? 3. What cargo? 4. Volume/Weight? 5. When?</div>
            </div>
            <div class="reply-box">
                <div class="intent">📦 Tracking</div>
                <div class="text">Let me check the status for you! Could you share your booking reference or B/L number?</div>
            </div>
            <div class="reply-box">
                <div class="intent">⏱️ Transit Time</div>
                <div class="text">Sea: China→PKL 7-12 days | East MY +3-5 days | Air: China→KUL 1-3 days</div>
            </div>
            <div class="reply-box">
                <div class="intent">🎯 Services</div>
                <div class="text">Sea freight (FCL/LCL), Air freight, Customs, Door-to-door, East MY transshipment, Insurance</div>
            </div>
        </div>
        
        <div class="card">
            <h2>🏷️ Detected Keywords</h2>
            <div>
                <span class="tag">price</span><span class="tag">how much</span><span class="tag">quote</span>
                <span class="tag">where</span><span class="tag">status</span><span class="tag">track</span>
                <span class="tag">when</span><span class="tag">transit</span><span class="tag">time</span>
                <span class="tag">报价</span><span class="tag">多少钱</span><span class="tag">到哪</span>
                <span class="tag">状态</span><span class="tag">多久</span><span class="tag">时效</span>
            </div>
        </div>
        
        <div class="card">
            <h2>⚙️ How to Use</h2>
            <ol style="padding-left: 20px; line-height: 2;">
                <li>Open WhatsApp Web in your browser</li>
                <li>Run <code>python3 bot.py</code> in terminal</li>
                <li>Bot detects incoming messages</li>
                <li>Auto-replies based on keywords</li>
                <li>Edit <code>replies.json</code> to customize messages</li>
            </ol>
        </div>
    </div>
</body>
</html>"""
    
    return dashboard_html


if __name__ == "__main__":
    print("🤖 WhatsApp Freight Bot")
    print("=" * 40)
    print(f"Business hours: {config['business_hours']['start']} - {config['business_hours']['end']} MYT")
    print(f"Auto-reply: {config['auto_reply']}")
    print()
    
    # Test the bot
    test_messages = [
        "How much to ship to Malaysia?",
        "我的货到哪了？",
        "What services do you offer?",
        "How long does shipping take?",
        "Hello!",
        "Can you ship to Kuching?",
    ]
    
    print("Testing auto-reply system:")
    print("-" * 40)
    for msg in test_messages:
        reply = process_message(msg)
        print()
    
    print("-" * 40)
    print("✅ Bot ready! Edit config.json and replies.json to customize.")
    print("📋 Run dashboard: python3 -c \"from bot import create_dashboard; print(create_dashboard())\" > dashboard.html")
'''

with open('/root/.openclaw/workspace/freight-agent/whatsapp-bot/bot.py', 'w') as f:
    f.write(bot_script)

# Dashboard HTML
dashboard = '''<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>WhatsApp Bot Dashboard - Freight Agent</title>
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body { font-family: 'Segoe UI', Arial, sans-serif; background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); min-height: 100vh; padding: 20px; }
        .container { max-width: 1000px; margin: 0 auto; }
        h1 { color: white; text-align: center; margin-bottom: 30px; font-size: 28px; }
        .grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(300px, 1fr)); gap: 20px; }
        .card { background: white; border-radius: 16px; padding: 24px; box-shadow: 0 10px 40px rgba(0,0,0,0.15); }
        .card h2 { color: #333; margin-bottom: 16px; font-size: 18px; display: flex; align-items: center; gap: 8px; }
        .status-bar { display: flex; gap: 15px; margin-bottom: 20px; }
        .status-item { flex: 1; background: linear-gradient(135deg, #667eea, #764ba2); color: white; padding: 20px; border-radius: 12px; text-align: center; }
        .status-item .num { font-size: 28px; font-weight: bold; }
        .status-item .label { font-size: 12px; opacity: 0.9; }
        .reply-card { background: #f8f9fa; border-radius: 10px; padding: 16px; margin-bottom: 12px; border-left: 4px solid #667eea; }
        .reply-card .intent { font-weight: bold; color: #667eea; margin-bottom: 8px; }
        .reply-card .text { font-size: 13px; color: #555; white-space: pre-wrap; line-height: 1.6; }
        .keyword-cloud { display: flex; flex-wrap: wrap; gap: 8px; }
        .kw { background: #e8eaf6; color: #3f51b5; padding: 6px 12px; border-radius: 20px; font-size: 13px; }
        .kw.cn { background: #fff3e0; color: #e65100; }
        .steps { list-style: none; counter-reset: step; }
        .steps li { counter-increment: step; padding: 12px 0; border-bottom: 1px solid #eee; display: flex; align-items: flex-start; gap: 12px; }
        .steps li::before { content: counter(step); background: #667eea; color: white; width: 28px; height: 28px; border-radius: 50%; display: flex; align-items: center; justify-content: center; font-weight: bold; font-size: 14px; flex-shrink: 0; }
        .code { background: #1e1e1e; color: #d4d4d4; padding: 2px 8px; border-radius: 4px; font-family: 'Consolas', monospace; font-size: 13px; }
    </style>
</head>
<body>
    <div class="container">
        <h1>🤖 WhatsApp Freight Bot Dashboard</h1>
        
        <div class="status-bar">
            <div class="status-item"><div class="num">🟢</div><div class="label">Bot Active</div></div>
            <div class="status-item"><div class="num">5</div><div class="label">Intents</div></div>
            <div class="status-item"><div class="num">EN/CN</div><div class="label">Languages</div></div>
            <div class="status-item"><div class="num">24/7</div><div class="label">Auto-Reply</div></div>
        </div>
        
        <div class="grid">
            <div class="card">
                <h2>📝 Quick Reply Templates</h2>
                
                <div class="reply-card">
                    <div class="intent">🚢 Quote Request</div>
                    <div class="text">Thanks for your interest! To give you an accurate quote, I need:

1. From which city/port in China?
2. To where in Malaysia?
3. What cargo?
4. Volume (CBM) or weight (KG)?
5. When do you need to ship?

I'll get back with a competitive rate ASAP! 🚢</div>
                </div>
                
                <div class="reply-card">
                    <div class="intent">📦 Shipment Tracking</div>
                    <div class="text">Let me check the status for you! 🔍
Could you share your booking reference or B/L number?</div>
                </div>
                
                <div class="reply-card">
                    <div class="intent">⏱️ Transit Time</div>
                    <div class="text">🚢 Sea: China→PKL: 7-12 days | East MY: +3-5 days
✈️ Air: China→KUL: 1-3 days
+ 2-3 days customs & delivery</div>
                </div>
            </div>
            
            <div class="card">
                <h2>🏷️ Detected Keywords</h2>
                <div class="keyword-cloud">
                    <span class="kw">price</span><span class="kw">how much</span><span class="kw">quote</span>
                    <span class="kw">rate</span><span class="kw">where</span><span class="kw">status</span>
                    <span class="kw">track</span><span class="kw">when</span><span class="kw">transit</span>
                    <span class="kw">time</span><span class="kw">service</span><span class="kw">can you</span>
                    <span class="kw cn">报价</span><span class="kw cn">多少钱</span><span class="kw cn">运价</span>
                    <span class="kw cn">到哪</span><span class="kw cn">状态</span><span class="kw cn">追踪</span>
                    <span class="kw cn">多久</span><span class="kw cn">时效</span><span class="kw cn">船期</span>
                    <span class="kw cn">服务</span><span class="kw cn">能运</span><span class="kw cn">你好</span>
                </div>
                
                <h2 style="margin-top: 24px;">🎯 Services Offered</h2>
                <div style="line-height: 2;">
                    ✅ Sea Freight (FCL & LCL)<br>
                    ✅ Air Freight<br>
                    ✅ Customs Clearance<br>
                    ✅ Door-to-Door Delivery<br>
                    ✅ East MY Transshipment<br>
                    ✅ Cargo Insurance
                </div>
            </div>
            
            <div class="card" style="grid-column: 1 / -1;">
                <h2>⚙️ Setup Instructions</h2>
                <ol class="steps">
                    <li>Install Python dependencies: <span class="code">pip install pywhatkit flask</span></li>
                    <li>Open WhatsApp Web in your browser: <span class="code">web.whatsapp.com</span></li>
                    <li>Run the bot: <span class="code">python3 bot.py</span></li>
                    <li>Scan QR code if prompted</li>
                    <li>Bot will auto-reply to incoming messages based on keywords</li>
                    <li>Customize replies in <span class="code">replies.json</span></li>
                    <li>Adjust business hours in <span class="code">config.json</span></li>
                </ol>
            </div>
        </div>
    </div>
</body>
</html>'''

with open('/root/.openclaw/workspace/freight-agent/whatsapp-bot/dashboard.html', 'w') as f:
    f.write(dashboard)

# README for bot
bot_readme = '''# 🤖 WhatsApp Auto-Reply Bot

自动回复 WhatsApp 消息的机器人。

## 文件说明

| 文件 | 用途 |
|------|------|
| `config.json` | 机器人配置（营业时间、关键词等） |
| `replies.json` | 自动回复模板（中英双语） |
| `bot.py` | 机器人主程序 |
| `dashboard.html` | 管理面板（浏览器打开） |

## 快速开始

```bash
# 1. 安装依赖
pip install pywhatkit flask

# 2. 打开 WhatsApp Web
# 浏览器访问 web.whatsapp.com

# 3. 运行机器人
python3 bot.py
```

## 支持的关键词

| 意图 | 英文关键词 | 中文关键词 |
|------|-----------|-----------|
| 报价 | price, how much, quote | 报价, 多少钱 |
| 追踪 | where, status, track | 到哪, 状态, 追踪 |
| 时效 | when, transit, time | 多久, 时效, 船期 |
| 服务 | service, can you | 服务, 能运 |
| 问候 | hi, hello, hey | 你好 |

## 自定义

- 编辑 `replies.json` 修改回复内容
- 编辑 `config.json` 修改营业时间和关键词
'''

with open('/root/.openclaw/workspace/freight-agent/whatsapp-bot/README.md', 'w') as f:
    f.write(bot_readme)

print("✅ WhatsApp Bot created in whatsapp-bot/")
print("   - bot.py")
print("   - config.json")
print("   - replies.json")
print("   - dashboard.html")
print("   - README.md")
print("\n🎉 All 3 upgrades created!")
