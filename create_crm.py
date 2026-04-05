"""
Freight Forwarder CRM - Customer Management
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime

wb = openpyxl.Workbook()

# Styles
header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
blue_fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
normal_font = Font(name='Arial', size=10)
bold_font = Font(name='Arial', bold=True, size=10)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

def style_header(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

def style_data(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = normal_font
        cell.alignment = center_align
        cell.border = thin_border

# ============================================================
# Sheet 1: 客户总览 Customer Overview
# ============================================================
ws_cust = wb.active
ws_cust.title = "客户总览"

headers = [
    "No.", "公司名\nCompany", "联系人\nContact", 
    "电话\nPhone", "微信\nWeChat", "邮箱\nEmail",
    "客户来源\nSource", "主要需求\nMain Need",
    "客户状态\nStatus", "上次联系\nLast Contact",
    "下次跟进\nNext Follow-up", "备注\nNotes"
]
for col, h in enumerate(headers, 1):
    ws_cust.cell(row=1, column=col, value=h)
style_header(ws_cust, 1, len(headers))

# Sample data
samples = [
    [1, "ABC Trading Sdn Bhd", "Mr. Tan", "+60123456789", "tan_abc", "tan@abc.com", 
     "LinkedIn", "Sea FCL Shanghai→PKL", "🟢 活跃", "2026-04-01", "2026-04-08", "每月2-3柜"],
    [2, "XYZ Electronics", "Ms. Lim", "+60198765432", "lim_xyz", "lim@xyz.com",
     "WhatsApp", "Air PVG→KUL", "🟡 跟进中", "2026-03-28", "2026-04-05", "报价已发，等回复"],
    [3, "DEF Manufacturing", "Mr. Wong", "+60112233445", "wong_def", "", 
     "转介绍", "Sea LCL Shenzhen→PKL", "🔴 未跟进", "2026-03-15", "2026-04-05", "上次报价偏高"],
    [4, "", "", "", "", "", "", "", "", "", "", ""],
    [5, "", "", "", "", "", "", "", "", "", "", ""],
]
for r, data in enumerate(samples, 2):
    for c, val in enumerate(data, 1):
        ws_cust.cell(row=r, column=c, value=val)
    style_data(ws_cust, r, len(headers))

# Color the status column
for r in range(2, 2 + len(samples)):
    status_cell = ws_cust.cell(row=r, column=9)
    val = str(status_cell.value)
    if "活跃" in val:
        status_cell.fill = green_fill
    elif "跟进" in val:
        status_cell.fill = yellow_fill
    elif "未跟进" in val:
        status_cell.fill = red_fill

# Data validation for status
dv_status = DataValidation(
    type="list",
    formula1='"🟢 活跃,🟡 跟进中,🔴 未跟进,⚫ 已成交,⚪ 已流失"',
    allow_blank=True
)
dv_status.error = "请从下拉列表选择"
dv_status.errorTitle = "无效状态"
ws_cust.add_data_validation(dv_status)
dv_status.add(f'I2:I100')

# Data validation for source
dv_source = DataValidation(
    type="list",
    formula1='"WhatsApp,微信,Email,LinkedIn,转介绍,展会,Google,其他"',
    allow_blank=True
)
ws_cust.add_data_validation(dv_source)
dv_source.add(f'G2:G100')

# Data validation for main need
dv_need = DataValidation(
    type="list",
    formula1='"Sea FCL,Sea LCL,Air,Sea+Air,不确定"',
    allow_blank=True
)
ws_cust.add_data_validation(dv_need)
dv_need.add(f'H2:H100')

# Column widths
widths = [5, 22, 14, 16, 14, 22, 12, 22, 14, 14, 14, 25]
for i, w in enumerate(widths, 1):
    ws_cust.column_dimensions[get_column_letter(i)].width = w

# Freeze top row
ws_cust.freeze_panes = 'A2'

# ============================================================
# Sheet 2: 报价记录 Quote History
# ============================================================
ws_quote = wb.create_sheet("报价记录")

q_headers = [
    "No.", "日期\nDate", "公司名\nCompany", "联系人\nContact",
    "运输方式\nMode", "航线\nRoute", "柜型/货量\nVolume",
    "成本价\nCost (USD)", "报价\nQuote (USD)", "Margin\n%",
    "状态\nStatus", "报价单号\nQuote No.", "备注\nNotes"
]
for col, h in enumerate(q_headers, 1):
    ws_quote.cell(row=1, column=col, value=h)
style_header(ws_quote, 1, len(q_headers))

q_samples = [
    [1, "2026-04-01", "ABC Trading", "Mr. Tan", "Sea FCL", "Shanghai→PKL", "1x40HQ", 580, 720, "24%", "✅ 已成交", "QT-2026-001", ""],
    [2, "2026-04-02", "XYZ Electronics", "Ms. Lim", "Air", "PVG→KUL", "500kg", 1750, 2100, "20%", "⏳ 等回复", "QT-2026-002", ""],
    [3, "2026-04-03", "DEF Manufacturing", "Mr. Wong", "Sea LCL", "Shenzhen→PKL", "8 CBM", 240, 280, "17%", "❌ 未成交", "QT-2026-003", "客户觉得贵"],
    [4, "", "", "", "", "", "", "", "", "", "", "", ""],
]
for r, data in enumerate(q_samples, 2):
    for c, val in enumerate(data, 1):
        ws_quote.cell(row=r, column=c, value=val)
    style_data(ws_quote, r, len(q_headers))

# Color status
for r in range(2, 2 + len(q_samples)):
    status_cell = ws_quote.cell(row=r, column=11)
    val = str(status_cell.value)
    if "成交" in val:
        status_cell.fill = green_fill
    elif "等回复" in val:
        status_cell.fill = yellow_fill
    elif "未成交" in val:
        status_cell.fill = red_fill

# Data validation for quote status
dv_qstatus = DataValidation(
    type="list",
    formula1='"✅ 已成交,⏳ 等回复,❌ 未成交,📝 已报价,🔄 重新报价"',
    allow_blank=True
)
ws_quote.add_data_validation(dv_qstatus)
dv_qstatus.add(f'K2:K100')

# Data validation for mode
dv_mode = DataValidation(
    type="list",
    formula1='"Sea FCL,Sea LCL,Air,Sea+Air"',
    allow_blank=True
)
ws_quote.add_data_validation(dv_mode)
dv_mode.add(f'E2:E100')

q_widths = [5, 12, 18, 12, 12, 18, 14, 14, 14, 10, 14, 16, 20]
for i, w in enumerate(q_widths, 1):
    ws_quote.column_dimensions[get_column_letter(i)].width = w
ws_quote.freeze_panes = 'A2'

# ============================================================
# Sheet 3: 跟进提醒 Follow-up Reminders
# ============================================================
ws_follow = wb.create_sheet("跟进提醒")

f_headers = [
    "No.", "日期\nDate", "公司名\nCompany", "联系人\nContact",
    "联系方式\nChannel", "跟进内容\nAction", "结果\nResult",
    "下次跟进\nNext Action", "截止日期\nDeadline", "优先级\nPriority"
]
for col, h in enumerate(f_headers, 1):
    ws_follow.cell(row=1, column=col, value=h)
style_header(ws_follow, 1, len(f_headers))

f_samples = [
    [1, "2026-04-01", "ABC Trading", "Mr. Tan", "WhatsApp", "发送4月报价", "客户确认", "确认出货时间", "2026-04-08", "🟢 低"],
    [2, "2026-04-02", "XYZ Electronics", "Ms. Lim", "微信", "跟进报价反馈", "未回复", "再次跟进", "2026-04-05", "🔴 高"],
    [3, "2026-04-03", "DEF Manufacturing", "Mr. Wong", "Email", "发送降价方案", "已读未回", "电话跟进", "2026-04-06", "🟡 中"],
    [4, "", "", "", "", "", "", "", "", ""],
]
for r, data in enumerate(f_samples, 2):
    for c, val in enumerate(data, 1):
        ws_follow.cell(row=r, column=c, value=val)
    style_data(ws_follow, r, len(f_headers))

# Color priority
for r in range(2, 2 + len(f_samples)):
    p_cell = ws_follow.cell(row=r, column=10)
    val = str(p_cell.value)
    if "高" in val:
        p_cell.fill = red_fill
    elif "中" in val:
        p_cell.fill = yellow_fill
    elif "低" in val:
        p_cell.fill = green_fill

# Data validation
dv_channel = DataValidation(type="list", formula1='"WhatsApp,微信,Email,电话,LinkedIn,面谈"', allow_blank=True)
ws_follow.add_data_validation(dv_channel)
dv_channel.add(f'E2:E100')

dv_priority = DataValidation(type="list", formula1='"🔴 高,🟡 中,🟢 低"', allow_blank=True)
ws_follow.add_data_validation(dv_priority)
dv_priority.add(f'J2:J100')

f_widths = [5, 12, 18, 12, 12, 20, 14, 16, 14, 10]
for i, w in enumerate(f_widths, 1):
    ws_follow.column_dimensions[get_column_letter(i)].width = w
ws_follow.freeze_panes = 'A2'

# ============================================================
# Sheet 4: 统计看板 Dashboard
# ============================================================
ws_dash = wb.create_sheet("统计看板")

# Title
ws_dash.merge_cells('A1:F1')
ws_dash.cell(row=1, column=1, value="📊 业务统计看板 Business Dashboard")
ws_dash.cell(row=1, column=1).font = Font(name='Arial', bold=True, size=16, color='2F5496')

# Customer stats
ws_dash.cell(row=3, column=1, value="👥 客户统计").font = Font(name='Arial', bold=True, size=13)
stats = [
    ("总客户数", '=COUNTA(客户总览!B2:B100)-COUNTBLANK(客户总览!B2:B100)'),
    ("活跃客户", '=COUNTIF(客户总览!I2:I100,"*活跃*")'),
    ("跟进中", '=COUNTIF(客户总览!I2:I100,"*跟进*")'),
    ("未跟进 ⚠️", '=COUNTIF(客户总览!I2:I100,"*未跟进*")'),
    ("已成交", '=COUNTIF(客户总览!I2:I100,"*成交*")'),
]
for i, (label, formula) in enumerate(stats):
    r = 4 + i
    ws_dash.cell(row=r, column=1, value=label).font = bold_font
    ws_dash.cell(row=r, column=1).border = thin_border
    ws_dash.cell(row=r, column=2, value=formula).font = normal_font
    ws_dash.cell(row=r, column=2).border = thin_border
    ws_dash.cell(row=r, column=2).alignment = center_align

# Quote stats
ws_dash.cell(row=11, column=1, value="💰 报价统计").font = Font(name='Arial', bold=True, size=13)
q_stats = [
    ("总报价数", '=COUNTA(报价记录!B2:B100)-COUNTBLANK(报价记录!B2:B100)'),
    ("已成交", '=COUNTIF(报价记录!K2:K100,"*成交*")'),
    ("等回复", '=COUNTIF(报价记录!K2:K100,"*等回复*")'),
    ("未成交", '=COUNTIF(报价记录!K2:K100,"*未成交*")'),
    ("成交率", '=IF(B12=0,0,B13/B12)'),
]
for i, (label, formula) in enumerate(q_stats):
    r = 12 + i
    ws_dash.cell(row=r, column=1, value=label).font = bold_font
    ws_dash.cell(row=r, column=1).border = thin_border
    ws_dash.cell(row=r, column=2, value=formula).font = normal_font
    ws_dash.cell(row=r, column=2).border = thin_border
    ws_dash.cell(row=r, column=2).alignment = center_align
# Format percentage
ws_dash.cell(row=16, column=2).number_format = '0%'

# Follow-up stats
ws_dash.cell(row=19, column=1, value="📋 跟进统计").font = Font(name='Arial', bold=True, size=13)
f_stats = [
    ("待跟进总数", '=COUNTA(跟进提醒!B2:B100)-COUNTBLANK(跟进提醒!B2:B100)'),
    ("🔴 高优先级", '=COUNTIF(跟进提醒!J2:J100,"*高*")'),
    ("🟡 中优先级", '=COUNTIF(跟进提醒!J2:J100,"*中*")'),
    ("🟢 低优先级", '=COUNTIF(跟进提醒!J2:J100,"*低*")'),
]
for i, (label, formula) in enumerate(f_stats):
    r = 20 + i
    ws_dash.cell(row=r, column=1, value=label).font = bold_font
    ws_dash.cell(row=r, column=1).border = thin_border
    ws_dash.cell(row=r, column=2, value=formula).font = normal_font
    ws_dash.cell(row=r, column=2).border = thin_border
    ws_dash.cell(row=r, column=2).alignment = center_align

ws_dash.column_dimensions['A'].width = 20
ws_dash.column_dimensions['B'].width = 15

# Save
wb.save('/root/.openclaw/workspace/freight-agent/Customer_CRM.xlsx')
print("✅ Customer_CRM.xlsx created")
