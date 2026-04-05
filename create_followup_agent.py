"""
Follow-up Agent - 客户跟进管理
自动生成跟进提醒、邮件/消息模板、客户活跃度分析
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime, timedelta

header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
input_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
blue_fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
grey_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
bold_font = Font(name='Arial', bold=True, size=10)
big_bold = Font(name='Arial', bold=True, size=13, color='2F5496')
title_font = Font(name='Arial', bold=True, size=18, color='2F5496')
normal_font = Font(name='Arial', size=10)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_wrap = Alignment(horizontal='left', vertical='center', wrap_text=True)

wb = openpyxl.Workbook()

# ============================================================
# Sheet 1: 今日跟进 Today's Follow-ups
# ============================================================
ws_today = wb.active
ws_today.title = "今日跟进 Today"

ws_today.merge_cells('A1:H1')
ws_today.cell(row=1, column=1, value="📋 今日跟进任务 Today's Follow-up Tasks").font = title_font

ws_today.merge_cells('A2:H2')
ws_today.cell(row=2, column=1, value='=CONCATENATE("📅 ",TEXT(TODAY(),"yyyy-mm-dd"),"  |  ",COUNTA(A5:A30)-COUNTBLANK(A5:A30)," 个任务待处理")').font = Font(name='Arial', size=12, color='E65100')

t_headers = ["⏰", "客户\nCustomer", "联系人\nContact", "方式\nChannel",
             "上次内容\nLast Action", "今日任务\nToday's Task", "优先级\nPriority", "状态\nDone"]
for col, h in enumerate(t_headers, 1):
    cell = ws_today.cell(row=4, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = center

for r in range(5, 35):
    for c in range(1, 9):
        cell = ws_today.cell(row=r, column=c)
        cell.border = thin_border
        cell.alignment = center if c != 5 else left_wrap
        cell.font = normal_font
    ws_today.cell(row=r, column=8).fill = input_fill

dv_priority = DataValidation(type="list", formula1='"🔴 高,🟡 中,🟢 低"', allow_blank=True)
ws_today.add_data_validation(dv_priority)
dv_priority.add('G5:G34')

dv_done = DataValidation(type="list", formula1='"✅ 完成,⏳ 进行中,❌ 跳过"', allow_blank=True)
ws_today.add_data_validation(dv_done)
dv_done.add('H5:H34')

t_widths = [5, 18, 12, 10, 25, 25, 10, 10]
for i, w in enumerate(t_widths, 1):
    ws_today.column_dimensions[get_column_letter(i)].width = w

# ============================================================
# Sheet 2: 跟进记录 Follow-up Log
# ============================================================
ws_log = wb.create_sheet("跟进记录 Log")

ws_log.merge_cells('A1:J1')
ws_log.cell(row=1, column=1, value="📝 客户跟进记录 Customer Follow-up Log").font = title_font

l_headers = ["日期\nDate", "公司\nCompany", "联系人\nContact", "渠道\nChannel",
             "类型\nType", "内容摘要\nSummary", "客户反馈\nResponse",
             "下次行动\nNext Action", "跟进日期\nFollow Date", "优先级\nPriority"]
for col, h in enumerate(l_headers, 1):
    cell = ws_log.cell(row=3, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = center

for r in range(4, 54):
    for c in range(1, 11):
        cell = ws_log.cell(row=r, column=c)
        cell.border = thin_border
        cell.alignment = center if c not in [6, 7, 8] else left_wrap
        cell.font = normal_font
    ws_log.cell(row=r, column=1).fill = input_fill
    ws_log.cell(row=r, column=9).fill = input_fill

dv_channel = DataValidation(type="list", formula1='"WhatsApp,微信,Email,电话,LinkedIn,面谈"', allow_blank=True)
ws_log.add_data_validation(dv_channel)
dv_channel.add('D4:D53')

dv_type = DataValidation(type="list", formula1='"报价跟进,询价回复,日常问候,投诉处理,催款,成交后回访,重新激活"', allow_blank=True)
ws_log.add_data_validation(dv_type)
dv_type.add('E4:E53')

dv_lp = DataValidation(type="list", formula1='"🔴 高,🟡 中,🟢 低"', allow_blank=True)
ws_log.add_data_validation(dv_lp)
dv_lp.add('J4:J53')

l_widths = [12, 18, 12, 10, 14, 28, 20, 20, 12, 10]
for i, w in enumerate(l_widths, 1):
    ws_log.column_dimensions[get_column_letter(i)].width = w
ws_log.freeze_panes = 'A4'

# ============================================================
# Sheet 3: 客户分级 Customer Tiers
# ============================================================
ws_tier = wb.create_sheet("客户分级 Tiers")

ws_tier.merge_cells('A1:H1')
ws_tier.cell(row=1, column=1, value="🏆 客户分级管理 Customer Tier Management").font = title_font

ws_tier.merge_cells('A2:H2')
ws_tier.cell(row=2, column=1, value="根据客户价值和活跃度分级，不同级别用不同跟进频率").font = Font(name='Arial', size=10, color='666666')

tier_headers = ["公司\nCompany", "联系人\nContact", "级别\nTier",
                "最近下单\nLast Order", "订单次数\nOrders", "累计金额\nRevenue (USD)",
                "跟进频率\nFrequency", "下次联系\nNext Contact"]
for col, h in enumerate(tier_headers, 1):
    cell = ws_tier.cell(row=4, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = center

tier_samples = [
    ["ABC Trading", "Mr. Tan", "🥇 VIP", "2026-04-01", 8, 12000, "每周", "2026-04-08"],
    ["XYZ Electronics", "Ms. Lim", "🥈 A级", "2026-03-20", 3, 5000, "每2周", "2026-04-10"],
    ["DEF Manufacturing", "Mr. Wong", "🥉 B级", "2026-03-01", 1, 1500, "每月", "2026-04-15"],
    ["", "", "", "", "", "", "", ""],
    ["", "", "", "", "", "", "", ""],
]

for r, data in enumerate(tier_samples, 5):
    for c, val in enumerate(data, 1):
        cell = ws_tier.cell(row=r, column=c, value=val)
        cell.font = normal_font
        cell.border = thin_border
        cell.alignment = center
    # Color tier
    tier_val = str(data[2])
    tier_cell = ws_tier.cell(row=r, column=3)
    if "VIP" in tier_val:
        tier_cell.fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')
    elif "A级" in tier_val:
        tier_cell.fill = silver_fill = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')
    elif "B级" in tier_val:
        tier_cell.fill = PatternFill(start_color='CD7F32', end_color='CD7F32', fill_type='solid')

dv_tier = DataValidation(type="list", formula1='"🥇 VIP,🥈 A级,🥉 B级,⚪ 潜在,💀 流失"', allow_blank=True)
ws_tier.add_data_validation(dv_tier)
dv_tier.add('C5:C25')

dv_freq = DataValidation(type="list", formula1='"每周,每2周,每月,每季度,按需"', allow_blank=True)
ws_tier.add_data_validation(dv_freq)
dv_freq.add('G5:G25')

t_widths = [20, 12, 10, 14, 10, 16, 10, 14]
for i, w in enumerate(t_widths, 1):
    ws_tier.column_dimensions[get_column_letter(i)].width = w

# Legend
ws_tier.cell(row=12, column=1, value="📌 分级标准").font = big_bold
ws_tier.cell(row=13, column=1, value="🥇 VIP: 月出货 ≥3单 或 累计 ≥USD 10,000").font = normal_font
ws_tier.cell(row=14, column=1, value="🥈 A级: 月出货 1-2单 或 累计 ≥USD 3,000").font = normal_font
ws_tier.cell(row=15, column=1, value="🥉 B级: 有成交但量不大").font = normal_font
ws_tier.cell(row=16, column=1, value="⚪ 潜在: 报过价但未成交").font = normal_font
ws_tier.cell(row=17, column=1, value="💀 流失: 超过3个月没联系").font = normal_font

# ============================================================
# Sheet 4: 邮件/消息模板 Email & Message Templates
# ============================================================
ws_tpl = wb.create_sheet("消息模板 Templates")

ws_tpl.merge_cells('A1:C1')
ws_tpl.cell(row=1, column=1, value="✉️ 跟进消息模板 Follow-up Templates").font = title_font

tpl_headers = ["场景 Scenario", "语言 Lang", "模板 Template"]
for col, h in enumerate(tpl_headers, 1):
    cell = ws_tpl.cell(row=3, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = center

templates = [
    # 报价后跟进
    ["报价后1天\n1 day after quote", "EN",
     "Hi [Name],\n\n"
     "Just following up on the quotation I sent yesterday for [Route].\n\n"
     "Have you had a chance to review it? Happy to clarify any questions.\n\n"
     "Best regards,\n[Your Name]"],

    ["报价后1天\n1 day after quote", "CN",
     "Hi [名字]，\n\n"
     "昨天给你的 [航线] 报价有看了吗？\n"
     "有什么问题随时问我，价格有效期还有 [X] 天。\n\n"
     "[你的名字]"],

    ["报价后3天\n3 days after quote", "EN",
     "Hi [Name],\n\n"
     "Hope you're doing well! Following up on the shipment quote for [Route].\n\n"
     "Shipping rates are changing frequently these days — "
     "I wanted to make sure you can still lock in the current rate before it expires on [Date].\n\n"
     "Ready to proceed? Just let me know! 🙏"],

    ["报价后3天\n3 days after quote", "CN",
     "Hi [名字]，\n\n"
     "运价最近变动比较大，想帮你确认一下之前报的价格还能不能拿到。\n"
     "报价到 [日期] 过期，要走的话尽快告诉我 👍\n\n"
     "[你的名字]"],

    ["报价后7天（最后跟进）\n7 days (last follow-up)", "EN",
     "Hi [Name],\n\n"
     "I understand you might be busy or considering other options. No pressure at all!\n\n"
     "If your shipping needs change in the future, feel free to reach out anytime. "
     "I'm always happy to help with competitive rates.\n\n"
     "Wishing you all the best! 🙏"],

    ["报价后7天（最后跟进）\n7 days (last follow-up)", "CN",
     "Hi [名字]，\n\n"
     "理解你可能在比较其他方案，没关系！\n"
     "以后有任何货运需求随时找我，随时帮你查最新价格 😊\n\n"
     "[你的名字]"],

    # 成交后回访
    ["成交后 - 确认出货\nPost-booking confirmation", "EN",
     "Hi [Name],\n\n"
     "Great news! Your shipment is confirmed. Here are the details:\n\n"
     "📦 Booking Ref: [Ref]\n"
     "🚢 Vessel: [Vessel/Voy]\n"
     "📅 Cut-off: [Date]\n"
     "📅 ETD: [Date]\n"
     "📅 ETA: [Date]\n\n"
     "I'll keep you updated on the progress. Feel free to message me anytime!\n\n"
     "Best,\n[Your Name]"],

    ["成交后 - 到港通知\nArrival notification", "EN",
     "Hi [Name],\n\n"
     "Good news! Your shipment has arrived at [Port]. 🎉\n\n"
     "📦 Booking Ref: [Ref]\n"
     "📍 Status: Arrived\n"
     "📋 Next: Customs clearance (est. 2-3 days)\n\n"
     "I'll update you once it's cleared for delivery.\n\n"
     "Best,\n[Your Name]"],

    # 重新激活
    ["重新激活老客户\nRe-activate dormant", "EN",
     "Hi [Name],\n\n"
     "It's been a while! Hope everything is going well with your business.\n\n"
     "I wanted to share that shipping rates from China have been quite competitive recently. "
     "If you have any upcoming shipments, I'd love to help you get the best deal.\n\n"
     "Just a quick reply and I'll get you a quote right away! 😊"],

    ["重新激活老客户\nRe-activate dormant", "CN",
     "Hi [名字]，\n\n"
     "好久没联系了！最近生意怎么样？😊\n\n"
     "中国到马来西亚的运价最近比较有竞争力，\n"
     "如果有出货计划的话，随时找我报价！\n\n"
     "[你的名字]"],

    # 节日问候
    ["节日问候\nHoliday greeting", "EN",
     "Hi [Name],\n\n"
     "Wishing you and your team a wonderful [Holiday]! 🎉\n\n"
     "Thank you for your continued trust and partnership. "
     "Looking forward to serving you in the coming year!\n\n"
     "Best wishes,\n[Your Name]"],

    ["节日问候\nHoliday greeting", "CN",
     "Hi [名字]，\n\n"
     "祝你和团队 [节日] 快乐！🎉\n\n"
     "感谢一直以来的信任与支持，新的一年继续合作愉快！\n\n"
     "[你的名字]"],
]

for r, data in enumerate(templates, 4):
    ws_tpl.cell(row=r, column=1, value=data[0]).font = bold_font
    ws_tpl.cell(row=r, column=1).border = thin_border
    ws_tpl.cell(row=r, column=1).alignment = left_wrap
    ws_tpl.cell(row=r, column=2, value=data[1]).font = normal_font
    ws_tpl.cell(row=r, column=2).border = thin_border
    ws_tpl.cell(row=r, column=2).alignment = center
    ws_tpl.cell(row=r, column=3, value=data[2]).font = normal_font
    ws_tpl.cell(row=r, column=3).border = thin_border
    ws_tpl.cell(row=r, column=3).alignment = left_wrap
    ws_tpl.row_dimensions[r].height = 100
    # Color by language
    if data[1] == "CN":
        ws_tpl.cell(row=r, column=2).fill = PatternFill(start_color='FFE0B2', end_color='FFE0B2', fill_type='solid')
    else:
        ws_tpl.cell(row=r, column=2).fill = PatternFill(start_color='BBDEFB', end_color='BBDEFB', fill_type='solid')

tpl_widths = [22, 8, 70]
for i, w in enumerate(tpl_widths, 1):
    ws_tpl.column_dimensions[get_column_letter(i)].width = w

# ============================================================
# Sheet 5: 跟进看板 Dashboard
# ============================================================
ws_dash = wb.create_sheet("跟进看板 Dashboard")

ws_dash.merge_cells('A1:F1')
ws_dash.cell(row=1, column=1, value="📊 跟进看板 Follow-up Dashboard").font = title_font

ws_dash.cell(row=3, column=1, value="📋 跟进统计").font = big_bold

stats = [
    (4, "总跟进次数", '=COUNTA(跟进记录 Log!A4:A53)-COUNTBLANK(跟进记录 Log!A4:A53)'),
    (5, "本周跟进", '=COUNTIFS(跟进记录 Log!A4:A53,">="&(TODAY()-WEEKDAY(TODAY(),2)+1),跟进记录 Log!A4:A53,"<="&TODAY())'),
    (6, "报价跟进", '=COUNTIF(跟进记录 Log!E4:E53,"报价跟进")'),
    (7, "成交后回访", '=COUNTIF(跟进记录 Log!E4:E53,"成交后回访")'),
    (8, "重新激活", '=COUNTIF(跟进记录 Log!E4:E53,"重新激活")'),
]
for row, label, formula in stats:
    ws_dash.cell(row=row, column=1, value=label).font = bold_font
    ws_dash.cell(row=row, column=1).border = thin_border
    ws_dash.cell(row=row, column=2, value=formula).font = Font(name='Arial', bold=True, size=14, color='2F5496')
    ws_dash.cell(row=row, column=2).border = thin_border
    ws_dash.cell(row=row, column=2).alignment = center

ws_dash.cell(row=10, column=1, value="👥 客户分级统计").font = big_bold

tier_stats = [
    (11, "🥇 VIP 客户", '=COUNTIF(客户分级 Tiers!C5:C25,"*VIP*")'),
    (12, "🥈 A级 客户", '=COUNTIF(客户分级 Tiers!C5:C25,"*A级*")'),
    (13, "🥉 B级 客户", '=COUNTIF(客户分级 Tiers!C5:C25,"*B级*")'),
    (14, "⚪ 潜在客户", '=COUNTIF(客户分级 Tiers!C5:C25,"*潜在*")'),
    (15, "💀 流失客户", '=COUNTIF(客户分级 Tiers!C5:C25,"*流失*")'),
]
for row, label, formula in tier_stats:
    ws_dash.cell(row=row, column=1, value=label).font = bold_font
    ws_dash.cell(row=row, column=1).border = thin_border
    ws_dash.cell(row=row, column=2, value=formula).font = Font(name='Arial', bold=True, size=14)
    ws_dash.cell(row=row, column=2).border = thin_border
    ws_dash.cell(row=row, column=2).alignment = center

ws_dash.column_dimensions['A'].width = 20
ws_dash.column_dimensions['B'].width = 15

# Save
wb.save('/root/.openclaw/workspace/freight-agent/Followup_Agent.xlsx')
print("✅ Followup_Agent.xlsx created")
