"""
WhatsApp Quick Reply Templates for Freight Forwarder
East Malaysia Transshipment Calculator
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
orange_fill = PatternFill(start_color='F4B084', end_color='F4B084', fill_type='solid')
blue_fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
normal_font = Font(name='Arial', size=10)
bold_font = Font(name='Arial', bold=True, size=10)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

def style_header(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

def style_row(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = normal_font
        cell.alignment = left_align if col >= 3 else center_align
        cell.border = thin_border

# ============================================================
# PART 1: WhatsApp Quick Reply Templates
# ============================================================
wb = openpyxl.Workbook()

# --- Sheet 1: 询价回复 Inquiry Replies ---
ws_inq = wb.active
ws_inq.title = "询价回复"

inq_headers = ["场景\nScenario", "触发词\nTrigger", "回复模板\nReply Template"]
for col, h in enumerate(inq_headers, 1):
    ws_inq.cell(row=1, column=col, value=h)
style_header(ws_inq, 1, len(inq_headers))

templates = [
    # 初次询价
    ["初次询价 - 海运整柜",
     "How much / 多少钱 / 报价 / FCL / 整柜",
     "Hi! Thanks for reaching out 😊\n\n"
     "To quote you accurately, I need a few details:\n\n"
     "📦 1. From where? (City/Port in China)\n"
     "📍 2. To where? (Peninsular / East Malaysia?)\n"
     "📋 3. What cargo? (Product type)\n"
     "📐 4. 20GP / 40GP / 40HQ?\n"
     "📅 5. When do you need to ship?\n\n"
     "Once I have these, I'll get back to you with the best rate ASAP 👍"],

    ["初次询价 - 海运拼柜",
     "LCL / 拼柜 / CBM / 立方",
     "Hi! Happy to help with LCL shipment 😊\n\n"
     "Please share:\n\n"
     "📦 1. Loading port (China)?\n"
     "📍 2. Destination (PKL / Kuching / KK)?\n"
     "📐 3. Volume in CBM & weight in KG?\n"
     "📋 4. What goods? (for customs purpose)\n"
     "📅 5. Ready date?\n\n"
     "I'll quote you the best rate! 💪"],

    ["初次询价 - 空运",
     "Air / 空运 / kg / 公斤",
     "Hi! I can help with air freight ✈️\n\n"
     "Need these info:\n\n"
     "📦 1. From which airport? (PVG/CAN/SZX?)\n"
     "📍 2. To KUL? Or other MY airport?\n"
     "⚖️ 3. Total weight (KG)?\n"
     "📐 4. Dimensions (LxWxH)?\n"
     "📋 5. What product?\n\n"
     "Will get you a competitive rate! 🚀"],

    # 询问运价详情
    ["问All-in价还是分开报",
     "all in / total / 总共 / 一共多少",
     "The all-in rate includes:\n"
     "✅ Ocean/Air freight\n"
     "✅ THC (Terminal Handling)\n"
     "✅ Documentation fee\n"
     "✅ Customs clearance\n\n"
     "NOT included:\n"
     "❌ Trucking (depends on distance)\n"
     "❌ Insurance (optional, ~0.3% cargo value)\n"
     "❌ Storage / Inspection (if any)\n\n"
     "Want me to include trucking in the quote?"],

    ["问拖车费",
     "trucking / 拖车 / 内陆 / 到门",
     "Trucking cost depends on destination:\n\n"
     "📍 From Port Klang:\n"
     "→ KL area: ~MYR 400-600\n"
     "→ Johor: ~MYR 800-1200\n"
     "→ Penang: ~MYR 900-1300\n\n"
     "East Malaysia (via PKL):\n"
     "→ Kuching: MYR 2500-3500\n"
     "→ Kota Kinabalu: MYR 2800-4000\n\n"
     "Exact price depends on container size & delivery address. "
     "Share your delivery address and I'll get the exact rate 👍"],

    ["问保险",
     "insurance / 保险 / 货损",
     "I strongly recommend cargo insurance! 🛡️\n\n"
     "Cost: ~0.3% of cargo value\n"
     "Example: Cargo worth USD 10,000 → Insurance ~USD 30\n\n"
     "Covers: Damage, loss, theft during transit\n"
     "Without insurance: Any damage = your own risk\n\n"
     "Want me to include insurance in the quote?"],

    # 比价/砍价
    ["客户说别家更便宜",
     "cheaper / 更便宜 / 别人报价 / other forwarder",
     "I understand price matters! 💰\n\n"
     "May I ask what rate you were quoted? "
     "I'll check if I can match or beat it.\n\n"
     "Also, a few things to check when comparing:\n"
     "1. Is it truly all-in? (some quotes exclude THC, DOC)\n"
     "2. Transit time?\n"
     "3. Which shipping line?\n"
     "4. Is customs clearance included?\n\n"
     "Happy to work with your budget 😊"],

    ["客户要求降价",
     "discount / cheaper / lower / 便宜点 / 打折",
     "Let me check with my shipping line and see what I can do 🤝\n\n"
     "A few options:\n"
     "1. If you can confirm shipment today, I can try for a better rate\n"
     "2. If you have regular shipments, we can discuss volume discount\n"
     "3. I can check alternative routes/schedules\n\n"
     "Give me a moment, I'll get back to you shortly 👍"],

    # 跟进
    ["报价后跟进（1-2天）",
     "（手动触发）",
     "Hi [Name]! Just checking in on the quote I sent for [Route] 😊\n\n"
     "Have you had a chance to review it?\n"
     "Any questions, I'm happy to clarify!\n\n"
     "The rate is valid until [Date] 📅"],

    ["报价后跟进（3-5天）",
     "（手动触发）",
     "Hi [Name], hope you're doing well!\n\n"
     "Following up on the [Route] shipment quote.\n"
     "Rates are changing frequently, so wanted to make sure "
     "you don't miss the current offer.\n\n"
     "Ready to proceed? Or need any adjustments? 🙏"],

    ["成交后感谢",
     "（手动触发 - 确认成交后）",
     "Thank you for booking with us! 🎉\n\n"
     "Here's what happens next:\n"
     "1. I'll send booking confirmation shortly\n"
     "2. Shipping schedule & cut-off date will be shared\n"
     "3. I'll keep you updated on the shipment status\n\n"
     "Feel free to WhatsApp me anytime if you have questions!\n"
     "Looking forward to a long-term partnership 🤝"],

    # 特殊情况
    ["客户问时效",
     "how long / transit time / 多久 / 几天到",
     "Transit time (approx):\n\n"
     "🚢 Sea freight:\n"
     "China → Port Klang: 7-12 days\n"
     "China → East MY (via PKL): +3-5 days\n\n"
     "✈️ Air freight:\n"
     "China → KUL: 1-3 days\n\n"
     "These are port-to-port / airport-to-airport. "
     "Add 2-3 days for customs & delivery.\n\n"
     "Need it faster? Let me check express options 🚀"],

    ["客户问禁运品",
     "prohibited / banned / 禁运 / 不能运",
     "Prohibited / Restricted items include:\n\n"
     "🚫 Strictly prohibited:\n"
     "Drugs, weapons, counterfeit goods, hazardous materials\n\n"
     "⚠️ Need special license:\n"
     "Food, medicine, cosmetics, electronics (some)\n"
     "Chemicals, batteries (lithium), flammable items\n\n"
     "📋 What are you planning to ship?\n"
     "I'll check if any special permits are needed 👍"],

    ["非工作时间自动回复",
     "（触发条件：非工作时间）",
     "Hi! Thanks for your message 🙏\n\n"
     "I've received your inquiry and will get back to you "
     "during business hours (Mon-Fri 9am-6pm, Sat 9am-1pm).\n\n"
     "For urgent matters, please call: [Your Phone]\n\n"
     "Talk to you soon! 😊"],
]

for r, data in enumerate(templates, 2):
    ws_inq.cell(row=r, column=1, value=data[0])
    ws_inq.cell(row=r, column=2, value=data[1])
    ws_inq.cell(row=r, column=3, value=data[2])
    style_row(ws_inq, r, 3)
    # Color by category
    cat = data[0]
    if "初次" in cat:
        ws_inq.cell(row=r, column=1).fill = green_fill
    elif "跟进" in cat:
        ws_inq.cell(row=r, column=1).fill = yellow_fill
    elif "砍价" in cat or "便宜" in cat or "降价" in cat:
        ws_inq.cell(row=r, column=1).fill = orange_fill
    elif "成交" in cat:
        ws_inq.cell(row=r, column=1).fill = blue_fill

ws_inq.column_dimensions['A'].width = 25
ws_inq.column_dimensions['B'].width = 30
ws_inq.column_dimensions['C'].width = 65
ws_inq.freeze_panes = 'A2'

# --- Sheet 2: 中文话术 Chinese Templates ---
ws_cn = wb.create_sheet("中文话术")

cn_headers = ["场景", "回复模板"]
for col, h in enumerate(cn_headers, 1):
    ws_cn.cell(row=1, column=col, value=h)
style_header(ws_cn, 1, len(cn_headers))

cn_templates = [
    ["收到询价",
     "收到！我帮你查一下价格 👍\n\n"
     "请确认一下信息：\n"
     "1. 从中国哪里出？（哪个城市/港口）\n"
     "2. 到马来西亚哪里？\n"
     "3. 什么货？整柜还是拼柜？\n"
     "4. 大概什么时候出货？\n\n"
     "确认后马上给你报价！"],

    ["发送报价",
     "Hi [名字]，报价来了 👇\n\n"
     "📦 [柜型] [起运地] → [目的地]\n"
     "💰 USD [总价] (all-in)\n\n"
     "包含：海运费、THC、文件费、报关费\n"
     "不含：拖车费、保险\n\n"
     "有效期7天，有问题随时联系我 🙏"],

    ["客户说贵",
     "理解你的顾虑 💰\n\n"
     "方便告诉我别家报多少吗？\n"
     "我看看能不能帮你争取更好的价格。\n\n"
     "另外提醒一下，比较价格时注意看：\n"
     "1. 是否真的 all-in？（有些报价不含THC和文件费）\n"
     "2. 航程多少天？\n"
     "3. 用的是哪家船公司？\n\n"
     "我这边保证透明报价，没有隐藏费用 ✅"],

    ["跟进未回复客户",
     "Hi [名字]，之前给你的报价有看了吗？😊\n\n"
     "运价最近变动比较大，想帮你确认一下\n"
     "现在的价格还能不能拿到。\n\n"
     "有什么问题直接问我就好！"],

    ["确认成交",
     "收到确认！🎉 谢谢你的信任！\n\n"
     "接下来的流程：\n"
     "1. 我会发 booking confirmation\n"
     "2. 告诉你截关时间和船期\n"
     "3. 全程跟踪，有更新随时通知你\n\n"
     "有任何问题随时 WhatsApp 我！"],

    ["客户问时效",
     "时效参考：\n\n"
     "🚢 海运：中国 → 巴生港 7-12天\n"
     "中国 → 东马（经PKL中转）再加3-5天\n\n"
     "✈️ 空运：中国 → 吉隆坡 1-3天\n\n"
     "以上是港到港/机场到机场的时间，\n"
     "加上清关和派送一般还要2-3天。\n\n"
     "需要加急的方案也可以帮你查 🚀"],
]

for r, data in enumerate(cn_templates, 2):
    ws_cn.cell(row=r, column=1, value=data[0])
    ws_cn.cell(row=r, column=2, value=data[1])
    style_row(ws_cn, r, 2)

ws_cn.column_dimensions['A'].width = 20
ws_cn.column_dimensions['B'].width = 65
ws_cn.freeze_panes = 'A2'

# --- Sheet 3: 微信话术 WeChat Templates ---
ws_wc = wb.create_sheet("微信话术")
# Same structure as Chinese, copy content
for col, h in enumerate(cn_headers, 1):
    ws_wc.cell(row=1, column=col, value=h)
style_header(ws_wc, 1, len(cn_headers))
for r, data in enumerate(cn_templates, 2):
    ws_wc.cell(row=r, column=1, value=data[0])
    ws_wc.cell(row=r, column=2, value=data[1])
    style_row(ws_wc, r, 2)
ws_wc.column_dimensions['A'].width = 20
ws_wc.column_dimensions['B'].width = 65
ws_wc.freeze_panes = 'A2'

# Save
wb.save('/root/.openclaw/workspace/freight-agent/WhatsApp_Templates.xlsx')
print("✅ WhatsApp_Templates.xlsx created")

# ============================================================
# PART 2: East Malaysia Transshipment Calculator
# ============================================================
wb2 = openpyxl.Workbook()

# --- Sheet 1: Calculator ---
ws_calc = wb2.active
ws_calc.title = "计算器 Calculator"

# Title
ws_calc.merge_cells('A1:G1')
ws_calc.cell(row=1, column=1, value="🚢 East Malaysia 转运费用计算器")
ws_calc.cell(row=1, column=1).font = Font(name='Arial', bold=True, size=16, color='2F5496')

ws_calc.merge_cells('A2:G2')
ws_calc.cell(row=2, column=1, value="China → Port Klang → East Malaysia (Kuching / Kota Kinabalu / Bintulu / Miri / Sibu)")
ws_calc.cell(row=2, column=1).font = Font(name='Arial', size=10, color='666666')

# Input section
ws_calc.cell(row=4, column=1, value="📥 输入区 INPUT").font = Font(name='Arial', bold=True, size=13, color='2F5496')

input_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')

inputs = [
    (5, "运输方式 Mode:", "Sea FCL", "Sea FCL / Sea LCL / Air"),
    (6, "起运港 POL:", "Shanghai", "Shanghai / Ningbo / Shenzhen / Guangzhou"),
    (7, "目的地 POD:", "Kuching", "Kuching / Kota Kinabalu / Bintulu / Miri / Sibu"),
    (8, "柜型 Container:", "40HQ", "20GP / 40GP / 40HQ (FCL) 或 CBM (LCL)"),
    (9, "货量 Volume:", "1", "FCL: 1柜 / LCL: CBM数 / Air: KG数"),
    (10, "货值 Cargo Value (USD):", "10000", "用于计算保险费"),
    (11, "Margin %:", "20", "你的利润率 10%-30%"),
    (12, "是否含保险 Insurance:", "No", "Yes / No"),
    (13, "是否含拖车 Trucking:", "No", "Yes / No (East MY delivery)"),
]
for row, label, default, note in inputs:
    ws_calc.cell(row=row, column=1, value=label).font = bold_font
    ws_calc.cell(row=row, column=1).border = thin_border
    ws_calc.cell(row=row, column=2, value=default).font = normal_font
    ws_calc.cell(row=row, column=2).fill = input_fill
    ws_calc.cell(row=row, column=2).border = thin_border
    ws_calc.cell(row=row, column=2).alignment = center_align
    ws_calc.cell(row=row, column=4, value=note).font = Font(name='Arial', italic=True, size=9, color='999999')

# Data validation for inputs
dv_mode = DataValidation(type="list", formula1='"Sea FCL,Sea LCL,Air"', allow_blank=True)
ws_calc.add_data_validation(dv_mode)
dv_mode.add('B5')

dv_pol = DataValidation(type="list", formula1='"Shanghai,Ningbo,Shenzhen,Guangzhou,Xiamen,Qingdao"', allow_blank=True)
ws_calc.add_data_validation(dv_pol)
dv_pol.add('B6')

dv_pod = DataValidation(type="list", formula1='"Kuching,Kota Kinabalu,Bintulu,Miri,Sibu"', allow_blank=True)
ws_calc.add_data_validation(dv_pod)
dv_pod.add('B7')

dv_container = DataValidation(type="list", formula1='"20GP,40GP,40HQ,LCL"', allow_blank=True)
ws_calc.add_data_validation(dv_container)
dv_container.add('B8')

dv_yn = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
ws_calc.add_data_validation(dv_yn)
dv_yn.add('B12')
dv_yn.add('B13')

# Output section - Cost Breakdown
ws_calc.cell(row=15, column=1, value="💰 费用拆解 COST BREAKDOWN").font = Font(name='Arial', bold=True, size=13, color='2F5496')

out_headers = ["费用项目 Charge", "币种 Currency", "成本 Cost", "备注 Notes"]
for col, h in enumerate(out_headers, 1):
    cell = ws_calc.cell(row=16, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = center_align

# Leg 1: China to PKL
ws_calc.cell(row=17, column=1, value="── 第一段：中国 → 巴生港 ──").font = Font(name='Arial', bold=True, size=10, color='2F5496')

leg1_items = [
    (18, "海运费 Ocean Freight", "USD", '=IF(B5="Sea FCL",VLOOKUP(B6,Rate_Ref_FCL,IF(B8="20GP",2,IF(B8="40GP",3,4)),FALSE),IF(B5="Sea LCL",VLOOKUP(B6,Rate_Ref_LCL,2,FALSE)*B9,VLOOKUP(B6,Rate_Ref_Air,2,FALSE)*B9))',
     "从运价表自动查"),
    (19, "THC 码头费", "USD", '=IF(B5="Air","-",IF(B5="Sea FCL",150,150))', "Per shipment"),
    (20, "DOC 文件费", "USD", 50, "Per shipment"),
    (21, "BAF 燃油附加费", "USD", '=IF(B5="Air","Included",50)', "估算值，按实际调整"),
    (22, "ISPS 安保费", "USD", '=IF(B5="Air","-",15)', "Per container"),
    (23, "第一段小计 Subtotal 1", "USD",
     '=IF(C18="-",0,C18)+IF(C19="-",0,C19)+C20+IF(C21="-",0,C21)+IF(C22="-",0,C22)',
     ""),
]

for row, item, curr, cost, note in leg1_items:
    ws_calc.cell(row=row, column=1, value=item).font = normal_font
    ws_calc.cell(row=row, column=1).border = thin_border
    ws_calc.cell(row=row, column=2, value=curr).font = normal_font
    ws_calc.cell(row=row, column=2).border = thin_border
    ws_calc.cell(row=row, column=2).alignment = center_align
    ws_calc.cell(row=row, column=3, value=cost).font = normal_font
    ws_calc.cell(row=row, column=3).border = thin_border
    ws_calc.cell(row=row, column=3).alignment = center_align
    ws_calc.cell(row=row, column=3).number_format = '#,##0.00'
    ws_calc.cell(row=row, column=4, value=note).font = Font(name='Arial', italic=True, size=9, color='666666')
    ws_calc.cell(row=row, column=4).border = thin_border

ws_calc.cell(row=23, column=1).font = bold_font

# Leg 2: PKL to East MY
ws_calc.cell(row=25, column=1, value="── 第二段：巴生港 → 东马 ──").font = Font(name='Arial', bold=True, size=10, color='2F5496')

leg2_items = [
    (26, "本地运费 Local Freight (PKL→POD)", "MYR",
     '=IF(B7="Kuching",2800,IF(B7="Kota Kinabalu",3200,IF(B7="Bintulu",3500,IF(B7="Miri",3800,3000))))',
     "参考价格，按实际调整"),
    (27, "报关费 Customs (PKL)", "MYR", 200, "Per shipment"),
    (28, "报关费 Customs (East MY)", "MYR", 250, "Per shipment"),
    (29, "港口费 Port Charges (East MY)", "MYR", 150, "Estimated"),
    (30, "第二段小计 Subtotal 2", "MYR", '=C26+C27+C28+C29', ""),
]
for row, item, curr, cost, note in leg2_items:
    ws_calc.cell(row=row, column=1, value=item).font = normal_font
    ws_calc.cell(row=row, column=1).border = thin_border
    ws_calc.cell(row=row, column=2, value=curr).font = normal_font
    ws_calc.cell(row=row, column=2).border = thin_border
    ws_calc.cell(row=row, column=2).alignment = center_align
    ws_calc.cell(row=row, column=3, value=cost).font = normal_font
    ws_calc.cell(row=row, column=3).border = thin_border
    ws_calc.cell(row=row, column=3).alignment = center_align
    ws_calc.cell(row=row, column=3).number_format = '#,##0.00'
    ws_calc.cell(row=row, column=4, value=note).font = Font(name='Arial', italic=True, size=9, color='666666')
    ws_calc.cell(row=row, column=4).border = thin_border
ws_calc.cell(row=30, column=1).font = bold_font

# Optional costs
ws_calc.cell(row=32, column=1, value="── 可选费用 Optional ──").font = Font(name='Arial', bold=True, size=10, color='2F5496')

opt_items = [
    (33, "保险 Insurance", "USD", '=IF(B12="Yes",B11*0.003,"-")', "0.3% of cargo value"),
    (34, "拖车费 Trucking (East MY delivery)", "MYR",
     '=IF(B13="Yes",IF(B7="Kuching",600,IF(B7="Kota Kinabalu",800,500)),"-")',
     "市区内参考价"),
]
for row, item, curr, cost, note in opt_items:
    ws_calc.cell(row=row, column=1, value=item).font = normal_font
    ws_calc.cell(row=row, column=1).border = thin_border
    ws_calc.cell(row=row, column=2, value=curr).font = normal_font
    ws_calc.cell(row=row, column=2).border = thin_border
    ws_calc.cell(row=row, column=2).alignment = center_align
    ws_calc.cell(row=row, column=3, value=cost).font = normal_font
    ws_calc.cell(row=row, column=3).border = thin_border
    ws_calc.cell(row=row, column=3).alignment = center_align
    ws_calc.cell(row=row, column=3).number_format = '#,##0.00'
    ws_calc.cell(row=row, column=4, value=note).font = Font(name='Arial', italic=True, size=9, color='666666')
    ws_calc.cell(row=row, column=4).border = thin_border

# Total & Quote
ws_calc.cell(row=36, column=1, value="📊 报价汇总 QUOTE SUMMARY").font = Font(name='Arial', bold=True, size=13, color='2F5496')

summary_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')

summaries = [
    (37, "总成本 Total Cost (USD)", "USD",
     '=C23+C30/4.5+IF(C33="-",0,C33)+IF(C34="-",0,C34/4.5)',
     "MYR÷4.5 粗略换算，按实际汇率调整"),
    (38, "Margin 金额", "USD", '=C37*B11/100', ""),
    (39, "报价 Selling Price (USD)", "USD", '=C37+C38', "发给客户的价格"),
    (40, "汇率参考 Rate", "MYR/USD", 4.5, "更新为当前汇率"),
]
for row, item, curr, cost, note in summaries:
    ws_calc.cell(row=row, column=1, value=item).font = bold_font
    ws_calc.cell(row=row, column=1).border = thin_border
    ws_calc.cell(row=row, column=2, value=curr).font = bold_font
    ws_calc.cell(row=row, column=2).border = thin_border
    ws_calc.cell(row=row, column=2).alignment = center_align
    ws_calc.cell(row=row, column=3, value=cost).font = Font(name='Arial', bold=True, size=12)
    ws_calc.cell(row=row, column=3).border = thin_border
    ws_calc.cell(row=row, column=3).alignment = center_align
    ws_calc.cell(row=row, column=3).number_format = '#,##0.00'
    ws_calc.cell(row=row, column=3).fill = summary_fill
    ws_calc.cell(row=row, column=4, value=note).font = Font(name='Arial', italic=True, size=9, color='666666')
    ws_calc.cell(row=row, column=4).border = thin_border

# Highlight final quote
ws_calc.cell(row=39, column=1).fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
ws_calc.cell(row=39, column=3).fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

# Column widths
ws_calc.column_dimensions['A'].width = 35
ws_calc.column_dimensions['B'].width = 14
ws_calc.column_dimensions['C'].width = 18
ws_calc.column_dimensions['D'].width = 35

# --- Sheet 2: East MY Route Reference ---
ws_route = wb2.create_sheet("东马航线参考")

ws_route.merge_cells('A1:F1')
ws_route.cell(row=1, column=1, value="🚢 东马转运参考信息 East Malaysia Transshipment Reference")
ws_route.cell(row=1, column=1).font = Font(name='Arial', bold=True, size=14, color='2F5496')

route_headers = ["目的地\nDestination", "港口代码\nPort Code", "转运方式\nVia", "额外航程\nExtra Days",
                 "本地运费参考\nLocal Freight (MYR)", "备注\nNotes"]
for col, h in enumerate(route_headers, 1):
    ws_route.cell(row=3, column=col, value=h)
style_header(ws_route, 3, len(route_headers))

routes = [
    ["Kuching 古晋", "MYKCH", "PKL → Kuching (feeder)", "3-5 days", "2,500 - 3,500", "东马最大城市，班次较多"],
    ["Kota Kinabalu 亚庇", "MYBKI", "PKL → KK (feeder)", "4-6 days", "2,800 - 4,000", "沙巴州首府"],
    ["Bintulu 民都鲁", "MYBTU", "PKL → Bintulu (feeder)", "4-6 days", "3,200 - 4,200", "LNG 工业区"],
    ["Miri 美里", "MYMYY", "PKL → Miri (feeder)", "5-7 days", "3,500 - 4,500", "石油城市"],
    ["Sibu 诗巫", "MYSBW", "PKL → Sibu (feeder)", "4-6 days", "2,800 - 3,800", "木材/农业区"],
]
for r, data in enumerate(routes, 4):
    for c, val in enumerate(data, 1):
        ws_route.cell(row=r, column=c, value=val)
    style_row(ws_route, r, len(route_headers))

route_widths = [22, 12, 25, 14, 22, 30]
for i, w in enumerate(route_widths, 1):
    ws_route.column_dimensions[get_column_letter(i)].width = w

# Additional info
ws_route.cell(row=10, column=1, value="📝 重要提醒 Important Notes").font = Font(name='Arial', bold=True, size=12)
notes = [
    "1. 东马转运必须经 Port Klang 中转，没有直航",
    "2. 转运时间受 feeder 船期影响，可能有延误",
    "3. 东马属于不同关税区，需要额外报关",
    "4. East MY 有 Borneo 特殊政策，某些货物有税收优惠",
    "5. 本地运费以整柜为单位，拼柜按 CBM 比例分摊",
    "6. 以上价格为参考值，实际价格请跟当地代理确认",
]
for i, n in enumerate(notes):
    ws_route.cell(row=11 + i, column=1, value=n).font = normal_font

# --- Sheet 3: Margin Calculator ---
ws_margin = wb2.create_sheet("Margin参考")

ws_margin.merge_cells('A1:E1')
ws_margin.cell(row=1, column=1, value="💰 Margin 快速参考 Quick Reference")
ws_margin.cell(row=1, column=1).font = Font(name='Arial', bold=True, size=14, color='2F5496')

m_headers = ["成本\nCost", "10% Margin", "15% Margin", "20% Margin", "25% Margin", "30% Margin"]
for col, h in enumerate(m_headers, 1):
    ws_margin.cell(row=3, column=col, value=h)
style_header(ws_margin, 3, len(m_headers))

costs = [100, 200, 300, 400, 500, 600, 700, 800, 900, 1000, 1500, 2000, 3000, 5000]
for r, cost in enumerate(costs, 4):
    ws_margin.cell(row=r, column=1, value=cost).font = bold_font
    ws_margin.cell(row=r, column=1).border = thin_border
    ws_margin.cell(row=r, column=1).number_format = '#,##0.00'
    for c, pct in enumerate([10, 15, 20, 25, 30], 2):
        cell = ws_margin.cell(row=r, column=c, value=round(cost / (1 - pct/100), 2))
        cell.font = normal_font
        cell.border = thin_border
        cell.number_format = '#,##0.00'
        cell.alignment = center_align

m_widths = [12, 14, 14, 14, 14, 14]
for i, w in enumerate(m_widths, 1):
    ws_margin.column_dimensions[get_column_letter(i)].width = w

ws_margin.cell(row=r+2, column=1, value="公式 Formula:").font = bold_font
ws_margin.cell(row=r+2, column=2, value="报价 = 成本 ÷ (1 - Margin%)").font = normal_font
ws_margin.cell(row=r+3, column=1, value="例如:").font = bold_font
ws_margin.cell(row=r+3, column=2, value="成本500 + 20% margin = 500 ÷ 0.8 = 625 USD").font = normal_font

wb2.save('/root/.openclaw/workspace/freight-agent/East_MY_Calculator.xlsx')
print("✅ East_MY_Calculator.xlsx created")
print("\n🎉 All files created!")
