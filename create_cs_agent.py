"""
Customer Service Agent - 客服管理
常见问题、自动回复、货物追踪、工单管理
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
input_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
blue_fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
grey_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
bold_font = Font(name='Arial', bold=True, size=10)
big_bold = Font(name='Arial', bold=True, size=13, color='2F5496')
title_font = Font(name='Arial', bold=True, size=18, color='2F5496')
normal_font = Font(name='Arial', size=10)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_wrap = Alignment(horizontal='left', vertical='center', wrap_text=True)

wb = openpyxl.Workbook()

# ============================================================
# Sheet 1: FAQ 常见问题
# ============================================================
ws_faq = wb.active
ws_faq.title = "FAQ 常见问题"

ws_faq.merge_cells('A1:C1')
ws_faq.cell(row=1, column=1, value="❓ FAQ 常见问题 Quick Response").font = title_font

ws_faq.merge_cells('A2:C2')
ws_faq.cell(row=2, column=1, value="客户问这些问题时，直接复制回复").font = Font(name='Arial', size=10, color='666666')

faq_headers = ["问题 Question (EN)", "问题 Question (CN)", "回复 Reply"]
for col, h in enumerate(faq_headers, 1):
    cell = ws_faq.cell(row=4, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = center

faqs = [
    # 运价相关
    ["How much to ship from China to MY?",
     "从中国运到马来西亚多少钱？",
     "Price depends on: cargo type, volume/weight, origin port, destination, "
     "and shipping method (sea/air).\n\n"
     "Please share:\n1. From which city/port?\n2. To where in MY?\n3. What cargo?\n"
     "4. Volume (CBM) or weight (KG)?\n5. When do you need it?\n\n"
     "I'll get back with a quote ASAP! 🚢"],

    ["How long does shipping take?",
     "运输要多久？",
     "🚢 Sea freight: China → Port Klang: 7-12 days\n"
     "China → East MY: add 3-5 days (transshipment)\n\n"
     "✈️ Air freight: China → KUL: 1-3 days\n\n"
     "Plus 2-3 days for customs clearance & delivery.\n"
     "Express options available if urgent! 🚀"],

    ["What documents do I need?",
     "需要什么文件？",
     "For import to Malaysia, you'll need:\n\n"
     "1. Commercial Invoice (商业发票)\n"
     "2. Packing List (装箱单)\n"
     "3. Bill of Lading / Air Waybill (提单)\n"
     "4. Certificate of Origin (if applicable)\n\n"
     "Don't worry — I'll help you prepare everything! 📋"],

    ["Do you handle customs clearance?",
     "你们帮忙清关吗？",
     "Yes! We handle the full process:\n"
     "✅ Customs declaration\n"
     "✅ Duty & tax calculation\n"
     "✅ Document preparation\n"
     "✅ Coordination with customs\n\n"
     "Just provide the cargo details and we take care of the rest 👍"],

    ["Can you ship to East Malaysia?",
     "能运到东马吗？",
     "Absolutely! We regularly ship to all East MY destinations:\n"
     "📍 Kuching, Kota Kinabalu, Bintulu, Miri, Sibu\n\n"
     "Route: China → Port Klang → East MY (feeder vessel)\n"
     "Extra transit: 3-5 days\n"
     "Both FCL and LCL available 🚢"],

    # 操作相关
    ["Where is my shipment?",
     "我的货到哪了？",
     "Let me check the latest status for you! 🔍\n\n"
     "Could you share your booking reference number?\n"
     "Or the B/L number?\n\n"
     "I'll get back to you with the update shortly."],

    ["My cargo is damaged!",
     "我的货损坏了！",
     "I'm sorry to hear that! Let me help resolve this. 🙏\n\n"
     "Please provide:\n"
     "1. Photos of the damage\n"
     "2. Booking/B/L reference\n"
     "3. Description of what happened\n\n"
     "I'll file a claim with the carrier immediately.\n"
     "Did you purchase cargo insurance?"],

    ["Can I change the delivery address?",
     "可以改送货地址吗？",
     "It depends on the shipment status:\n\n"
     "📦 Before departure: Usually yes, with possible extra charges\n"
     "🚢 After departure: Limited, depends on carrier\n"
     "📍 At destination: Can usually adjust final delivery\n\n"
     "Share your booking ref and new address, I'll check what's possible."],

    # 付款相关
    ["What payment methods do you accept?",
     "接受什么付款方式？",
     "We accept:\n"
     "✅ Bank Transfer (T/T)\n"
     "✅ Online Banking (FPX)\n"
     "✅ For regular clients: 30-day credit terms available\n\n"
     "Payment terms: Prepaid or upon B/L release\n"
     "Details will be in the invoice. 💰"],

    ["Do you offer credit terms?",
     "可以月结吗？",
     "For regular clients, yes! 🤝\n\n"
     "After 3 successful shipments, we can discuss:\n"
     "• 15-day or 30-day credit terms\n"
     "• Monthly consolidated billing\n\n"
     "Let's build the partnership first!"],

    # 特殊货物
    ["Can you ship dangerous goods / batteries?",
     "能运危险品/电池吗？",
     "Dangerous goods (DG) require special handling:\n"
     "✅ MSDS (Material Safety Data Sheet) required\n"
     "✅ DG declaration & packaging\n"
     "✅ Special container/routing if needed\n\n"
     "⚠️ Some items are strictly prohibited.\n"
     "Please share the MSDS and I'll check feasibility."],

    ["Can you ship food / cosmetics?",
     "能运食品/化妆品吗？",
     "Yes, but with requirements:\n\n"
     "🍎 Food: Needs MESTI/Halal certification, lab testing\n"
     "💄 Cosmetics: Needs NPRA registration\n\n"
     "Process takes longer and costs more.\n"
     "Share the product details and I'll advise the requirements."],
]

for r, data in enumerate(faqs, 5):
    for c, val in enumerate(data, 1):
        cell = ws_faq.cell(row=r, column=c, value=val)
        cell.font = normal_font
        cell.border = thin_border
        cell.alignment = left_wrap
    ws_faq.row_dimensions[r].height = 120

faq_widths = [30, 25, 55]
for i, w in enumerate(faq_widths, 1):
    ws_faq.column_dimensions[get_column_letter(i)].width = w

# ============================================================
# Sheet 2: 货物追踪 Shipment Tracker
# ============================================================
ws_track = wb.create_sheet("货物追踪 Tracker")

ws_track.merge_cells('A1:I1')
ws_track.cell(row=1, column=1, value="📍 货物追踪 Shipment Tracker").font = title_font

track_headers = ["订单号\nBooking Ref", "客户\nCustomer", "起运地\nPOL", "目的地\nPOD",
                 "柜号\nContainer", "ETD", "ETA", "状态\nStatus", "更新时间\nUpdated"]
for col, h in enumerate(track_headers, 1):
    cell = ws_track.cell(row=3, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = center

for r in range(4, 24):
    for c in range(1, 10):
        cell = ws_track.cell(row=r, column=c)
        cell.border = thin_border
        cell.alignment = center
        cell.font = normal_font
    ws_track.cell(row=r, column=8).fill = input_fill

dv_track_status = DataValidation(
    type="list",
    formula1='"📋 已订舱,📦 已装柜,🚢 已离港,🌊 在途,📍 已到港,🛃 清关中,✅ 已放行,🚚 派送中,🏁 已签收"',
    allow_blank=True
)
ws_track.add_data_validation(dv_track_status)
dv_track_status.add('H4:H23')

track_widths = [16, 16, 12, 12, 16, 12, 12, 12, 14]
for i, w in enumerate(track_widths, 1):
    ws_track.column_dimensions[get_column_letter(i)].width = w
ws_track.freeze_panes = 'A4'

# ============================================================
# Sheet 3: 工单管理 Support Tickets
# ============================================================
ws_ticket = wb.create_sheet("工单管理 Tickets")

ws_ticket.merge_cells('A1:J1')
ws_ticket.cell(row=1, column=1, value="🎫 工单管理 Support Tickets").font = title_font

ticket_headers = ["工单号\nTicket No.", "日期\nDate", "客户\nCustomer",
                  "渠道\nChannel", "问题类型\nType", "问题描述\nDescription",
                  "处理人\nHandler", "状态\nStatus", "解决时间\nResolved", "备注\nNotes"]
for col, h in enumerate(ticket_headers, 1):
    cell = ws_ticket.cell(row=3, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = center

for r in range(4, 34):
    for c in range(1, 11):
        cell = ws_ticket.cell(row=r, column=c)
        cell.border = thin_border
        cell.alignment = center if c not in [6, 10] else left_wrap
        cell.font = normal_font
    ws_ticket.cell(row=r, column=1).fill = grey_fill
    ws_ticket.cell(row=r, column=2).fill = input_fill
    ws_ticket.cell(row=r, column=8).fill = input_fill

dv_ticket_type = DataValidation(
    type="list",
    formula1='"询价,货物追踪,货物损坏,清关问题,账单问题,改地址,投诉,其他"',
    allow_blank=True
)
ws_ticket.add_data_validation(dv_ticket_type)
dv_ticket_type.add('E4:E33')

dv_ticket_status = DataValidation(
    type="list",
    formula1='"🆕 新建,⏳ 处理中,✅ 已解决,❌ 无法解决,🔄 升级处理"',
    allow_blank=True
)
ws_ticket.add_data_validation(dv_ticket_status)
dv_ticket_status.add('H4:H33')

ticket_widths = [12, 12, 16, 10, 12, 30, 10, 12, 12, 20]
for i, w in enumerate(ticket_widths, 1):
    ws_ticket.column_dimensions[get_column_letter(i)].width = w
ws_ticket.freeze_panes = 'A4'

# ============================================================
# Sheet 4: 客服话术 Service Scripts
# ============================================================
ws_script = wb.create_sheet("客服话术 Scripts")

ws_script.merge_cells('A1:C1')
ws_script.cell(row=1, column=1, value="🎙️ 客服话术 Service Scripts").font = title_font

script_headers = ["场景 Scenario", "语气 Tone", "话术 Script"]
for col, h in enumerate(script_headers, 1):
    cell = ws_script.cell(row=3, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = center

scripts = [
    ["开场问候\nGreeting", "友好",
     "Hi! Thank you for contacting us. 🙏\n"
     "How can I help you today?\n\n"
     "您好！感谢联系我们，请问有什么可以帮您？"],

    ["无法立即回答\nNeed time", "专业",
     "Thank you for your patience. Let me check with our team "
     "and get back to you within [X hours].\n\n"
     "感谢您的耐心，我需要跟团队确认一下，[X小时]内回复您。"],

    ["投诉处理-第一步\nComplaint step 1", "同理心",
     "I completely understand your frustration, and I'm sorry for the inconvenience. "
     "Let me look into this right away and find a solution for you.\n\n"
     "非常理解您的感受，对此造成的不便深表歉意。我马上查看并为您解决。"],

    ["投诉处理-道歉\nComplaint apology", "真诚",
     "I sincerely apologize for this experience. "
     "This is not the standard of service we aim to provide. "
     "Here's what I'm going to do to fix it:\n"
     "1. [Action 1]\n"
     "2. [Action 2]\n\n"
     "对此我们深表歉意，以下是我们的解决方案..."],

    ["拒绝不合理要求\nDecline politely", "坚定但礼貌",
     "I understand your request, and I wish I could accommodate it. "
     "Unfortunately, due to [reason], we're unable to do so. "
     "However, here's an alternative option that might work:\n"
     "[Alternative]\n\n"
     "理解您的需求，但由于[原因]无法满足，但我们有以下替代方案..."],

    ["催款\nPayment reminder", "专业友好",
     "Hi [Name], gentle reminder that Invoice [No.] dated [Date] "
     "for [Amount] is due on [Due Date].\n\n"
     "Please arrange payment at your earliest convenience. "
     "If already paid, kindly ignore this message. Thank you! 🙏\n\n"
     "您好，温馨提醒 Invoice [编号] 将于 [日期] 到期，请安排付款。"],

    ["结束对话\nClosing",
     "友好",
     "Is there anything else I can help you with?\n\n"
     "If not, thank you for reaching out! Don't hesitate to contact us "
     "anytime. Have a great day! 😊\n\n"
     "还有其他需要帮助的吗？没有的话感谢您的联系，祝您愉快！"],
]

for r, data in enumerate(scripts, 4):
    ws_script.cell(row=r, column=1, value=data[0]).font = bold_font
    ws_script.cell(row=r, column=1).border = thin_border
    ws_script.cell(row=r, column=1).alignment = center
    ws_script.cell(row=r, column=2, value=data[1]).font = normal_font
    ws_script.cell(row=r, column=2).border = thin_border
    ws_script.cell(row=r, column=2).alignment = center
    ws_script.cell(row=r, column=3, value=data[2]).font = normal_font
    ws_script.cell(row=r, column=3).border = thin_border
    ws_script.cell(row=r, column=3).alignment = left_wrap
    ws_script.row_dimensions[r].height = 100

script_widths = [18, 10, 70]
for i, w in enumerate(script_widths, 1):
    ws_script.column_dimensions[get_column_letter(i)].width = w

# ============================================================
# Sheet 5: 客服看板 Dashboard
# ============================================================
ws_dash = wb.create_sheet("客服看板 Dashboard")

ws_dash.merge_cells('A1:F1')
ws_dash.cell(row=1, column=1, value="📊 客服看板 Service Dashboard").font = title_font

ws_dash.cell(row=3, column=1, value="🎫 工单统计 Ticket Stats").font = big_bold

stats = [
    (4, "总工单数", '=COUNTA(工单管理 Tickets!A4:A33)-COUNTBLANK(工单管理 Tickets!A4:A33)'),
    (5, "🆕 新建", '=COUNTIF(工单管理 Tickets!H4:H33,"*新建*")'),
    (6, "⏳ 处理中", '=COUNTIF(工单管理 Tickets!H4:H33,"*处理中*")'),
    (7, "✅ 已解决", '=COUNTIF(工单管理 Tickets!H4:H33,"*已解决*")'),
    (8, "❌ 无法解决", '=COUNTIF(工单管理 Tickets!H4:H33,"*无法解决*")'),
]
for row, label, formula in stats:
    ws_dash.cell(row=row, column=1, value=label).font = bold_font
    ws_dash.cell(row=row, column=1).border = thin_border
    ws_dash.cell(row=row, column=2, value=formula).font = Font(name='Arial', bold=True, size=14, color='2F5496')
    ws_dash.cell(row=row, column=2).border = thin_border
    ws_dash.cell(row=row, column=2).alignment = center

ws_dash.cell(row=10, column=1, value="📍 货物追踪统计").font = big_bold

t_stats = [
    (11, "在途货物", '=COUNTIF(货物追踪 Tracker!H4:H23,"*在途*")+COUNTIF(货物追踪 Tracker!H4:H23,"*已离港*")'),
    (12, "清关中", '=COUNTIF(货物追踪 Tracker!H4:H23,"*清关*")'),
    (13, "已签收", '=COUNTIF(货物追踪 Tracker!H4:H23,"*已签收*")'),
]
for row, label, formula in t_stats:
    ws_dash.cell(row=row, column=1, value=label).font = bold_font
    ws_dash.cell(row=row, column=1).border = thin_border
    ws_dash.cell(row=row, column=2, value=formula).font = Font(name='Arial', bold=True, size=14)
    ws_dash.cell(row=row, column=2).border = thin_border
    ws_dash.cell(row=row, column=2).alignment = center

ws_dash.column_dimensions['A'].width = 20
ws_dash.column_dimensions['B'].width = 15

wb.save('/root/.openclaw/workspace/freight-agent/Customer_Service_Agent.xlsx')
print("✅ Customer_Service_Agent.xlsx created")
