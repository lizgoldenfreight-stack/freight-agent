"""
Social Media Agent - 社交媒体内容管理
LinkedIn / Facebook 内容规划、帖子模板、行业素材
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime

header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
input_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
blue_fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
orange_fill = PatternFill(start_color='F4B084', end_color='F4B084', fill_type='solid')
bold_font = Font(name='Arial', bold=True, size=10)
big_bold = Font(name='Arial', bold=True, size=13, color='2F5496')
title_font = Font(name='Arial', bold=True, size=18, color='2F5496')
normal_font = Font(name='Arial', size=10)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_wrap = Alignment(horizontal='left', vertical='center', wrap_text=True)

wb = openpyxl.Workbook()

# ============================================================
# Sheet 1: 内容日历 Content Calendar
# ============================================================
ws_cal = wb.active
ws_cal.title = "内容日历 Calendar"

ws_cal.merge_cells('A1:G1')
ws_cal.cell(row=1, column=1, value="📅 内容发布日历 Content Calendar").font = title_font

ws_cal.merge_cells('A2:G2')
ws_cal.cell(row=2, column=1, value="每周发布 2-3 条，保持活跃度。提前规划，批量创作").font = Font(name='Arial', size=10, color='666666')

cal_headers = ["日期\nDate", "平台\nPlatform", "类型\nType", "主题\nTopic",
               "内容摘要\nSummary", "状态\nStatus", "发布链接\nLink"]
for col, h in enumerate(cal_headers, 1):
    cell = ws_cal.cell(row=4, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = center

# Pre-fill 4 weeks of content ideas
cal_samples = [
    # Week 1
    ["2026-04-07", "LinkedIn", "行业知识", "海运 vs 空运：怎么选？",
     "帮客户分析海运空运的选择标准，适合中小企业", "⬜ 待写", ""],
    ["2026-04-09", "Facebook", "案例分享", "成功案例：紧急空运",
     "分享一个客户紧急空运的案例，体现专业性", "⬜ 待写", ""],
    ["2026-04-11", "LinkedIn", "市场动态", "2026 Q2 运价趋势",
     "分享近期运价走势，给客户参考", "⬜ 待写", ""],
    # Week 2
    ["2026-04-14", "LinkedIn", "专业知识", "报关避坑指南",
     "新手进口商常见的报关错误和解决方案", "⬜ 待写", ""],
    ["2026-04-16", "Facebook", "互动话题", "你最头疼的物流问题？",
     "发起互动，增加曝光和评论", "⬜ 待写", ""],
    ["2026-04-18", "LinkedIn", "个人故事", "为什么我选择做 Freight Forwarder",
     "个人创业故事，建立信任和人设", "⬜ 待写", ""],
    # Week 3
    ["2026-04-21", "LinkedIn", "行业知识", "FOB vs CIF：贸易术语解释",
     "通俗解释常见贸易术语，帮助客户理解", "⬜ 待写", ""],
    ["2026-04-23", "Facebook", "实用贴", "马来西亚进口流程图解",
     "图文并茂展示进口流程，价值内容", "⬜ 待写", ""],
    ["2026-04-25", "LinkedIn", "客户故事", "客户好评分享",
     "分享客户正面反馈，建立口碑", "⬜ 待写", ""],
    # Week 4
    ["2026-04-28", "LinkedIn", "市场动态", "东马物流特殊政策解读",
     "沙巴砂拉越的特殊进口政策", "⬜ 待写", ""],
    ["2026-04-30", "Facebook", "节日/热点", "劳动节特别内容",
     "结合节日热点发布内容", "⬜ 待写", ""],
]

for r, data in enumerate(cal_samples, 5):
    for c, val in enumerate(data, 1):
        cell = ws_cal.cell(row=r, column=c, value=val)
        cell.font = normal_font
        cell.border = thin_border
        cell.alignment = center if c not in [4, 5, 7] else left_wrap
    ws_cal.cell(row=r, column=1).fill = input_fill
    ws_cal.cell(row=r, column=6).fill = input_fill

# Dropdowns
dv_platform = DataValidation(type="list", formula1='"LinkedIn,Facebook,Instagram,微信公众号,小红书,TikTok"', allow_blank=True)
ws_cal.add_data_validation(dv_platform)
dv_platform.add('B5:B50')

dv_type = DataValidation(type="list", formula1='"行业知识,案例分享,市场动态,个人故事,互动话题,实用贴,客户故事,产品/服务介绍,节日/热点,视频"', allow_blank=True)
ws_cal.add_data_validation(dv_type)
dv_type.add('C5:C50')

dv_status = DataValidation(type="list", formula1='"⬜ 待写,📝 草稿,👀 审核中,✅ 已发布,❌ 已废弃"', allow_blank=True)
ws_cal.add_data_validation(dv_status)
dv_status.add('F5:F50')

cal_widths = [12, 12, 12, 22, 35, 12, 25]
for i, w in enumerate(cal_widths, 1):
    ws_cal.column_dimensions[get_column_letter(i)].width = w
ws_cal.freeze_panes = 'A5'

# ============================================================
# Sheet 2: LinkedIn 帖子模板
# ============================================================
ws_li = wb.create_sheet("LinkedIn 模板")

ws_li.merge_cells('A1:C1')
ws_li.cell(row=1, column=1, value="💼 LinkedIn 帖子模板").font = title_font

li_headers = ["类型 Type", "标题/钩子 Hook", "完整模板 Full Template"]
for col, h in enumerate(li_headers, 1):
    cell = ws_li.cell(row=3, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = center

li_templates = [
    ["行业知识\nEducational",
     "90%的人不知道的海运冷知识 ⬇️",
     "90%的人不知道的海运冷知识 ⬇️\n\n"
     "做了X年货运，分享几个实用的小知识：\n\n"
     "1️⃣ 海运计费是取「实际重量」和「体积重量」中较大的那个\n"
     "2️⃣ 1 CBM = 167 KG（海运标准换算）\n"
     "3️⃣ 40HQ 比 40GP 多装 10% 的货，但价格差不多\n"
     "4️⃣ LCL 拼柜有个最低收费，通常 1-2 CBM 起\n\n"
     "💡 省钱小技巧：\n"
     "货量在 15-18 CBM 之间时，拼一个 20GP 可能比 LCL 更划算\n\n"
     "你还有什么想了解的？评论区告诉我 👇\n\n"
     "#FreightForwarding #Logistics #ImportExport #Malaysia"],

    ["案例分享\nCase Study",
     "客户说：这批货3天内必须到！我们这样搞定的 👇",
     "客户说：这批货3天内必须到！我们这样搞定的 👇\n\n"
     "上周收到一个紧急询价：\n"
     "📦 500kg 电子配件\n"
     "📍 深圳 → 吉隆坡\n"
     "⏰ 3天内必须到达\n\n"
     "方案对比：\n"
     "❌ 海运：7-12天，来不及\n"
     "❌ 普通空运：2-3天，但仓位紧张\n"
     "✅ 包机/快航：1-2天，价格高但能赶上\n\n"
     "最终方案：走次日达空运 + 预清关\n"
     "客户第2天就收到了货 ✅\n\n"
     "关键时刻，靠谱的物流伙伴很重要 🤝\n\n"
     "#Logistics #AirFreight #Emergency #Malaysia"],

    ["市场动态\nMarket Update",
     "2026年Q2 运价走势，外贸人必看 📊",
     "2026年Q2 运价走势，外贸人必看 📊\n\n"
     "中国 → 马来西亚 近期运价参考：\n\n"
     "🚢 海运 FCL：\n"
     "• 20GP: USD XXX - XXX\n"
     "• 40HQ: USD XXX - XXX\n"
     "趋势：[上涨/平稳/下跌]，原因是...\n\n"
     "✈️ 空运：\n"
     "• USD X.X - X.X /kg\n"
     "趋势：...\n\n"
     "💡 建议：\n"
     "• 有出货计划的尽早订舱\n"
     "• Q2 是旺季前，价格可能...\n\n"
     "需要最新报价？DM 我 📩\n\n"
     "#ShippingRates #FreightForwarding #Trade"],

    ["个人故事\nPersonal",
     "从打工到创业，我的 Freight Forwarder 之路",
     "从打工到创业，我的 Freight Forwarder 之路 🚢\n\n"
     "X年前，我只是一个普通的货代业务员...\n\n"
     "[你的故事：为什么入行、遇到的困难、学到的东西]\n\n"
     "创业后最大的感悟：\n"
     "1. 人脉就是一切\n"
     "2. 服务比价格更重要\n"
     "3. 数字化工具让一个人也能做大团队的事\n\n"
     "现在我用 AI 工具管理报价、客户、文档...\n"
     "效率比以前高了 10 倍 💪\n\n"
     "如果你也在创业路上，一起加油！\n\n"
     "#Entrepreneur #FreightForwarding #Startup #Malaysia"],

    ["互动话题\nEngagement",
     "做外贸/进口的朋友们，你们选 Freight Forwarder 最看重什么？",
     "做外贸/进口的朋友们，你们选 Freight Forwarder 最看重什么？🤔\n\n"
     "A. 价格最低\n"
     "B. 服务稳定可靠\n"
     "C. 沟通响应快\n"
     "D. 一站式全包\n\n"
     "说实话，大多数人选 B 和 C 😊\n\n"
     "因为物流这行，便宜不等于好。\n"
     "船期延误、清关出问题、货损... \n"
     "有一个靠谱的代理，关键时刻能救命。\n\n"
     "你选什么？评论区见 👇\n\n"
     "#FreightForwarding #ImportExport #Logistics"],
]

for r, data in enumerate(li_templates, 4):
    ws_li.cell(row=r, column=1, value=data[0]).font = bold_font
    ws_li.cell(row=r, column=1).border = thin_border
    ws_li.cell(row=r, column=1).alignment = center
    ws_li.cell(row=r, column=2, value=data[1]).font = Font(name='Arial', bold=True, size=11)
    ws_li.cell(row=r, column=2).border = thin_border
    ws_li.cell(row=r, column=2).alignment = left_wrap
    ws_li.cell(row=r, column=3, value=data[2]).font = normal_font
    ws_li.cell(row=r, column=3).border = thin_border
    ws_li.cell(row=r, column=3).alignment = left_wrap
    ws_li.row_dimensions[r].height = 200

li_widths = [14, 30, 65]
for i, w in enumerate(li_widths, 1):
    ws_li.column_dimensions[get_column_letter(i)].width = w

# ============================================================
# Sheet 3: Facebook 模板
# ============================================================
ws_fb = wb.create_sheet("Facebook 模板")

ws_fb.merge_cells('A1:C1')
ws_fb.cell(row=1, column=1, value="📘 Facebook 帖子模板").font = title_font

for col, h in enumerate(li_headers, 1):
    cell = ws_fb.cell(row=3, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = center

fb_templates = [
    ["促销/服务介绍\nPromo",
     "🚢 中国→马来西亚 海运空运，找我就对了！",
     "🚢 中国→马来西亚 海运空运\n\n"
     "✅ 海运整柜 FCL (20GP/40GP/40HQ)\n"
     "✅ 海运拼柜 LCL\n"
     "✅ 空运\n"
     "✅ 东马转运 (Kuching/KK/Miri/Sibu/Bintulu)\n\n"
     "📋 服务包括：\n"
     "• 订舱 • 报关 • 拖车 • 保险\n"
     "• 全程跟踪 • 到门配送\n\n"
     "💰 价格透明，没有隐藏费用\n"
     "⚡ 快速报价，5分钟回复\n\n"
     "PM 我获取最新运价 📩"],

    ["实用知识\nTips",
     "📦 第一次进口到马来西亚？看这篇就够了！",
     "📦 第一次进口到马来西亚？看这篇就够了！\n\n"
     "进口流程 5 步走：\n\n"
     "1️⃣ 确认货品 & HS Code\n"
     "→ 决定关税税率\n\n"
     "2️⃣ 找 Freight Forwarder 报价\n"
     "→ 海运/空运，全包价\n\n"
     "3️⃣ 安排出货\n"
     "→ 订舱 → 装柜 → 发货\n\n"
     "4️⃣ 清关\n"
     "→ 需要：Invoice, Packing List, B/L\n\n"
     "5️⃣ 提货/派送\n"
     "→ 清关完成后安排运输\n\n"
     "有疑问？评论区问我 👇"],

    ["客户好评\nTestimonial",
     "客户说：这是我合作过最靠谱的货代！",
     "客户说：这是我合作过最靠谱的货代！\n\n"
     "⭐⭐⭐⭐⭐\n\n"
     "\"之前换了3个货代，不是价格不透明就是服务跟不上。\n"
     "自从找了 [你的名字]，再也不用担心物流问题了。\n"
     "报价快、价格公道、有问必答。强烈推荐！\"\n\n"
     "—— [客户名], [公司]\n\n"
     "感谢客户的信任！我会继续做好服务 🤝\n\n"
     "需要货运服务？PM 我 📩"],
]

for r, data in enumerate(fb_templates, 4):
    ws_fb.cell(row=r, column=1, value=data[0]).font = bold_font
    ws_fb.cell(row=r, column=1).border = thin_border
    ws_fb.cell(row=r, column=1).alignment = center
    ws_fb.cell(row=r, column=2, value=data[1]).font = Font(name='Arial', bold=True, size=11)
    ws_fb.cell(row=r, column=2).border = thin_border
    ws_fb.cell(row=r, column=2).alignment = left_wrap
    ws_fb.cell(row=r, column=3, value=data[2]).font = normal_font
    ws_fb.cell(row=r, column=3).border = thin_border
    ws_fb.cell(row=r, column=3).alignment = left_wrap
    ws_fb.row_dimensions[r].height = 180

for i, w in enumerate(li_widths, 1):
    ws_fb.column_dimensions[get_column_letter(i)].width = w

# ============================================================
# Sheet 4: 行业素材库 Content Bank
# ============================================================
ws_bank = wb.create_sheet("素材库 Content Bank")

ws_bank.merge_cells('A1:D1')
ws_bank.cell(row=1, column=1, value="📚 行业素材库 Content Bank").font = title_font

bank_headers = ["类别 Category", "素材标题 Title", "内容/要点 Content", "可用于 For"]
for col, h in enumerate(bank_headers, 1):
    cell = ws_bank.cell(row=3, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = center

bank_items = [
    ["行业数据", "马来西亚 Top 10 进口商品",
     "1. 电子产品 2. 机械设备 3. 石油产品 4. 塑料制品 5. 钢铁\n"
     "6. 化学品 7. 车辆 8. 橡胶 9. 纺织品 10. 粮食",
     "LinkedIn, Facebook"],
    ["行业数据", "中国是马来西亚最大贸易伙伴",
     "2025年双边贸易额超 USD 100 billion\n"
     "主要进口：电子元件、机械设备、纺织品、钢铁",
     "LinkedIn"],
    ["港口信息", "马来西亚主要港口",
     "西马：Port Klang, Port of Tanjung Pelepas, Penang Port\n"
     "东马：Kuching Port, Kota Kinabalu Port, Bintulu Port, Miri Port",
     "Facebook, 客户教育"],
    ["流程知识", "海运 vs 空运选择指南",
     "海运：>500kg or >1CBM, 不急, 省钱\n"
     "空运：<500kg or 急件, 高价值货物\n"
     "Breakpoint: 约 100-150kg 空运更划算",
     "LinkedIn, Facebook"],
    ["流程知识", "东马转运注意事项",
     "1. 必须经 Port Klang 中转\n"
     "2. 东马属不同关税区\n"
     "3. 额外 3-5 天转运时间\n"
     "4. 需要两地报关",
     "LinkedIn, 客户教育"],
    ["法规更新", "马来西亚进口关税",
     "0-60% 不等，取决于 HS Code\n"
     "Sales Tax: 5-10%\n"
     "Service Tax: 6%\n"
     "FTA 优惠：中国-东盟 ACFTA 可减免",
     "LinkedIn"],
    ["热点话题", "红海危机对运价的影响",
     "2024-2025 红海危机绕道好望角\n"
     "航程增加 10-15 天，运价上涨 30-50%\n"
     "关注最新动态",
     "LinkedIn, Facebook"],
    ["省钱技巧", "如何降低物流成本",
     "1. 合并发货（consolidate）\n"
     "2. 提前订舱\n"
     "3. 比较多家报价\n"
     "4. 优化包装减少体积\n"
     "5. 利用 FTA 减免关税",
     "LinkedIn, Facebook"],
]

for r, data in enumerate(bank_items, 4):
    for c, val in enumerate(data, 1):
        cell = ws_bank.cell(row=r, column=c, value=val)
        cell.font = normal_font
        cell.border = thin_border
        cell.alignment = left_wrap
    ws_bank.row_dimensions[r].height = 60

bank_widths = [14, 25, 50, 20]
for i, w in enumerate(bank_widths, 1):
    ws_bank.column_dimensions[get_column_letter(i)].width = w

# ============================================================
# Sheet 5: Hashtag & 最佳发布时间
# ============================================================
ws_misc = wb.create_sheet("Hashtag & 时间")

ws_misc.merge_cells('A1:B1')
ws_misc.cell(row=1, column=1, value="#️⃣ Hashtag 库 & 最佳发布时间").font = title_font

ws_misc.cell(row=3, column=1, value="📌 LinkedIn 常用 Hashtags").font = big_bold
hashtags = [
    "#FreightForwarding", "#Logistics", "#Shipping", "#ImportExport",
    "#SupplyChain", "#Malaysia", "#Trade", "#Freight",
    "#ContainerShipping", "#AirFreight", "#CustomsClearance",
    "#CargoShipping", "#InternationalTrade", "#Ecommerce",
    "#Entrepreneur", "#SmallBusiness", "#Solopreneur"
]
for i, tag in enumerate(hashtags):
    ws_misc.cell(row=4 + i, column=1, value=tag).font = normal_font
    ws_misc.cell(row=4 + i, column=1).border = thin_border

ws_misc.cell(row=3, column=3, value="📌 Facebook 常用 Hashtags").font = big_bold
fb_hashtags = [
    "#FreightForwarding", "#MalaysiaShipping", "#SeaFreight",
    "#AirFreight", "#ImportMalaysia", "#LogisticsMY",
    "#ChinaToMalaysia", "#CargoServices", "#ShippingRates"
]
for i, tag in enumerate(fb_hashtags):
    ws_misc.cell(row=4 + i, column=3, value=tag).font = normal_font
    ws_misc.cell(row=4 + i, column=3).border = thin_border

ws_misc.cell(row=22, column=1, value="⏰ 最佳发布时间 Best Posting Times").font = big_bold
times = [
    ["LinkedIn", "周二/三/四", "8:00-9:00 AM / 12:00-1:00 PM", "MYT"],
    ["Facebook", "周三/四/五", "1:00-4:00 PM", "MYT"],
    ["Instagram", "周二/三", "11:00 AM-1:00 PM", "MYT"],
]
time_headers = ["平台", "最佳日期", "最佳时间", "时区"]
for col, h in enumerate(time_headers, 1):
    cell = ws_misc.cell(row=23, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = center

for r, data in enumerate(times, 24):
    for c, val in enumerate(data, 1):
        cell = ws_misc.cell(row=r, column=c, value=val)
        cell.font = normal_font
        cell.border = thin_border
        cell.alignment = center

ws_misc.column_dimensions['A'].width = 25
ws_misc.column_dimensions['B'].width = 25
ws_misc.column_dimensions['C'].width = 28
ws_misc.column_dimensions['D'].width = 8

wb.save('/root/.openclaw/workspace/freight-agent/Social_Media_Agent.xlsx')
print("✅ Social_Media_Agent.xlsx created")
