"""
Freight Forwarder Quotation System for Brandon
Creates: Rate Database + Quotation Template
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from datetime import datetime

# ============================================================
# PART 1: Rate Database (运价数据库)
# ============================================================
wb_rates = openpyxl.Workbook()

# --- Colors & Styles ---
header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
sub_header_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
sub_header_font = Font(name='Arial', bold=True, size=10)
normal_font = Font(name='Arial', size=10)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

def style_header(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

def style_row(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = normal_font
        cell.alignment = center_align
        cell.border = thin_border

# ===== Sheet 1: Sea Freight FCL (海运整柜) =====
ws_fcl = wb_rates.active
ws_fcl.title = "Sea FCL"

fcl_headers = [
    "No.", "供应商\nSupplier", "航线\nRoute", "POL\n装货港", "POD\n卸货港",
    "柜型\nContainer", "20GP\n(USD)", "40GP\n(USD)", "40HQ\n(USD)",
    "有效期\nValid Until", "备注\nRemarks"
]
for col, h in enumerate(fcl_headers, 1):
    ws_fcl.cell(row=1, column=col, value=h)
style_header(ws_fcl, 1, len(fcl_headers))

# Sample data
fcl_samples = [
    [1, "Supplier A", "China → MY", "Shanghai", "Port Klang", "GP/HQ", 350, 550, 580, "2026-04-30", "每周二截关"],
    [2, "Supplier A", "China → MY", "Ningbo", "Port Klang", "GP/HQ", 380, 580, 610, "2026-04-30", ""],
    [3, "Supplier B", "China → MY", "Shenzhen", "Port Klang", "GP/HQ", 320, 500, 530, "2026-05-15", "含THC"],
    [4, "Supplier B", "China → MY", "Guangzhou", "Port Klang", "GP/HQ", 340, 520, 550, "2026-05-15", ""],
    [5, "Supplier A", "China → East MY", "Shanghai", "Kuching", "GP/HQ", 550, 800, 850, "2026-04-30", "经PKG中转"],
    [6, "Supplier B", "China → East MY", "Shenzhen", "Kuching", "GP/HQ", 520, 780, 820, "2026-05-15", "经PKG中转"],
    [7, "", "", "", "", "", "", "", "", "", ""],
    [8, "", "", "", "", "", "", "", "", "", ""],
]
for r, data in enumerate(fcl_samples, 2):
    for c, val in enumerate(data, 1):
        ws_fcl.cell(row=r, column=c, value=val)
    style_row(ws_fcl, r, len(fcl_headers))

# Column widths
fcl_widths = [5, 15, 15, 12, 12, 12, 12, 12, 12, 14, 20]
for i, w in enumerate(fcl_widths, 1):
    ws_fcl.column_dimensions[get_column_letter(i)].width = w

# ===== Sheet 2: Sea Freight LCL (海运拼柜) =====
ws_lcl = wb_rates.create_sheet("Sea LCL")

lcl_headers = [
    "No.", "供应商\nSupplier", "航线\nRoute", "POL\n装货港", "POD\n卸货港",
    "计费方式\nBasis", "费率\nRate (USD)", "最低收费\nMin (USD)",
    "有效期\nValid Until", "备注\nRemarks"
]
for col, h in enumerate(lcl_headers, 1):
    ws_lcl.cell(row=1, column=col, value=h)
style_header(ws_lcl, 1, len(lcl_headers))

lcl_samples = [
    [1, "Supplier A", "China → MY", "Shanghai", "Port Klang", "Per CBM", 35, 80, "2026-04-30", ""],
    [2, "Supplier A", "China → MY", "Ningbo", "Port Klang", "Per CBM", 38, 85, "2026-04-30", ""],
    [3, "Supplier B", "China → MY", "Shenzhen", "Port Klang", "Per CBM", 30, 75, "2026-05-15", "含THC"],
    [4, "", "", "", "", "", "", "", "", ""],
]
for r, data in enumerate(lcl_samples, 2):
    for c, val in enumerate(data, 1):
        ws_lcl.cell(row=r, column=c, value=val)
    style_row(ws_lcl, r, len(lcl_headers))

lcl_widths = [5, 15, 15, 12, 12, 12, 12, 12, 14, 20]
for i, w in enumerate(lcl_widths, 1):
    ws_lcl.column_dimensions[get_column_letter(i)].width = w

# ===== Sheet 3: Air Freight (空运) =====
ws_air = wb_rates.create_sheet("Air Freight")

air_headers = [
    "No.", "供应商\nSupplier", "航线\nRoute", "Origin\n出发地", "Dest\n目的地",
    "计费方式\nBasis", "费率\nRate (USD/kg)",
    "最低收费\nMin (USD)", "有效期\nValid Until", "备注\nRemarks"
]
for col, h in enumerate(air_headers, 1):
    ws_air.cell(row=1, column=col, value=h)
style_header(ws_air, 1, len(air_headers))

air_samples = [
    [1, "Supplier A", "China → MY", "PVG", "KUL", "Per KG", 3.5, 50, "2026-04-30", ""],
    [2, "Supplier A", "China → MY", "CAN", "KUL", "Per KG", 3.2, 50, "2026-04-30", ""],
    [3, "Supplier B", "China → MY", "SZX", "KUL", "Per KG", 3.0, 45, "2026-05-15", "含燃油"],
    [4, "", "", "", "", "", "", "", "", ""],
]
for r, data in enumerate(air_samples, 2):
    for c, val in enumerate(data, 1):
        ws_air.cell(row=r, column=c, value=val)
    style_row(ws_air, r, len(air_headers))

air_widths = [5, 15, 15, 10, 10, 12, 15, 12, 14, 20]
for i, w in enumerate(air_widths, 1):
    ws_air.column_dimensions[get_column_letter(i)].width = w

# ===== Sheet 4: Surcharges (附加费) =====
ws_sur = wb_rates.create_sheet("Surcharges")

sur_headers = [
    "No.", "费用项目\nCharge", "海运FCL\nSea FCL", "海运LCL\nSea LCL",
    "空运\nAir", "备注\nRemarks"
]
for col, h in enumerate(sur_headers, 1):
    ws_sur.cell(row=1, column=col, value=h)
style_header(ws_sur, 1, len(sur_headers))

sur_samples = [
    [1, "THC 码头费", "Included / USD 150", "Included", "-", "Per container"],
    [2, "DOC 文件费", "USD 50", "USD 50", "USD 50", "Per shipment"],
    [3, "BAF 燃油附加费", "Varies", "Varies", "Included", "Check monthly"],
    [4, "ISPS 安保费", "USD 15", "USD 15", "-", "Per container"],
    [5, "报关费 Customs", "MYR 200", "MYR 200", "MYR 200", "Per shipment"],
    [6, "拖车费 Trucking", "Check route", "Check route", "-", "PKL vs East MY"],
    [7, "保险 Insurance", "0.3% cargo value", "0.3% cargo value", "0.3% cargo value", "Optional"],
]
for r, data in enumerate(sur_samples, 2):
    for c, val in enumerate(data, 1):
        ws_sur.cell(row=r, column=c, value=val)
    style_row(ws_sur, r, len(sur_headers))

sur_widths = [5, 20, 18, 18, 18, 20]
for i, w in enumerate(sur_widths, 1):
    ws_sur.column_dimensions[get_column_letter(i)].width = w

# Save rate database
wb_rates.save('/root/.openclaw/workspace/freight-agent/Rate_Database.xlsx')
print("✅ Rate_Database.xlsx created")

# ============================================================
# PART 2: Quotation Template (报价单模板)
# ============================================================
wb_quote = openpyxl.Workbook()
ws_q = wb_quote.active
ws_q.title = "Quotation"

# Company header
ws_q.merge_cells('A1:F1')
ws_q.cell(row=1, column=1, value="[YOUR COMPANY NAME]")
ws_q.cell(row=1, column=1).font = Font(name='Arial', bold=True, size=16, color='2F5496')

ws_q.merge_cells('A2:F2')
ws_q.cell(row=2, column=1, value="Freight Forwarding Services")
ws_q.cell(row=2, column=1).font = Font(name='Arial', size=11, color='666666')

# Quote info
info_rows = [
    (4, "报价单号 Quote No:", "[QT-2026-001]"),
    (5, "日期 Date:", datetime.now().strftime("%Y-%m-%d")),
    (6, "有效期 Valid Until:", "[2026-05-05]"),
    (7, "业务员 Sales:", "[Your Name]"),
]
for row, label, val in info_rows:
    ws_q.cell(row=row, column=1, value=label).font = Font(name='Arial', bold=True, size=10)
    ws_q.cell(row=row, column=3, value=val).font = normal_font

# Customer info
ws_q.cell(row=9, column=1, value="客户信息 Customer Info").font = Font(name='Arial', bold=True, size=12, color='2F5496')
cust_rows = [
    (10, "公司 Company:", "[Customer Company]"),
    (11, "联系人 Contact:", "[Contact Person]"),
    (12, "电话 Phone:", "[Phone]"),
    (13, "邮箱 Email:", "[Email]"),
]
for row, label, val in cust_rows:
    ws_q.cell(row=row, column=1, value=label).font = Font(name='Arial', bold=True, size=10)
    ws_q.cell(row=row, column=3, value=val).font = normal_font

# Shipment details
ws_q.cell(row=15, column=1, value="货物信息 Shipment Details").font = Font(name='Arial', bold=True, size=12, color='2F5496')

ship_headers = ["项目 Item", "详情 Details"]
for col, h in enumerate(ship_headers, 1):
    cell = ws_q.cell(row=16, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = center_align

ship_items = [
    ("运输方式 Mode", "[Sea FCL / Sea LCL / Air]"),
    ("起运地 POL", "[Shanghai, China]"),
    ("目的地 POD", "[Port Klang / Kuching / Kota Kinabalu]"),
    ("柜型 Container", "[20GP / 40GP / 40HQ / LCL CBM / Air KG]"),
    ("货品 Cargo", "[Description]"),
    ("货值 Cargo Value", "[USD ...]"),
    ("HS Code", "[If applicable]"),
]
for r, (item, detail) in enumerate(ship_items, 17):
    ws_q.cell(row=r, column=1, value=item).font = normal_font
    ws_q.cell(row=r, column=1).border = thin_border
    ws_q.cell(row=r, column=2, value=detail).font = normal_font
    ws_q.cell(row=r, column=2).border = thin_border

# Cost breakdown
cost_start = 25
ws_q.cell(row=cost_start, column=1, value="费用明细 Cost Breakdown").font = Font(name='Arial', bold=True, size=12, color='2F5496')

cost_headers = ["费用项目 Charge", "币种 Currency", "金额 Amount", "备注 Remarks"]
for col, h in enumerate(cost_headers, 1):
    cell = ws_q.cell(row=cost_start + 1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = center_align

cost_items = [
    ["海运费 Ocean Freight", "USD", "", ""],
    ["THC 码头费", "USD", "", ""],
    ["DOC 文件费", "USD", "", ""],
    ["报关费 Customs", "MYR", "", ""],
    ["拖车费 Trucking", "MYR", "", "PKG to destination"],
    ["保险 Insurance", "USD", "", "Optional"],
    ["其他 Others", "", "", ""],
]
for r, data in enumerate(cost_items, cost_start + 2):
    for c, val in enumerate(data, 1):
        cell = ws_q.cell(row=r, column=c, value=val)
        cell.font = normal_font
        cell.border = thin_border
        cell.alignment = center_align

# Total row
total_row = cost_start + 2 + len(cost_items)
ws_q.merge_cells(f'A{total_row}:B{total_row}')
ws_q.cell(row=total_row, column=1, value="合计 Total").font = Font(name='Arial', bold=True, size=11)
ws_q.cell(row=total_row, column=1).border = thin_border
ws_q.cell(row=total_row, column=3).border = thin_border
ws_q.cell(row=total_row, column=3).font = Font(name='Arial', bold=True, size=11)
ws_q.cell(row=total_row, column=4).border = thin_border

# Terms
terms_start = total_row + 2
ws_q.cell(row=terms_start, column=1, value="条款 Terms & Conditions").font = Font(name='Arial', bold=True, size=12, color='2F5496')
terms = [
    "1. 以上报价有效期为7天，过期需重新确认",
    "2. 报价不含海关查验费、仓储费等意外费用",
    "3. 如遇船公司/航空公司调价，以实际为准",
    "4. 运费到付或预付请提前说明",
    "5. 货物保险建议购买，未购买保险货损自负",
]
for i, t in enumerate(terms):
    ws_q.cell(row=terms_start + 1 + i, column=1, value=t).font = normal_font

# Contact
contact_row = terms_start + 1 + len(terms) + 1
ws_q.merge_cells(f'A{contact_row}:F{contact_row}')
ws_q.cell(row=contact_row, column=1, value="如有任何疑问请随时联系 | For any inquiries, please don't hesitate to contact us")
ws_q.cell(row=contact_row, column=1).font = Font(name='Arial', italic=True, size=10, color='666666')
ws_q.cell(row=contact_row, column=1).alignment = Alignment(horizontal='center')

# Column widths
ws_q.column_dimensions['A'].width = 25
ws_q.column_dimensions['B'].width = 25
ws_q.column_dimensions['C'].width = 20
ws_q.column_dimensions['D'].width = 20
ws_q.column_dimensions['E'].width = 15
ws_q.column_dimensions['F'].width = 15

# Save
wb_quote.save('/root/.openclaw/workspace/freight-agent/Quotation_Template.xlsx')
print("✅ Quotation_Template.xlsx created")
print("\n🎉 All files created in /root/.openclaw/workspace/freight-agent/")
