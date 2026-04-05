"""
Sales Hunter Agent - 主动获客引擎
目标：马来西亚中小企业进口商/出口商
渠道：LinkedIn、Google 搜索、SEO/AEO
输出：存入 CRM → 自动发 WhatsApp/邮件 → 生成跟进任务
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime

# ============================================================
# Styles
# ============================================================
header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
input_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
blue_fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
grey_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
orange_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
purple_fill = PatternFill(start_color='E2D0F0', end_color='E2D0F0', fill_type='solid')
bold_font = Font(name='Arial', bold=True, size=10)
big_bold = Font(name='Arial', bold=True, size=13, color='2F5496')
title_font = Font(name='Arial', bold=True, size=18, color='2F5496')
small_font = Font(name='Arial', size=9, color='666666')
normal_font = Font(name='Arial', size=10)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_wrap = Alignment(horizontal='left', vertical='center', wrap_text=True)

wb = openpyxl.Workbook()

# ============================================================
# Sheet 1: Lead Database 全部 Leads
# ============================================================
ws1 = wb.active
ws1.title = "Lead Database"

ws1.merge_cells('A1:S1')
ws1.cell(row=1, column=1, value="🎯 Sales Hunter — Lead Database").font = title_font
ws1.row_dimensions[1].height = 35

ws1.merge_cells('A2:S2')
ws1.cell(row=2, column=1,
    value="从 LinkedIn/Google/展会 等渠道找到的潜在客户 → 评分 → 外展 → 跟进").font = small_font

headers1 = [
    "No.",                       # A
    "Company Name\n公司名",       # B
    "Industry\n行业",             # C
    "Company Size\n公司规模",      # D
    "Website\n网站",              # E
    "LinkedIn\n领英",             # F
    "Contact Person\n联系人",      # G
    "Position\n职位",             # H
    "Phone/WhatsApp\n电话",        # I
    "Email\n邮箱",                # J
    "City/Country\n城市/国家",     # K
    "Import/Export\n进出口",       # L
    "Source\n来源渠道",            # M
    "Lead Score\n评分",            # N
    "Status\n状态",               # O
    "First Contact\n首次联系",     # P
    "Last Contact\n上次联系",      # Q
    "Assigned To\n负责人",         # R
    "Notes\n备注",                # S
]

for col, h in enumerate(headers1, 1):
    cell = ws1.cell(row=4, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = center

# Dropdowns
source_options = (
    "LinkedIn,Google Search,Google SEO,AEO,"
    "Alibaba,Trade Show,Referral,Cold Call,"
    "Facebook,Instagram,Industry Directory,"
    "Customs Data,Other"
)
dv_source = DataValidation(type="list", formula1=f'"{source_options}"', allow_blank=True)
ws1.add_data_validation(dv_source)

status_options = (
    "New 新,Contacted 已联系,Interested 有兴趣,"
    "Quoted 已报价,Negotiating 谈判中,Won 已成交,"
    "Lost 已丢失,Dormant 休眠"
)
dv_status = DataValidation(type="list", formula1=f'"{status_options}"', allow_blank=True)
ws1.add_data_validation(dv_status)

imp_exp = "Importer 进口商,Exporter 出口商,Both 进出口"
dv_impexp = DataValidation(type="list", formula1=f'"{imp_exp}"', allow_blank=True)
ws1.add_data_validation(dv_impexp)

size_options = "Micro (1-10),Small (11-50),Medium (51-250),Unknown"
dv_size = DataValidation(type="list", formula1=f'"{size_options}"', allow_blank=True)
ws1.add_data_validation(dv_size)

for row in range(5, 55):
    ws1.cell(row=row, column=1, value=row - 4).font = normal_font
    ws1.cell(row=row, column=1).alignment = center
    for col in range(1, 20):
        ws1.cell(row=row, column=col).border = thin_border
        ws1.cell(row=row, column=col).font = normal_font
        if col in [2,3,4,5,6,7,8,9,10,11,12,13,14,18,19]:
            ws1.cell(row=row, column=col).fill = input_fill
    dv_source.add(ws1.cell(row=row, column=13))
    dv_status.add(ws1.cell(row=row, column=15))
    dv_impexp.add(ws1.cell(row=row, column=12))
    dv_size.add(ws1.cell(row=row, column=4))

col_widths_1 = [5, 20, 16, 14, 22, 22, 14, 14, 18, 22, 14, 12, 14, 10, 14, 14, 14, 10, 20]
for i, w in enumerate(col_widths_1, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

ws1.freeze_panes = 'A5'

# Lead Score explanation
ws1.cell(row=56, column=1, value="📊 Lead Score 评分标准:").font = bold_font
score_desc = [
    ("5 ⭐⭐⭐⭐⭐", "完美匹配 — 进出口 + 有明确需求 + 可联系到决策人"),
    ("4 ⭐⭐⭐⭐", "强匹配 — 进出口商 + 公司规模合适 + 有联系方式"),
    ("3 ⭐⭐⭐", "中等 — 行业相关但信息不全"),
    ("2 ⭐⭐", "弱 — 仅知道公司名，无联系方式"),
    ("1 ⭐", "待验证 — 需要更多信息"),
]
for i, (score, desc) in enumerate(score_desc):
    ws1.cell(row=57 + i, column=1, value=score).font = bold_font
    ws1.cell(row=57 + i, column=3, value=desc).font = small_font

# ============================================================
# Sheet 2: Google Search Queries 搜索指令库
# ============================================================
ws2 = wb.create_sheet("Google Dorks")

ws2.merge_cells('A1:D1')
ws2.cell(row=1, column=1, value="🔍 Google 高级搜索指令库 — 找马来西亚进出口商").font = title_font
ws2.row_dimensions[1].height = 35

ws2.merge_cells('A2:D2')
ws2.cell(row=2, column=1,
    value="复制搜索指令到 Google → 找到公司 → 填入 Lead Database").font = small_font

headers2 = ["Category 分类", "Search Query 搜索指令", "What It Finds 找什么", "Used 已用"]
for col, h in enumerate(headers2, 1):
    cell = ws2.cell(row=4, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = center

queries = [
    # 进口商搜索
    ["🇲🇾 MY 进口商",
     '"freight forwarder" OR "shipping agent" site:.my import',
     "马来西亚做进口的公司"],
    ["🇲🇾 MY 进口商",
     '"import" "Kuala Lumpur" OR "Penang" OR "Johor" "company" logistics',
     "吉隆坡/槟城/新山进口公司"],
    ["🇲🇾 MY 进口商",
     '"customs clearance" site:.my "import" OR "freight"',
     "做清关的马来西亚公司"],
    ["🇲🇾 MY 进口商",
     '"Malaysia importer" "China" OR "shipping" "contact"',
     "从中国进口的马来公司"],

    # 出口商搜索
    ["🇲🇾 MY 出口商",
     '"Malaysia exporter" "manufacturing" OR "factory" "contact us"',
     "马来西亚工厂/出口商"],
    ["🇲🇾 MY 出口商",
     '"export" site:.my "shipping" OR "freight" "company"',
     "马来西亚做出口的公司"],
    ["🇲🇾 MY 出口商",
     '"made in Malaysia" "wholesale" OR "distributor" "export"',
     "马来西亚批发/分销出口商"],

    # 电商 / 跨境电商
    ["🛒 电商卖家",
     '"Shopee seller" OR "Lazada seller" Malaysia "supplier"',
     "Shopee/Lazada 上的马来西亚卖家"],
    ["🛒 电商卖家",
     '"cross border ecommerce" Malaysia "logistics" OR "fulfillment"',
     "马来西亚跨境电商公司"],

    # LinkedIn 搜索
    ["🔗 LinkedIn",
     'site:linkedin.com "freight forwarder" OR "shipping" Malaysia',
     "LinkedIn 上的马来西亚货代从业者"],
    ["🔗 LinkedIn",
     'site:linkedin.com "import export" Malaysia "manager" OR "director"',
     "马来西亚进出口公司管理层"],
    ["🔗 LinkedIn",
     'site:linkedin.com "supply chain" Malaysia "logistics"',
     "马来西亚供应链/物流专业人士"],

    # 行业目录
    ["📁 行业目录",
     'site:yellowpages.my "freight" OR "shipping" OR "logistics"',
     "马来黄页上的货运公司"],
    ["📁 行业目录",
     'site:matrade.gov.my exporter directory',
     "马来西亚外贸发展局出口商目录"],
    ["📁 行业目录",
     'site:fmm.org.my manufacturer directory',
     "马来西亚制造商联合会会员"],

    # 本地 SEO 关键词
    ["🔎 SEO 关键词",
     '"freight forwarding Malaysia" OR "shipping agent KL"',
     "搜索量高的行业关键词（SEO 竞品分析）"],
    ["🔎 SEO 关键词",
     '"sea freight Malaysia" OR "air freight Malaysia" company',
     "海运/空运相关关键词"],
    ["🔎 AEO 优化",
     '"how to import from China to Malaysia" OR "shipping cost China Malaysia"',
     "客户常问的问题（用于 AEO 内容）"],
    ["🔎 AEO 优化",
     '"customs clearance Malaysia" OR "how long shipping China Malaysia"',
     "FAQ 类问题（抢 featured snippet）"],
]

for i, row_data in enumerate(queries):
    row = 5 + i
    for col, val in enumerate(row_data, 1):
        cell = ws2.cell(row=row, column=col, value=val)
        cell.font = normal_font
        cell.border = thin_border
        if col == 4:
            cell.fill = input_fill
        cell.alignment = left_wrap
    ws2.row_dimensions[row].height = 30

col_widths_2 = [18, 60, 35, 8]
for i, w in enumerate(col_widths_2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

ws2.freeze_panes = 'A5'

# ============================================================
# Sheet 3: SEO / AEO Strategy
# ============================================================
ws3 = wb.create_sheet("SEO AEO Strategy")

ws3.merge_cells('A1:F1')
ws3.cell(row=1, column=1, value="🔎 SEO & AEO 内容策略 — 让客户主动找你").font = title_font
ws3.row_dimensions[1].height = 35

ws3.merge_cells('A2:F2')
ws3.cell(row=2, column=1,
    value="SEO = 搜索引擎优化 | AEO = 答案引擎优化（抢 Google 精选摘要/ChatGPT 回答）").font = small_font

# Section 1: Target Keywords
ws3.cell(row=4, column=1, value="🎯 目标关键词 Target Keywords").font = big_bold

kw_headers = ["Keyword 关键词", "Search Intent 搜索意图", "Monthly Vol 月搜索量(估)", "Difficulty 难度", "Content Plan 内容计划", "Status 状态"]
for col, h in enumerate(kw_headers, 1):
    cell = ws3.cell(row=6, column=col, value=h)
    cell.font = header_font
    cell.fill = PatternFill(start_color='548235', end_color='548235', fill_type='solid')
    cell.border = thin_border
    cell.alignment = center

keywords = [
    ["freight forwarding Malaysia", "Commercial 商业", "300+", "Medium", "Landing Page 着陆页", ""],
    ["sea freight China to Malaysia", "Commercial 商业", "200+", "Low", "Guide 指南文章", ""],
    ["air freight Malaysia", "Commercial 商业", "150+", "Medium", "Landing Page 着陆页", ""],
    ["shipping from China to Malaysia cost", "Informational 信息", "500+", "Low", "Calculator 计算器页面", ""],
    ["customs clearance Malaysia", "Commercial 商业", "200+", "Medium", "Service Page 服务页", ""],
    ["how long shipping China to Malaysia", "Informational 信息", "400+", "Low", "FAQ Page FAQ页面", ""],
    ["LCL shipping Malaysia", "Commercial 商业", "100+", "Low", "Blog 博客文章", ""],
    ["import from China to Malaysia", "Informational 信息", "600+", "Medium", "Ultimate Guide 终极指南", ""],
    ["freight forwarder KL", "Local 本地", "100+", "Low", "Google My Business", ""],
    ["logistics company Penang", "Local 本地", "80+", "Low", "Google My Business", ""],
    ["cheapest shipping to Malaysia", "Transactional 交易", "300+", "Medium", "Comparison 对比文章", ""],
    ["3PL Malaysia", "Commercial 商业", "200+", "Medium", "Service Page 服务页", ""],
]

for i, row_data in enumerate(keywords):
    row = 7 + i
    for col, val in enumerate(row_data, 1):
        cell = ws3.cell(row=row, column=col, value=val)
        cell.font = normal_font
        cell.border = thin_border
        cell.fill = input_fill
        cell.alignment = left_wrap

# Section 2: AEO Questions (抢 Featured Snippet)
aeo_start = 21
ws3.cell(row=aeo_start, column=1, value="❓ AEO 问答优化 — 抢 Google 精选摘要").font = big_bold

aeo_headers = ["Question 客户问题", "Target Keyword 目标关键词", "Answer Format 回答格式", "Content 内容位置", "Status 状态"]
for col, h in enumerate(aeo_headers, 1):
    cell = ws3.cell(row=aeo_start + 2, column=col, value=h)
    cell.font = header_font
    cell.fill = PatternFill(start_color='7B2D8E', end_color='7B2D8E', fill_type='solid')
    cell.border = thin_border
    cell.alignment = center

aeo_questions = [
    ["How much does it cost to ship from China to Malaysia?",
     "shipping cost China Malaysia",
     "Numbered List + Table 列表+表格",
     "Blog / Calculator Page", ""],
    ["How long does shipping take from China to Malaysia?",
     "shipping time China Malaysia",
     "Short Paragraph 短段落",
     "FAQ Section", ""],
    ["What documents do I need to import to Malaysia?",
     "import documents Malaysia",
     "Numbered List 列表",
     "Guide Page", ""],
    ["What is the cheapest way to ship to Malaysia?",
     "cheapest shipping Malaysia",
     "Comparison Table 对比表",
     "Blog Post", ""],
    ["Do I need a freight forwarder to import from China?",
     "freight forwarder import China",
     "Yes/No + Explanation",
     "FAQ / Blog", ""],
    ["How to calculate shipping cost for LCL?",
     "LCL shipping cost calculator",
     "Formula + Calculator",
     "Tool Page", ""],
    ["What is the difference between FCL and LCL?",
     "FCL vs LCL",
     "Table Comparison 对比表",
     "Blog Post", ""],
    ["How to clear customs in Malaysia for imports?",
     "customs clearance Malaysia",
     "Step-by-Step 步骤",
     "Guide Page", ""],
]

for i, row_data in enumerate(aeo_questions):
    row = aeo_start + 3 + i
    for col, val in enumerate(row_data, 1):
        cell = ws3.cell(row=row, column=col, value=val)
        cell.font = normal_font
        cell.border = thin_border
        cell.fill = input_fill
        cell.alignment = left_wrap
    ws3.row_dimensions[row].height = 40

# Section 3: Content Calendar
cal_start = 33
ws3.cell(row=cal_start, column=1, value="📅 内容发布计划 Content Calendar").font = big_bold

cal_headers = ["Week 周", "Content 内容", "Type 类型", "Target Keyword", "Platform 平台", "Status 状态"]
for col, h in enumerate(cal_headers, 1):
    cell = ws3.cell(row=cal_start + 2, column=col, value=h)
    cell.font = header_font
    cell.fill = PatternFill(start_color='BF8F00', end_color='BF8F00', fill_type='solid')
    cell.border = thin_border
    cell.alignment = center

content_plan = [
    ["Week 1", "Import from China to Malaysia: Complete Guide", "Blog 篇", "import China Malaysia", "Website + LinkedIn", ""],
    ["Week 1", "Freight Forwarding KL — Who We Are", "Landing Page", "freight forwarder KL", "Website + GMB", ""],
    ["Week 2", "Shipping Cost Calculator: China → MY", "Tool 工具", "shipping cost calculator", "Website", ""],
    ["Week 2", "5 Things to Know About Customs in MY", "Blog 篇", "customs clearance MY", "Website + LinkedIn", ""],
    ["Week 3", "FCL vs LCL: Which is Right for You?", "Blog 篇", "FCL vs LCL", "Website + LinkedIn", ""],
    ["Week 3", "Sea Freight vs Air Freight: Cost Comparison", "Blog 篇", "sea vs air freight", "Website", ""],
    ["Week 4", "How to Track Your Shipment in Real Time", "FAQ FAQ", "shipment tracking", "Website", ""],
    ["Week 4", "Customer Case Study: Successful Import", "Case Study", "import success MY", "LinkedIn + FB", ""],
]

for i, row_data in enumerate(content_plan):
    row = cal_start + 3 + i
    for col, val in enumerate(row_data, 1):
        cell = ws3.cell(row=row, column=col, value=val)
        cell.font = normal_font
        cell.border = thin_border
        cell.fill = input_fill
        cell.alignment = left_wrap

col_widths_3 = [22, 40, 16, 22, 20, 10]
for i, w in enumerate(col_widths_3, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

# ============================================================
# Sheet 4: Outreach Templates 外展话术
# ============================================================
ws4 = wb.create_sheet("Outreach 外展话术")

ws4.merge_cells('A1:D1')
ws4.cell(row=1, column=1, value="📧 外展话术模板 — First Contact").font = title_font
ws4.row_dimensions[1].height = 35

ws4.merge_cells('A2:D2')
ws4.cell(row=2, column=1,
    value="找到 Lead → 选模板 → 替换 [变量] → 发送").font = small_font

headers4 = ["Channel 渠道", "Scenario 场景", "Message 消息模板 (EN)", "Message 消息模板 (CN)"]
for col, h in enumerate(headers4, 1):
    cell = ws4.cell(row=4, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = center

outreach = [
    [
        "WhatsApp",
        "LinkedIn 找到的 Lead",
        "Hi [Name],\n\n"
        "I found your profile on LinkedIn and noticed you're in the "
        "[Industry] space in [City].\n\n"
        "We specialize in freight forwarding between China and Malaysia "
        "— sea & air freight, customs clearance, door-to-door delivery.\n\n"
        "Would you be open to a quick chat about your shipping needs? "
        "Happy to share a no-obligation quote. 🚢",

        "Hi [Name]，\n\n"
        "在 LinkedIn 上看到您的资料，了解到您在 [City] 做 [Industry]。\n\n"
        "我们专做中马货运 — 海运/空运、清关、门到门派送。\n\n"
        "方便聊聊您的运输需求吗？可以免费给您报个价 🚢"
    ],
    [
        "Email",
        "Google 搜索找到的公司",
        "Subject: Quick Question About Your Shipping Needs\n\n"
        "Hi [Name],\n\n"
        "I came across [Company] while researching top [Industry] "
        "companies in Malaysia. Impressive work!\n\n"
        "We help businesses like yours streamline imports from China — "
        "competitive rates, full customs clearance, real-time tracking.\n\n"
        "Would you be interested in a 10-minute call to see if we can "
        "save you money on your current shipping?\n\n"
        "Best regards,\n"
        "[Your Name]\n"
        "[Your Company]",

        "Subject: 关于您的运输需求\n\n"
        "Hi [Name]，\n\n"
        "在搜索马来西亚 [Industry] 公司时找到了 [Company]，很棒的业务！\n\n"
        "我们帮像您这样的企业简化从中国的进口流程 — 优惠运价、全程清关、实时追踪。\n\n"
        "有兴趣花 10 分钟聊聊，看看能不能帮您省运费吗？\n\n"
        "此致敬礼，\n[Your Name]\n[Your Company]"
    ],
    [
        "WhatsApp",
        "阿里巴巴上找到的卖家",
        "Hi [Name],\n\n"
        "I saw your products on Alibaba and thought your business might "
        "need a reliable shipping partner for Malaysia.\n\n"
        "We offer:\n"
        "🚢 Sea freight: China → MY, from [Price]/CBM\n"
        "✈️ Air freight: from [Price]/KG\n"
        "🛃 Full customs clearance\n"
        "📍 Door-to-door delivery\n\n"
        "Want me to send you our rate card?",

        "Hi [Name]，\n\n"
        "在阿里巴巴上看到您的产品，觉得您的业务可能需要一个靠谱的马来西亚运输伙伴。\n\n"
        "我们提供：\n"
        "🚢 海运：中国→马来西亚，[Price] 起/CBM\n"
        "✈️ 空运：[Price] 起/KG\n"
        "🛃 全程清关\n"
        "📍 门到门派送\n\n"
        "需要我发一份价目表给您吗？"
    ],
    [
        "LinkedIn DM",
        "同行业有人脉的人",
        "Hi [Name],\n\n"
        "I noticed we're both connected to [Mutual Connection] and "
        "you're in the [Industry] industry.\n\n"
        "I run a freight forwarding business focused on China-Malaysia "
        "routes. Always looking to connect with fellow professionals.\n\n"
        "Would love to add you to my network! 🤝",

        "Hi [Name]，\n\n"
        "看到我们共同连接了 [Mutual Connection]，而且您也在 [Industry] 行业。\n\n"
        "我做中马货运代理，很乐意认识同行的朋友。\n\n"
        "方便加个好友吗？🤝"
    ],
    [
        "Email",
        "参加展会的公司",
        "Subject: Great Meeting You at [Event]!\n\n"
        "Hi [Name],\n\n"
        "It was great meeting you at [Trade Show] last week. "
        "I enjoyed our conversation about [Topic].\n\n"
        "As discussed, I'm sending over our rate card for China-Malaysia "
        "shipping. We can definitely offer competitive pricing for your "
        "volume.\n\n"
        "Shall we schedule a follow-up call this week?\n\n"
        "Best,\n[Your Name]",

        "Subject: 很高兴在 [Event] 认识您！\n\n"
        "Hi [Name]，\n\n"
        "上周在 [Trade Show] 很高兴认识您，聊 [Topic] 很有收获。\n\n"
        "按约定发来中马货运的价目表，您的量我们可以给优惠价。\n\n"
        "这周方便安排个跟进电话吗？\n\n"
        "此致敬礼，\n[Your Name]"
    ],
    [
        "WhatsApp",
        "客户转介绍",
        "Hi [Name],\n\n"
        "Your friend [Referrer] suggested I reach out to you. "
        "They mentioned you might need help with shipping from China.\n\n"
        "We've been handling [Referrer]'s shipments for [X] months "
        "and they've been very happy with our service.\n\n"
        "Would you like to chat? Happy to offer you the same rates! 😊",

        "Hi [Name]，\n\n"
        "您的朋友 [Referrer] 建议我联系您，说您可能需要从中国运输的帮助。\n\n"
        "我们帮 [Referrer] 走货已经 [X] 个月了，合作一直很愉快。\n\n"
        "方便聊聊吗？可以给您同样的优惠价 😊"
    ],
]

for i, row_data in enumerate(outreach):
    row = 5 + i
    for col, val in enumerate(row_data, 1):
        cell = ws4.cell(row=row, column=col, value=val)
        cell.font = normal_font
        cell.border = thin_border
        cell.alignment = left_wrap
    ws4.row_dimensions[row].height = 150

col_widths_4 = [14, 18, 55, 55]
for i, w in enumerate(col_widths_4, 1):
    ws4.column_dimensions[get_column_letter(i)].width = w

# ============================================================
# Sheet 5: Pipeline 管道看板
# ============================================================
ws5 = wb.create_sheet("Pipeline 管道")

ws5.merge_cells('A1:H1')
ws5.cell(row=1, column=1, value="📊 Sales Pipeline 销售管道").font = title_font
ws5.row_dimensions[1].height = 35

ws5.merge_cells('A2:H2')
ws5.cell(row=2, column=1,
    value="每个 Lead 从 → New → Contacted → Interested → Quoted → Won/Lost").font = small_font

# Pipeline stages as visual board
stages = [
    ("🆕 New 新", 4, yellow_fill),
    ("📞 Contacted 已联系", 4, blue_fill),
    ("💡 Interested 有兴趣", 4, green_fill),
    ("💰 Quoted 已报价", 4, orange_fill),
    ("🤝 Negotiating 谈判中", 4, purple_fill),
]

col_idx = 1
for stage_name, _, fill in stages:
    cell = ws5.cell(row=4, column=col_idx, value=stage_name)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = center
    ws5.merge_cells(start_row=4, start_column=col_idx, end_row=4, end_column=col_idx + 1)
    col_idx += 2

# Pipeline detail headers
pipe_headers = [
    "Lead\n客户", "Score\n评分",
    "Lead\n客户", "Score\n评分",
    "Lead\n客户", "Score\n评分",
    "Lead\n客户", "Score\n评分",
    "Lead\n客户", "Score\n评分",
]
row = 5
for col, h in enumerate(pipe_headers, 1):
    cell = ws5.cell(row=row, column=col, value=h)
    cell.font = bold_font
    cell.fill = grey_fill
    cell.border = thin_border
    cell.alignment = center

for row in range(6, 36):
    for col in range(1, 11):
        ws5.cell(row=row, column=col).border = thin_border
        ws5.cell(row=row, column=col).fill = input_fill
        ws5.cell(row=row, column=col).font = normal_font

col_widths_5 = [20, 8, 20, 8, 20, 8, 20, 8, 20, 8]
for i, w in enumerate(col_widths_5, 1):
    ws5.column_dimensions[get_column_letter(i)].width = w

# Stats at bottom
ws5.cell(row=37, column=1, value="📊 汇总 Summary").font = big_bold
stats = [
    (38, "Total Leads 总 Leads:", ""),
    (39, "New 新:", ""),
    (40, "Contacted 已联系:", ""),
    (41, "Interested 有兴趣:", ""),
    (42, "Quoted 已报价:", ""),
    (43, "Won 已成交:", ""),
    (44, "Conversion Rate 成交率:", ""),
]
for row, label, default in stats:
    ws5.cell(row=row, column=1, value=label).font = bold_font
    ws5.cell(row=row, column=1).border = thin_border
    ws5.cell(row=row, column=2, value=default).fill = input_fill
    ws5.cell(row=row, column=2).border = thin_border

# ============================================================
# Sheet 6: Daily Routine 每日工作清单
# ============================================================
ws6 = wb.create_sheet("Daily Routine")

ws6.merge_cells('A1:E1')
ws6.cell(row=1, column=1, value="📋 Sales Hunter 每日工作清单").font = title_font
ws6.row_dimensions[1].height = 35

ws6.merge_cells('A2:E2')
ws6.cell(row=2, column=1,
    value="每天按清单执行 → 保持节奏 → 持续获客").font = small_font

headers6 = ["Time 时间", "Task 任务", "How 怎么做", "Target 目标", "Done 完成"]
for col, h in enumerate(headers6, 1):
    cell = ws6.cell(row=4, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = center

daily_tasks = [
    ["⏰ 9:00", "LinkedIn 搜索\nFind leads on LinkedIn",
     '用 Google Dorks Sheet 里的 LinkedIn 搜索指令\n或者直接在 LinkedIn 搜索 "import export Malaysia"',
     "每天找 5 个新 Lead", ""],
    ["⏰ 9:30", "Google 搜索\nGoogle search for leads",
     '用 Google Dorks Sheet 的搜索指令\n找马来西亚进出口公司',
     "每天找 5 个新 Lead", ""],
    ["⏰ 10:00", "填入 Lead Database\nLog leads into database",
     "把找到的公司信息填入 Lead Database Sheet\n填越多越好：公司名、联系人、邮箱、电话",
     "10 个新 Lead", ""],
    ["⏰ 10:30", "Lead 评分\nScore leads",
     "根据评分标准给每个 Lead 打分\n5 分 = 完美匹配，1 分 = 待验证",
     "所有新 Lead 评分完", ""],
    ["⏰ 11:00", "外展 Outreach\nSend first messages",
     "选评分 ≥ 3 的 Lead\n从 Outreach 话术模板选合适的\nWhatsApp/Email 发 first contact",
     "发 10 条消息", ""],
    ["⏰ 14:00", "跟进 Follow-up\nFollow up on previous contacts",
     "看 Pipeline 里 Contacted 状态的 Lead\n发 follow-up 消息",
     "跟进 5 个 Lead", ""],
    ["⏰ 15:00", "SEO 内容\nCreate SEO content",
     "根据 SEO AEO Strategy Sheet\n写博客文章/FAQ/landing page",
     "每周 1-2 篇内容", ""],
    ["⏰ 16:00", "更新 Pipeline\nUpdate pipeline",
     "更新每个 Lead 的状态\n把 Interested → Quoted 推进",
     "所有 Lead 状态更新", ""],
    ["⏰ 16:30", "复盘 Review\nDaily review",
     "看 Pipeline 汇总数据\n今天新增多少？成交多少？哪里可以改进？",
     "记录当天数据", ""],
]

for i, row_data in enumerate(daily_tasks):
    row = 5 + i
    for col, val in enumerate(row_data, 1):
        cell = ws6.cell(row=row, column=col, value=val)
        cell.font = normal_font
        cell.border = thin_border
        cell.alignment = left_wrap
        if col == 5:
            cell.fill = input_fill
    ws6.row_dimensions[row].height = 60

col_widths_6 = [12, 22, 45, 20, 8]
for i, w in enumerate(col_widths_6, 1):
    ws6.column_dimensions[get_column_letter(i)].width = w

# ============================================================
# Save
# ============================================================
output_path = "sales-support/Sales_Hunter_Agent.xlsx"
wb.save(output_path)
print(f"✅ Sales Hunter Agent saved to {output_path}")
print(f"   Sheets: {[ws.title for ws in wb.worksheets]}")
print(f"   Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
